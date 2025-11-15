"""
Test Sitemap Filtering - Recursive Expansion
=============================================

This script simulates the EXACT filtering logic from searching_pipeline.py:
1. Reads URLs from Excel
2. Fetches robots.txt for each
3. Applies word/year/date filters on robots.txt sitemaps
4. Recursively expands sitemaps (just like _expand_children_recursive)
5. Shows detailed filtering statistics at each level

NO LLM detection - only filtering test!

Usage:
    python test_sitemap_filtering_recursive.py
"""

import re
import os
import sys
import json
import random
from datetime import datetime
from typing import List, Tuple, Optional, Dict, Any
from openpyxl import load_workbook
from urllib.parse import urlparse, urlsplit, urlunsplit
from playwright.sync_api import sync_playwright
import concurrent.futures as cf
import csv
from overview_store import init_db as ov_init_db, upsert_overview as ov_upsert, export_csv as ov_export_csv
from urllib.parse import urljoin
try:
    import requests as _rq  # optional, used for status logging on fallback failures
except Exception:
    _rq = None

# Force unbuffered output for real-time logs (MUST be before any other stdout modifications)
os.environ['PYTHONUNBUFFERED'] = '1'
try:
    sys.stdout.reconfigure(encoding='utf-8', line_buffering=True, write_through=True)
    sys.stderr.reconfigure(encoding='utf-8', line_buffering=True, write_through=True)
except Exception:
    pass  # Python < 3.7 doesn't have reconfigure

# Fix Windows console encoding for emojis (AFTER unbuffering setup)
# PowerShell users should run: chcp 65001; $OutputEncoding = [Console]::OutputEncoding

# Load environment variables from .env
try:
    from dotenv import load_dotenv
    load_dotenv(override=False)  # Load .env file
    print(f"[env] Loaded .env file")
except Exception:
    print(f"[env] No .env file or dotenv not installed")
    pass

# Env flag to disable CSS fallback
CSS_FALLBACK_DISABLED = str(os.getenv("DISABLE_CSS_FALLBACK", "0")).strip().lower() in ("1", "true", "yes", "on")
# Enable browser retry for sitemap heuristics (like robots browser retry)
SITEMAP_BROWSER_RETRY_ENABLED = str(os.getenv("SITEMAP_BROWSER_RETRY", "1")).strip().lower() in ("1", "true", "yes", "on")
# Require title-like field in urlset items to keep leaf sitemaps
SITEMAP_REQUIRE_ANY_TITLE = str(os.getenv("SITEMAP_REQUIRE_ANY_TITLE", "1")).strip().lower() in ("1", "true", "yes", "on")
# Disable sitemap path entirely (force CSS fallback/crawler)
DISABLE_SITEMAP = str(os.getenv("DISABLE_SITEMAP", "0")).strip().lower() in ("1", "true", "yes", "on")

# Simple CSS crawler flags
CRAWLER_ENABLED = str(os.getenv("CRAWLER_ENABLED", "0")).strip().lower() in ("1", "true", "yes", "on")
CRAWLER_PAGE_TIMEOUT_SEC = 10.0
try:
    CSS_CHUNK_SIZE = max(10000, int(os.getenv("CSS_CHUNK_SIZE", "120000") or 120000))
except Exception:
    CSS_CHUNK_SIZE = 120000
try:
    NAV_SCAN_LIMIT = max(0, int(os.getenv("NAV_SCAN_LIMIT", "0") or 0))
except Exception:
    NAV_SCAN_LIMIT = 0

# NAV-only crawler (always on; header/navbar links only)
try:
    NAV_ALLOW_RE = re.compile(r"(?i)/(news|world|business|market|markets|tech|science|sport|sports|health|politics|opinion|culture)(/|$)")
except Exception:
    NAV_ALLOW_RE = None
try:
    NAV_DENY_RE = re.compile(r"(?i)(search|login|signin|subscribe|account|cart|shop|store|advert|ads|promo|sponsor|podcast|video|live|tv|radio|newsletter|lang|edition|about|contact|careers|privacy|terms)")
except Exception:
    NAV_DENY_RE = None

# Import sitemap discovery functions
from sitemap_discovery import (
    fetch_robots_txt,
    fetch_robots_txt_meta,
    parse_sitemaps_from_robots,
    fetch_bytes,
    maybe_decompress,
    parse_xml_bytes,
    child_text_any_ns,
    _parse_w3c_datetime,
)
import sitemap_discovery as sd


def _to_str(val, default: str) -> str:
    try:
        if val is None:
            return default
        if isinstance(val, bool):
            return "true" if val else "false"
        return str(val)
    except Exception:
        return default

# Import filter functions
from sitemap_filters import filter_by_words, filter_by_date, filter_sitemaps_by_year

# Import LLM function
import selector_scraper as ss


def _append_stream(record: dict, stream_path: str = "selection_extraction_report_stream.jsonl") -> None:
    """Append one JSON record to a newline-delimited stream file immediately."""
    try:
        with open(stream_path, "a", encoding="utf-8") as f:
            json.dump(record, f, ensure_ascii=False)
            f.write("\n")
            f.flush()
            try:
                os.fsync(f.fileno())
            except Exception:
                pass
    except Exception:
        # Best-effort stream; avoid crashing the pipeline due to I/O hiccups
        pass

def _clickable_path(path: str) -> str:
    try:
        from pathlib import Path
        return Path(os.path.abspath(path)).as_uri()
    except Exception:
        try:
            return os.path.abspath(path)
        except Exception:
            return str(path)
# Default stats skeleton for error cases to keep summaries robust
def _default_stats(url: Optional[str] = None, err: Optional[str] = None) -> dict:
    return {
        "url": (url or ""),
        "robotsTxt": {
            "found": False,
            "sitemapsTotal": 0,
            "afterWordFilter": 0,
            "afterYearFilter": 0,
            "afterDateFilter": 0,
            "rejected": [],
        },
        "recursiveExpansion": {
            "childrenFound": 0,
            "childrenRejectedWord": [],
            "childrenRejectedYear": [],
            "childrenRejectedDate": [],
            "childrenRejectedTitle": [],
            "leavesFound": [],
        },
        "finalStats": {
            "totalLeaves": 0,
            "afterWordFilter": 0,
            "afterYearFilter": 0,
            "afterDateFilter": 0,
        },
        "llmDetection": {
            "totalLeaves": 0,
            "successful": 0,
            "failed": 0,
            "selectors": [],
        },
        "cssFallback": {
            "triggered": False,
            "success": False,
            "selectors": None,
            "reason": "error",
            "failure": None,
        },
        "error": (err or ""),
    }


# ========================
# CSV upsert helper (per-site)
# ========================
_CSV_PATH = os.path.join(os.path.dirname(__file__) or ".", "pipelines_overview.csv")
_CSV_HEADER = [
    "Domain (sources)",
    "Selector Discovery Attempted",
    "Selector Discovery Not Attempted Reason",
    "Selector Discovery Attempt Error",
    "Selector Discovery Attempt Error Response",
    "Sitemap Processing Status",
    "Sitemap Processing Error Details",
    "leaf Sitemap URLs Discovered",
    "CSS Fallback Status",
    "CSS Fallback error Details",
    "Which Path Used for Final Extraction",
    "Total Time (sec) in scraping",
    "Raw Articles scraped",
    "Zero Raw Articles Reason",
    "Cleaning Status",
    "Cleaned Articles (Final)",
    "Duplicates Removed",
    "Missing Dates Removed",
    "Out of Range/Old Date Removed",
    "Overall pipelines Status",
    "Overall pipelines Error Details",
    "Overall pipelines Explanation",
    "Leaf Sitemap URLs",
]


def _normalize_domain(url: str) -> str:
    try:
        return urlparse(url).netloc.lower()
    except Exception:
        return ""


def _read_csv_map(path: str) -> Dict[str, Dict[str, str]]:
    rows: Dict[str, Dict[str, str]] = {}
    if not os.path.exists(path):
        return rows
    try:
        with open(path, "r", encoding="utf-8", newline="") as f:
            r = csv.DictReader(f)
            for row in r:
                domain = (row.get(_CSV_HEADER[0]) or "").strip()
                if domain:
                    rows[domain] = {k: (row.get(k) or "") for k in _CSV_HEADER}
    except Exception:
        pass
    return rows


def _default_row(domain: str) -> Dict[str, str]:
    return {
        "Domain (sources)": domain,
        "Selector Discovery Attempted": "No",
        "Selector Discovery Not Attempted Reason": "",
        "Selector Discovery Attempt Error": "",
        "Sitemap Processing Status": "Not Attempted",
        "Sitemap Processing Error Details": "",
        "leaf Sitemap URLs Discovered": "0",
        "CSS Fallback Status": "Not Attempted",
        "CSS Fallback error Details": "",
        "Which Path Used for Final Extraction": "Neither",
        "Total Time (sec) in scraping": "0",
        "Raw Articles scraped": "0",
        "Cleaning Status": "Not Attempted",
        "Cleaned Articles (Final)": "0",
        "Duplicates Removed": "0",
        "Missing Dates Removed": "0",
        "Out of Range/Old Date Removed": "0",
        "Overall pipelines Status": "Pending",
        "Leaf Sitemap URLs": "",
    }


def _write_csv_map(path: str, rows: Dict[str, Dict[str, str]]) -> None:
    tmp = path + ".tmp"
    try:
        with open(tmp, "w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=_CSV_HEADER)
            w.writeheader()
            for domain in sorted(rows.keys()):
                row = rows[domain]
                # Ensure all headers exist
                for h in _CSV_HEADER:
                    row.setdefault(h, "")
                w.writerow(row)
        # Atomic-ish replace
        try:
            os.replace(tmp, path)
        except Exception:
            # Fallback: remove then rename
            try:
                if os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass
            try:
                os.rename(tmp, path)
            except Exception:
                pass
    except Exception:
        # Best-effort; ignore CSV write errors so we don't impact pipeline
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass


def _merge_overall_error(prev: str, new_seg: str, max_len: int = 300) -> str:
    """Merge stage-tagged error segment into existing 'Overall pipelines Error Details'."""
    try:
        prev = (prev or "").strip()
        new_seg = (new_seg or "").strip()
        if not new_seg:
            return prev[:max_len]
        parts = [s.strip() for s in prev.split(" | ") if s.strip()] if prev else []
        if new_seg not in parts:
            parts.append(new_seg)
        out = " | ".join(parts)
        return out[:max_len]
    except Exception:
        return (new_seg or prev or "")[:max_len]


def _merge_friendly_explanation(prev: str, new_sentence: str, max_len: int = 300) -> str:
    """Merge a human-friendly sentence into Overall pipelines Explanation."""
    try:
        prev = (prev or "").strip()
        new_sentence = (new_sentence or "").strip()
        if not new_sentence:
            return prev[:max_len]
        parts = [s.strip() for s in prev.split(" | ") if s.strip()] if prev else []
        if new_sentence not in parts:
            parts.append(new_sentence)
        out = " | ".join(parts)
        return out[:max_len]
    except Exception:
        return (new_sentence or prev or "")[:max_len]


# ==============================
# Selector error detail helpers
# ==============================
def _allowlist_headers(h: Optional[Dict[str, str]]) -> Dict[str, str]:
    try:
        if not isinstance(h, dict):
            return {}
        allow = {"server", "cf-ray", "retry-after", "x-cache", "x-akamai-*", "via"}
        out: Dict[str, str] = {}
        for k, v in h.items():
            lk = str(k or "").lower()
            if lk in allow or any(lk.startswith(p.rstrip("*")) for p in ["x-akamai-"]):
                out[lk] = str(v)
        return out
    except Exception:
        return {}


def _http_probe(url: str, timeout: float = 5.0) -> Optional[Dict[str, Any]]:
    if _rq is None:
        return None
    try:
        r = _rq.get(url, timeout=timeout, allow_redirects=True, stream=True)
        try:
            status = int(getattr(r, "status_code", 0) or 0)
        except Exception:
            status = 0
        try:
            raw = r.content[:1024] if getattr(r, "content", None) is not None else b""
        except Exception:
            raw = b""
        try:
            body_snippet = raw.decode("utf-8", errors="replace").replace("\n", " ")
        except Exception:
            body_snippet = ""
        headers = {}
        try:
            headers = dict(getattr(r, "headers", {}) or {})
        except Exception:
            headers = {}
        return {
            "url": url,
            "status_code": status,
            "headers": _allowlist_headers(headers),
            "body_snippet": (body_snippet[:512] if isinstance(body_snippet, str) else "")
        }
    except Exception:
        return None


def _classify_probe(probe: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    try:
        if not isinstance(probe, dict):
            return None
        status = int(probe.get("status_code") or 0)
        headers = probe.get("headers") or {}
        body = (probe.get("body_snippet") or "").lower()
        server = str(headers.get("server", "")).lower()

        # Cloudflare signals
        if "cloudflare" in server or "cf-ray" in headers or "attention required" in body:
            return {
                "category": "access_blocked",
                "subtype": "cloudflare",
                "reason": "Blocked by Cloudflare / challenge page",
                "retryable": True,
            }

        if status == 429:
            return {
                "category": "http_error",
                "subtype": "429",
                "reason": "Too Many Requests (rate limited)",
                "retryable": True,
            }
        if status in (401, 403):
            return {
                "category": "http_error",
                "subtype": str(status),
                "reason": "Access forbidden/unauthorized",
                "retryable": False,
            }
        if status in (503, 521, 522, 523):
            return {
                "category": "http_error",
                "subtype": str(status),
                "reason": "Service unavailable / gateway error",
                "retryable": True,
            }
        # No strong signal
        if status >= 400:
            return {
                "category": "http_error",
                "subtype": str(status),
                "reason": "HTTP error",
                "retryable": status in (408, 429, 500, 502, 503, 504),
            }
        return None
    except Exception:
        return None


def _build_selector_error_detail(base_url: str,
                                 robots_found: bool,
                                 total_sm: int,
                                 leaves: int,
                                 after_date: int,
                                 llm: Dict[str, Any],
                                 cssf: Dict[str, Any],
                                 probe_timeout: float = 5.0) -> Optional[Dict[str, Any]]:
    try:
        # Primary state-derived category (will be refined by probe if possible)
        primary: Dict[str, Any] = {}
        total_leaves = int((llm or {}).get("totalLeaves") or 0)
        succ = int((llm or {}).get("successful") or 0)

        if not robots_found:
            primary = {"category": "robots_missing", "subtype": "robots_txt_absent", "reason": "robots.txt not found", "retryable": False}
        elif int(total_sm) == 0:
            primary = {"category": "sitemaps_missing", "subtype": "no_sitemaps_listed", "reason": "robots.txt has no sitemaps", "retryable": False}
        elif int(leaves) == 0 and int(after_date) == 0:
            primary = {"category": "no_recent_leaves", "subtype": "stale_or_empty", "reason": "no recent sitemap leaves", "retryable": False}
        elif total_leaves > 0 and succ == 0:
            # Extract first selector error if any
            first_err = ""
            try:
                for s in (llm.get("selectors") or []):
                    e = (s or {}).get("error")
                    if e:
                        first_err = str(e)
                        break
            except Exception:
                first_err = ""
            primary = {"category": "selector_detection_failed", "subtype": "all_failed", "reason": ("selector detection failed" + (f": {first_err}" if first_err else "")), "retryable": True}

        # Probe for network/WAF signals
        probe = None
        try:
            probe = _http_probe(base_url, timeout=probe_timeout)
        except Exception:
            probe = None
        classified = _classify_probe(probe)

        # Compose detail
        now_iso = datetime.utcnow().isoformat() + "Z"
        detail: Dict[str, Any] = {
            "timestamp": now_iso,
            "retryable": bool(primary.get("retryable", True) if primary else True),
        }
        if primary:
            detail.update({k: v for k, v in primary.items() if k in ("category", "subtype", "reason", "retryable")})
        if isinstance(classified, dict):
            # If probe indicates access_blocked etc., prefer it; else keep primary category and add probe as evidence
            if classified.get("category"):
                detail["category"] = classified["category"]
            if classified.get("subtype"):
                detail["subtype"] = classified["subtype"]
            if classified.get("reason"):
                detail["reason"] = classified["reason"]
            if "retryable" in classified:
                detail["retryable"] = bool(classified["retryable"]) or bool(detail.get("retryable"))

        # Evidence
        evidence: Dict[str, Any] = {"url": base_url}
        if isinstance(probe, dict):
            if probe.get("status_code") is not None:
                detail["status_code"] = int(probe.get("status_code") or 0)
            evidence["headers"] = probe.get("headers") or {}
            if probe.get("body_snippet"):
                evidence["body_snippet"] = probe.get("body_snippet")

        # Attach css fallback failure context (if any)
        if bool((cssf or {}).get("triggered")) and not bool((cssf or {}).get("success")):
            failure = (cssf or {}).get("failure") if isinstance((cssf or {}).get("failure"), dict) else None
            if failure and (failure.get("ok") is False):
                evidence["css_failure"] = {
                    "stage": failure.get("stage"),
                    "errorType": failure.get("errorType"),
                    "errorMessage": failure.get("errorMessage"),
                    "context": failure.get("context") or {},
                }

        if evidence:
            detail["evidence"] = evidence

        # If nothing set, return None
        if not detail.get("category") and not detail.get("status_code"):
            return None
        return detail
    except Exception:
        return None


def _summarize_error_response(detail: Optional[Dict[str, Any]], llm: Dict[str, Any], cssf: Dict[str, Any], max_len: int = 220) -> str:
    try:
        status_code = None
        message = None
        if isinstance(detail, dict):
            try:
                if detail.get("status_code") is not None:
                    status_code = int(detail.get("status_code") or 0)
            except Exception:
                status_code = detail.get("status_code")
            # Prefer CSS failure message if available
            failure = (cssf or {}).get("failure") if isinstance((cssf or {}).get("failure"), dict) else None
            if failure and (failure.get("ok") is False):
                message = failure.get("errorMessage") or failure.get("errorType") or failure.get("stage")
            if not message:
                message = detail.get("reason")
        # If still no message, fall back to first LLM selector error
        if not message:
            try:
                for s in (llm or {}).get("selectors", []) or []:
                    e = (s or {}).get("error")
                    if e:
                        message = str(e)
                        break
            except Exception:
                pass
        sc_str = str(status_code) if (status_code is not None and str(status_code)) else "unknown"
        msg_str = (message or "").strip()
        out = f"status={sc_str}; message={msg_str}" if msg_str else f"status={sc_str}"
        return out[:max_len]
    except Exception:
        return ""

def _upsert_csv_row(domain: str, updates: Dict[str, str]) -> None:
    # Small retry for occasional Windows file locks
    for _ in range(3):
        try:
            rows = _read_csv_map(_CSV_PATH)
            row = rows.get(domain) or _default_row(domain)
            row[_CSV_HEADER[0]] = domain
            for k, v in (updates or {}).items():
                if k in _CSV_HEADER and v is not None:
                    if k == "Overall pipelines Error Details":
                        row[k] = _merge_overall_error(row.get(k) or "", str(v))
                    elif k == "Overall pipelines Explanation":
                        row[k] = _merge_friendly_explanation(row.get(k) or "", str(v))
                    else:
                        row[k] = str(v)
            rows[domain] = row
            _write_csv_map(_CSV_PATH, rows)
            return
        except PermissionError:
            import time as _t
            _t.sleep(0.2)
        except Exception:
            return

def _extract_sample_from_sitemap(sitemap_url: str, timeout: float = 15.0, sample_count: int = 3) -> List[str]:
    """Extract sample <url> entries from sitemap for LLM analysis.

    Collects up to `sample_count` items from the top and up to `sample_count`
    items from the bottom of the sitemap (deduplicated, in that order).
    """
    try:
        raw = fetch_bytes(sitemap_url, timeout)
        raw = maybe_decompress(sitemap_url, raw)
        if not raw:
            return []

        root = parse_xml_bytes(raw)
        if root is None:
            return []

        url_elements = root.findall(".//{*}url")
        if not url_elements:
            return []

        # Take first N and last N, then deduplicate while preserving order
        first_part = url_elements[:sample_count]
        last_part = url_elements[-sample_count:] if len(url_elements) > sample_count else []

        selected_unique = []
        seen_ids = set()
        for el in (first_part + last_part):
            if el is None:
                continue
            key = id(el)
            if key in seen_ids:
                continue
            seen_ids.add(key)
            selected_unique.append(el)

        samples: List[str] = []
        for url_elem in selected_unique:
            # Convert XML element to string
            import xml.etree.ElementTree as ET
            sample_str = ET.tostring(url_elem, encoding='unicode', method='xml')
            samples.append(sample_str)

        return samples
    except Exception:
        return []


def _build_llm_prompt(samples: List[str], sitemap_url: str) -> str:
    """Build LLM prompt for sitemap field detection"""
    prompt = f"""You are analyzing a sitemap XML to discover field mappings.

Sitemap URL: {sitemap_url}

Here are {len(samples)} sample <url> entries from the sitemap:

"""
    for idx, sample in enumerate(samples, 1):
        prompt += f"\n--- Sample {idx} ---\n{sample}\n"
    
    prompt += """
Your task: Analyze these samples and identify the XML paths for common fields.

Return ONLY a valid JSON object with this structure:
{
  "type": "urlset" or "index",
  "item": "url" or "sitemap",
  "fields": {
    "url": "path to URL field",
    "date": "path to date field",
    "title": "path to title field (if present)",
    "description": "path to description (if present)",
    ... other fields ...
  },
  "confidence": 0.0 to 1.0
}

Field path format:
- For direct children: "tagname" (e.g., "loc", "lastmod")
- For nested: "parent/child" (e.g., "news:news/news:title")
- Use exact tag names including namespaces

Important:
- Only include fields that exist in the samples
- Be precise with nested paths
- Set confidence based on consistency across samples
"""
    return prompt


def _parse_llm_response(llm_response: str) -> Optional[dict]:
    """Parse LLM JSON response"""
    import json
    try:
        # Try to extract JSON from response
        start = llm_response.find('{')
        end = llm_response.rfind('}') + 1
        if start >= 0 and end > start:
            json_str = llm_response[start:end]
            data = json.loads(json_str)
            
            # Validate structure
            if 'fields' in data and isinstance(data['fields'], dict):
                return {
                    "type": data.get("type", "urlset"),
                    "item": data.get("item", "url"),
                    "fields": data['fields'],
                    "confidence": data.get("confidence", 0.7),
                    "detectionMethod": "llm"
                }
    except Exception:
        pass
    
    return None


def _detect_selectors_with_llm(sitemap_url: str, timeout: float = 15.0) -> Tuple[Optional[dict], Optional[str]]:
    """Detect sitemap field selectors using LLM. Returns (detected, error_message)."""
    try:
        # Extract samples
        samples = _extract_sample_from_sitemap(sitemap_url, timeout, sample_count=3)
        
        if not samples:
            return None, None
        
        # Build prompt
        prompt = _build_llm_prompt(samples, sitemap_url)
        
        # Call LLM
        llm_response = ss._call_llm(prompt, model=None)  # Use default model
        
        # Parse response
        detected = _parse_llm_response(llm_response)
        
        return detected, None
        
    except Exception as e:
        print(f"      LLM Error: {e}")
        return None, f"{type(e).__name__}: {str(e)}"


def _detect_selectors_basic(sitemap_url: str, timeout: float = 15.0) -> Optional[dict]:
    """Basic detection - inspect XML tags without LLM"""
    try:
        raw = fetch_bytes(sitemap_url, timeout)
        raw = maybe_decompress(sitemap_url, raw)
        if not raw:
            return None
        
        root = parse_xml_bytes(raw)
        if root is None:
            return None
        
        tag = (root.tag or "").lower()
        
        if tag.endswith("sitemapindex"):
            return {
                "type": "index",
                "item": "sitemap",
                "fields": {
                    "url": "loc",
                    "date": "lastmod"
                },
                "detectionMethod": "basic"
            }
        elif tag.endswith("urlset"):
            # Check what tags are present
            first_url = root.find(".//{*}url")
            if first_url is None:
                return None
            
            fields = {}
            
            # Check for common tags
            if first_url.find(".//{*}loc") is not None:
                fields["url"] = "loc"
            if first_url.find(".//{*}lastmod") is not None:
                fields["date"] = "lastmod"
            if first_url.find(".//{*}changefreq") is not None:
                fields["changefreq"] = "changefreq"
            if first_url.find(".//{*}priority") is not None:
                fields["priority"] = "priority"
            # News sitemap title (e.g., <news:news><news:title>)
            try:
                if first_url.find(".//{*}news/{*}title") is not None:
                    fields["title"] = "news:news/news:title"
            except Exception:
                pass
            
            return {
                "type": "urlset",
                "item": "url",
                "fields": fields,
                "detectionMethod": "basic"
            }
    except Exception:
        return None


def _chunk_html(html: str, chunk_size: int = 120_000) -> List[str]:
    """
    Split HTML into chunks of ~120KB
    
    Exactly like selector_scraper.py
    """
    chunks = []
    for i in range(0, len(html), chunk_size):
        chunks.append(html[i:i+chunk_size])
    return chunks


def _llm_prompt_for_html(batch_index: int, chunk: str) -> str:
    """
    Build LLM prompt for HTML chunk
    
    Exactly like selector_scraper.py line 129-144
    """
    return (
        "Goal: Find ONLY editorial/news article list sections on this page and RETURN CSS SELECTORS ONLY for fields.\n\n"
        "CRITICAL: Ignore and DO NOT return sections that are video, photo/gallery, live, carousel/slider, widgets, promos, sponsored, or recommendation carousels.\n\n"
        "Return JSON ONLY. Prefer multiple sections if present (but only editorial/news lists).\n\n"
        "Output format (either):\n"
        "- Single: { sectionName?, title, link, date?, description?, author?, category?, ticker? confidence }\n"
        "- Multi: { sections: [ { sectionName, selectors: { title, link, date?, description?, author?, category?, ticker? }, confidence } ] }\n\n"
        "STRICT RULES:\n"
        "- Return CSS SELECTORS (strings) for elements, NOT example values or text.\n"
        "- Avoid unstable hashed classes (e.g., .emotion-*) and avoid :nth-child.\n"
        "- At least one of title/link must use parent>child (>) for specificity.\n"
        "- Target visible DOM (no meta tags).\n"
        "- Aim for selectors that match many items (>=5).\n"
        "- Focus ONLY on editorial/news items; ignore ads/recommendations/widgets/sponsored/native/video/gallery/live/carousels.\n\n"
        f"HTML CHUNK ({batch_index + 1}):\n{chunk}"
    )


def _parse_candidates(llm_response: str) -> List[Dict[str, Any]]:
    """
    Parse LLM response to extract section candidates
    
    Exactly like selector_scraper.py line 183-215
    
    Handles both formats:
    - Single: { sectionName?, title, link, ... }
    - Multi: { sections: [ {...}, {...} ] }
    """
    import selector_scraper as ss
    data = ss._extract_json_from_text(llm_response)
    
    if not isinstance(data, dict):
        return []
    
    # Check if multi-section format
    if "sections" in data and isinstance(data["sections"], list):
        # Multi-section
        candidates = []
        for section in data["sections"]:
            if isinstance(section, dict):
                candidates.append({
                    "sectionName": section.get("sectionName"),
                    "selectors": section.get("selectors", {}),
                    "confidence": section.get("confidence", 0.7)
                })
        return candidates
    
    # Single section format
    if "title" in data or "link" in data or "selectors" in data:
        selectors = data.get("selectors") or {
            "title": data.get("title"),
            "link": data.get("link"),
            "date": data.get("date"),
            "description": data.get("description"),
            "author": data.get("author"),
            "category": data.get("category")
        }
        return [{
            "sectionName": data.get("sectionName"),
            "selectors": selectors,
            "confidence": data.get("confidence", 0.7)
        }]
    
    return []


def _acceptance_checks(title_sel: str, link_sel: str) -> Tuple[bool, bool]:
    """
    Validate selectors (strict/lenient)
    
    Exactly like selector_scraper.py line 250-270
    
    Returns: (is_strict, is_lenient)
    """
    # Must have both
    if not title_sel or not link_sel:
        return (False, False)
    
    # Check for parent>child in at least one
    has_child_combinator = (">" in title_sel) or (">" in link_sel)
    
    # Check for overly generic selectors
    too_generic = (
        title_sel.strip() in ["a", "div", "span", "h1", "h2", "h3"] or
        link_sel.strip() in ["a", "div"]
    )
    
    if too_generic:
        return (False, False)
    
    # Strict: Must have child combinator
    is_strict = has_child_combinator
    
    # Lenient: At least specific enough
    is_lenient = len(title_sel) > 3 and len(link_sel) > 3
    
    return (is_strict, is_lenient)


def _signature(selectors: dict) -> str:
    """
    Create unique signature for deduplication
    
    Exactly like selector_scraper._signature()
    """
    title = selectors.get("title", "")
    link = selectors.get("link", "")
    return f"{title}|{link}"


def _diagnose_repetition_hint(html: str, selectors: Dict[str, Any]) -> int:
    """Approx item repetition hint by counting occurrences of a key class/tag from selectors in HTML.
    Heuristic only; no DOM/CSS engine needed.
    """
    try:
        import re as _re
        sel = selectors or {}
        cand = str(sel.get("title") or sel.get("link") or "")
        # Prefer a class token from selector
        classes = _re.findall(r"\\.([a-zA-Z0-9_-]+)", cand)
        if classes:
            count = 0
            for c in set(classes[-3:]):  # last few tokens
                try:
                    count = max(count, len(_re.findall(r'class=["\'][^"\']*\b' + _re.escape(c) + r'\b', html)))
                except Exception:
                    pass
            if count:
                return int(count)
        # Fallback: tag-based hint
        low = cand.lower()
        for t in ("article", "li", "h3", "h2"):
            if t in low:
                try:
                    return int(len(_re.findall(r"<" + t + r"[\s>]", html, flags=_re.I)))
                except Exception:
                    return 0
        return 0
    except Exception:
        return 0


def _css_selector_fallback(base_url: str, timeout: float = 60.0, headful: bool = False, slowmo_ms: int = 0, relaxed: bool = False) -> Optional[dict]:
    """
    CSS Selector Fallback using EXACT selector_scraper.discover_selectors() approach
    
    Steps:
    1. Launch Playwright browser
    2. Navigate to page and wait for render
    3. Extract full HTML
    4. Save HTML to debug_html/
    5. Chunk HTML (~120KB chunks)
    6. For each chunk:
       - Build LLM prompt
       - Call LLM
       - Parse candidates (multi-section)
       - Validate (strict/lenient)
       - Save raw response to debug_llm/
    7. Deduplicate sections by signature
    8. Return all sections
    
    Returns:
    {
      "detectionMethod": "css_fallback_browser",
      "pageUrl": "...",
      "htmlLength": 123456,
      "chunksProcessed": 5,
      "sections": [...],
      "totalSections": 2,
      "debugFiles": {...}
    }
    """
    print(f"\n   🔄 CSS FALLBACK: Using browser-based detection (exact selector_scraper approach)...")
    
    # Parse URL for domain
    parsed = urlparse(base_url)
    domain = parsed.netloc or "unknown"
    ts = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S-%fZ")
    
    # Ensure debug directories
    os.makedirs("debug_html", exist_ok=True)
    os.makedirs("debug_llm", exist_ok=True)
    
    # Step 1: Launch Browser & Fetch HTML
    print(f"\n   🌐 [1/7] Launching browser...")
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=not headful, slow_mo=slowmo_ms or None)
            
            # Random viewport (like selector_scraper)
            viewport = {
                "width": random.randint(1200, 1440),
                "height": random.randint(800, 1000)
            }
            
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                viewport=viewport,
                extra_http_headers={"Referer": f"https://{domain}/"}
            )
            
            page = context.new_page()
            
            print(f"      Navigating to {base_url}...")
            try:
                page.goto(base_url, wait_until="domcontentloaded", timeout=int(timeout*1000))
            except Exception as e:
                # Structured failure: browser navigation
                print(f"      ❌ Browser error: {type(e).__name__}: {str(e)[:80]}")
                return {
                    "ok": False,
                    "stage": "browser_fetch",
                    "errorType": type(e).__name__,
                    "errorMessage": str(e)[:120],
                    "context": {"pageUrl": base_url}
                }
            
            # Wait for page to be ready (like _readiness_loop in selector_scraper)
            page.wait_for_timeout(2000)  # 2 second wait
            
            # Get HTML
            html = page.content()
            browser.close()
            
            print(f"      ✅ HTML fetched ({len(html)} chars)")
    
    except Exception as e:
        print(f"      ❌ Browser error: {type(e).__name__}: {str(e)[:80]}")
        return {
            "ok": False,
            "stage": "browser_fetch",
            "errorType": type(e).__name__,
            "errorMessage": str(e)[:120],
            "context": {"pageUrl": base_url}
        }
    
    # Step 2: Save HTML to debug
    print(f"\n   💾 [2/7] Saving HTML snapshot...")
    html_path = os.path.join("debug_html", f"{domain}_{ts}_HTML.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"      ✅ Saved: {_clickable_path(html_path)}")
    
    # Step 3: Chunk HTML
    print(f"\n   ✂️  [3/7] Chunking HTML...")
    chunks = _chunk_html(html, chunk_size=int(CSS_CHUNK_SIZE))
    print(f"      ✅ Created {len(chunks)} chunk(s) (~120KB each)")
    if len(html or "") == 0:
        return {
            "ok": False,
            "stage": "html_snapshot",
            "errorType": None,
            "errorMessage": "empty html",
            "context": {"htmlLength": 0, "pageUrl": base_url}
        }
    
    # Step 4: Process each chunk with LLM
    print(f"\n   🤖 [4/7] Processing chunks with LLM...")
    all_candidates = []
    debug_files = []
    
    for i, chunk in enumerate(chunks):
        print(f"\n      Batch {i+1}/{len(chunks)}: {len(chunk)} chars")
        
        # Build prompt
        prompt = _llm_prompt_for_html(i, chunk)
        
        # Call LLM
        try:
            llm_response = ss._call_llm(prompt, model=None)
            
            # Save raw response
            raw_path = os.path.join("debug_llm", f"{domain}_{ts}_batch{i+1}_raw.json")
            with open(raw_path, "w", encoding="utf-8") as f:
                f.write(llm_response)
            debug_files.append(raw_path)
            print(f"         💾 Saved: {_clickable_path(raw_path)}")
            
            # Parse candidates
            candidates = _parse_candidates(llm_response)
            print(f"         📋 Found {len(candidates)} candidate(s)")
            
            # Validate each candidate
            accepted_strict = []
            accepted_lenient = []
            
            for c in candidates:
                sel = c.get("selectors") or {}
                title_sel = str(sel.get("title", "")).strip()
                link_sel = str(sel.get("link", "")).strip()
                
                if not title_sel or not link_sel:
                    continue
                
                is_strict, is_lenient = _acceptance_checks(title_sel, link_sel)
                
                item = {
                    "sectionName": c.get("sectionName") or f"Batch {i+1}",
                    "selectors": sel,
                    "confidence": c.get("confidence", 0.7),
                    "detectedInChunk": i + 1
                }
                
                if is_strict:
                    # Diagnose repetition hint (likely editorial if higher)
                    rep = _diagnose_repetition_hint(html, sel)
                    if rep:
                        item["diagnosis"] = {"repeatHint": int(rep)}
                        print(f"         • diagnose: repeats≈{rep} (higher => likely editorial list)")
                    accepted_strict.append(item)
                elif is_lenient or relaxed:
                    rep = _diagnose_repetition_hint(html, sel)
                    if rep:
                        item["diagnosis"] = {"repeatHint": int(rep)}
                        print(f"         • diagnose: repeats≈{rep} (higher => likely editorial list)")
                    accepted_lenient.append(item)
            
            print(f"         ✅ Strict: {len(accepted_strict)}, Lenient: {len(accepted_lenient)}")
            all_candidates.extend(accepted_strict + accepted_lenient)
            
        except Exception as e:
            print(f"         ❌ LLM error: {type(e).__name__}: {str(e)[:50]}")
            return {
                "ok": False,
                "stage": "llm",
                "errorType": type(e).__name__,
                "errorMessage": str(e)[:120],
                "context": {"batchIndex": i + 1, "chunks": len(chunks), "pageUrl": base_url}
            }
    
    # Step 5: Deduplicate by signature
    print(f"\n   🔀 [5/7] Deduplicating sections...")
    seen = {}
    for c in all_candidates:
        sig = _signature(c["selectors"])
        if sig not in seen:
            seen[sig] = c
    
    unique_sections = list(seen.values())
    print(f"      ✅ Unique sections: {len(unique_sections)} (deduplicated from {len(all_candidates)})")
    
    if not unique_sections:
        print(f"\n   ❌ No valid sections found")
        return {
            "ok": False,
            "stage": "validate_selectors",
            "errorType": None,
            "errorMessage": "no strict or lenient selectors accepted",
            "context": {
                "htmlLength": len(html or ""),
                "chunksProcessed": len(chunks),
                "candidatesTotal": len(all_candidates),
                "strictAccepted": 0,
                "lenientAccepted": 0,
                "pageUrl": base_url
            }
        }
    
    # Step 6: Display results
    print(f"\n   📋 [6/7] Detected sections:")
    for idx, section in enumerate(unique_sections, 1):
        name = section.get("sectionName", f"Section {idx}")
        conf = section.get("confidence", 0)
        sel_count = len(section.get("selectors", {}))
        print(f"      {idx}. {name} (confidence={conf}, {sel_count} selectors)")
    
    # Step 7: Return result
    print(f"\n   ✅ [7/7] CSS Fallback Complete!")
    
    return {
        "detectionMethod": "css_fallback_browser",
        "pageUrl": base_url,
        "htmlLength": len(html),
        "chunksProcessed": len(chunks),
        "sections": unique_sections,
        "totalSections": len(unique_sections),
        "debugFiles": {
            "html": html_path,
            "llmResponses": debug_files
        }
    }


# ========================
# CSS Crawler (homepage + links)
# ========================
def _extract_internal_links_from_homepage(base_url: str, timeout: float, max_links: int) -> List[str]:
    out: List[str] = []
    try:
        raw = fetch_bytes(base_url, timeout)
        html = raw.decode("utf-8", errors="ignore") if raw else ""
        if not html:
            return out
        import re as _re
        hrefs = _re.findall(r'href=["\']([^"\']+)["\']', html, flags=_re.I)
        base = urlsplit(base_url)
        base_host = (base.netloc or "").lower()
        base_host_bare = base_host[4:] if base_host.startswith("www.") else base_host
        for h in hrefs:
            try:
                absu = urljoin(f"{base.scheme}://{base.netloc}/", h.strip())
                p = urlsplit(absu)
                if p.scheme not in ("http", "https"):
                    continue
                host = (p.netloc or "").lower()
                host_bare = host[4:] if host.startswith("www.") else host
                if host_bare != base_host_bare:
                    continue
                # drop fragment and query
                norm = f"{p.scheme}://{p.netloc}{p.path or '/'}"
                if norm and norm not in out:
                    out.append(norm)
                if len(out) >= max_links:
                    break
            except Exception:
                continue
    except Exception:
        pass
    return out


def _extract_nav_links_from_homepage(base_url: str, timeout: float) -> List[str]:
    """Extract only header/navbar links from homepage HTML (no JS interaction).

    Heuristics:
    - Prefer <nav>...</nav> blocks
    - Fallback to <header>...</header>
    - Also scan small windows around elements with nav-like classes/ids
    - Same-domain filter and allow/deny regexes applied
    """
    out: List[str] = []
    try:
        raw = fetch_bytes(base_url, timeout)
        html = raw.decode("utf-8", errors="ignore") if raw else ""
        if not html:
            return out
        import re as _re
        base = urlsplit(base_url)
        base_host = (base.netloc or "").lower()
        base_host_bare = base_host[4:] if base_host.startswith("www.") else base_host

        segments: List[str] = []
        # 1) All <nav> blocks
        try:
            segments.extend(_re.findall(r"<nav[\s\S]*?</nav>", html, flags=_re.I))
        except Exception:
            pass
        # 2) <header> block (top-most only)
        try:
            header_blocks = _re.findall(r"<header[\s\S]*?</header>", html, flags=_re.I)
            if header_blocks:
                # pick the smallest (likely top banner) or the first
                segments.append(min(header_blocks, key=len))
        except Exception:
            pass
        # 3) Containers with nav-like classes/ids; grab a window around them
        try:
            for m in _re.finditer(r"<[^>]+(?:class|id)=\"[^\"]*(?:navbar|primary-nav|site-nav|global-nav|top-nav|main-nav|menu)\"[^>]*>", html, flags=_re.I):
                start = max(0, m.start() - 1500)
                end = min(len(html), m.end() + 6000)
                segments.append(html[start:end])
        except Exception:
            pass

        # Try LLM-assisted extraction first
        try:
            seg_text = "\n\n".join(segments) if segments else html[:150000]
            if seg_text:
                prompt = (
                    "You are given an HTML snippet from the top of a news website.\n"
                    "Return ONLY PRIMARY NAVIGATION links from the site header/top navbar that lead to STOCKS/MARKETS pages.\n"
                    "Include market news, quotes, indices, tickers, equities. Ignore generic sections and any non-market pages.\n"
                    "Explicitly EXCLUDE: tourism, travel, photos, gallery, lifestyle, culture, entertainment, video, live, tv, radio, weather, games, crossword, horoscope(s), cooking, food, recipe(s), shop, store, classifieds, jobs, careers, about, contact.\n"
                    "Also ignore utility/secondary links: search, login, subscribe, account, cart, advert, ads, promo, sponsored, newsletter, language/edition, privacy, terms.\n\n"
                    "Output JSON ONLY in this format:\n"
                    "{\n  \"links\": [ { \"label\": \"...\", \"href\": \"https://...\", \"reason\": \"markets|stocks news|quotes|indices|tickers\" } ]\n}\n\n"
                    f"Base URL: {base_url}\nHTML:\n{seg_text}"
                )
                raw = ss._call_llm(prompt, model=None)
                # Save raw
                try:
                    parsed = urlparse(base_url)
                    domain = parsed.netloc or "unknown"
                    ts = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S-%fZ")
                    raw_path = os.path.join("debug_llm", f"{domain}_{ts}_nav_raw.json")
                    with open(raw_path, "w", encoding="utf-8") as f:
                        f.write(raw)
                    print(f"      [crawler] LLM nav JSON saved: {_clickable_path(raw_path)}")
                except Exception:
                    pass
                data = None
                try:
                    data = ss._extract_json_from_text(raw)
                except Exception:
                    data = None
                if isinstance(data, dict) and isinstance(data.get("links"), list):
                    for it in data["links"]:
                        try:
                            href = str((it or {}).get("href") or "").strip()
                            if not href:
                                continue
                            absu = urljoin(f"{base.scheme}://{base.netloc}/", href)
                            p = urlsplit(absu)
                            if p.scheme not in ("http", "https"):
                                continue
                            host = (p.netloc or "").lower()
                            host_bare = host[4:] if host.startswith("www.") else host
                            if host_bare != base_host_bare:
                                continue
                            norm = f"{p.scheme}://{p.netloc}{p.path or '/'}"
                            low = norm.lower()
                            label_low = str(((it or {}).get("label") or "")).lower()
                            reason_low = str(((it or {}).get("reason") or "")).lower()
                            if NAV_DENY_RE is not None and NAV_DENY_RE.search(low):
                                continue
                            # Allow shallow paths or allow-keywords
                            allowed = True
                            if NAV_ALLOW_RE is not None and not NAV_ALLOW_RE.search(low):
                                depth = len([s for s in (p.path or "/").split("/") if s])
                                if depth > 2:
                                    allowed = False
                            # Extra domain-specific deny list (non-stock categories)
                            deny_terms = (
                                "tourism", "travel", "photo", "photos", "gallery", "lifestyle", "culture",
                                "entertainment", "video", "live", "tv", "radio", "weather", "game", "games",
                                "crossword", "puzzle", "horoscope", "horoscopes", "cooking", "food", "recipe", "recipes",
                                "shop", "store", "classified", "classifieds", "job", "jobs", "career", "careers",
                                "about", "contact"
                            )
                            if any(t in label_low or t in low or t in reason_low for t in deny_terms):
                                continue
                            # Stocks/markets allow terms
                            allow_terms = (
                                "stock", "stocks", "market", "markets", "equities", "quote", "quotes",
                                "ticker", "symbol", "index", "indices", "invest", "investing"
                            )
                            if not any(t in label_low or t in low or t in reason_low for t in allow_terms):
                                continue
                            if allowed and norm not in out:
                                out.append(norm)
                        except Exception:
                            continue
                    if out:
                        # If LLM produced at least one, return (no further parsing)
                        return out
        except Exception:
            pass

        # Extract anchors from collected segments (regex fallback)
        seen: set = set()
        for seg in segments:
            try:
                hrefs = _re.findall(r'href=["\']([^"\']+)["\']', seg, flags=_re.I)
                for h in hrefs:
                    absu = urljoin(f"{base.scheme}://{base.netloc}/", h.strip())
                    p = urlsplit(absu)
                    if p.scheme not in ("http", "https"):
                        continue
                    host = (p.netloc or "").lower()
                    host_bare = host[4:] if host.startswith("www.") else host
                    # Always same-domain only (simplified)
                    if host_bare != base_host_bare:
                        continue
                    # Normalize: drop fragment
                    norm = f"{p.scheme}://{p.netloc}{p.path or '/'}"
                    # Apply deny/allow regexes
                    low = norm.lower()
                    if NAV_DENY_RE is not None and NAV_DENY_RE.search(low):
                        continue
                    if NAV_ALLOW_RE is not None and not NAV_ALLOW_RE.search(low):
                        # If no allow match, still keep top-level categories (depth ≤ 2)
                        depth = len([s for s in (p.path or "/").split("/") if s])
                        if depth > 2:
                            continue
                        if norm not in seen:
                            seen.add(norm)
                            out.append(norm)
            except Exception:
                continue
    except Exception:
        pass
    return out


def _css_crawler_collect(base_url: str, page_timeout: float) -> dict:
    def _safe_confidence(val: Any) -> float:
        try:
            if val is None:
                return 0.0
            if isinstance(val, (int, float)):
                return float(val)
            s = str(val).strip().lower()
            if not s:
                return 0.0
            # Map common categorical strings
            if s in ("very high", "vhigh", "v-high"):
                return 0.98
            if s in ("high",):
                return 0.9
            if s in ("medium", "med"):
                return 0.6
            if s in ("low",):
                return 0.3
            if s.endswith('%'):
                # e.g., "85%"
                num = float(s[:-1])
                return max(0.0, min(1.0, num / 100.0))
            # numeric string
            v = float(s)
            # clamp if someone returned 0-100
            if v > 1.0:
                return max(0.0, min(1.0, v / 100.0))
            return max(0.0, min(1.0, v))
        except Exception:
            return 0.0
    try:
        # Build target list: homepage + all header/navbar links
        print("\n   🧭 [crawler] Extracting header/navbar links...")
        links = _extract_nav_links_from_homepage(base_url, timeout=page_timeout)
        total_before = len(links)
        if NAV_SCAN_LIMIT > 0:
            links = links[:NAV_SCAN_LIMIT]
            print(f"      [crawler] nav_total_links={total_before}; applying NAV_SCAN_LIMIT={NAV_SCAN_LIMIT} -> using {len(links)} link(s)")
        else:
            print(f"      [crawler] nav_total_links={total_before}; NAV_SCAN_LIMIT=0 (no cap)")
        targets: List[str] = [base_url] + [u for u in links if u != base_url]
        print(f"\n   🔎 [crawler] Pages to scan ({len(targets)}):")
        for idx, u in enumerate(targets, 1):
            print(f"      {idx:2d}. {u}")
        sections_all: List[Dict[str, Any]] = []
        per_page: Dict[str, List[Dict[str, Any]]] = {}
        pages_visited = 0
        for u in targets:
            pages_visited += 1
            res = _css_selector_fallback(base_url=u, timeout=page_timeout, headful=False, slowmo_ms=0, relaxed=True)
            if isinstance(res, dict) and int(res.get("totalSections", 0)) > 0:
                # Tag each section with its source URL and collect per page
                page_sections: List[Dict[str, Any]] = []
                for s in res.get("sections", []) or []:
                    if isinstance(s, dict):
                        sc = dict(s)
                        sc.setdefault("sourceUrl", u)
                        page_sections.append(sc)
                        sections_all.append(sc)
                if page_sections:
                    # Deduplicate per-page by signature, keep best confidence
                    seen_page: Dict[str, Dict[str, Any]] = {}
                    for sec in page_sections:
                        sigp = _signature(sec.get("selectors") or {})
                        if not sigp:
                            continue
                        cur = seen_page.get(sigp)
                        if cur is None or _safe_confidence(sec.get("confidence")) > _safe_confidence(cur.get("confidence")):
                            seen_page[sigp] = sec
                    per_page[u] = list(seen_page.values())
        if not sections_all:
            return {
                "ok": False,
                "stage": "css_crawler",
                "errorType": None,
                "errorMessage": "no sections from homepage and links",
                "context": {"pagesVisited": pages_visited, "targets": len(targets)}
            }
        # Dedupe and pick best by signature (support_count, then confidence)
        by_sig: Dict[str, Dict[str, Any]] = {}
        for sec in sections_all:
            sel = sec.get("selectors") or {}
            sig = _signature(sel)
            if not sig:
                continue
            entry = by_sig.get(sig)
            if entry is None:
                by_sig[sig] = {"count": 1, "best": sec, "best_conf": _safe_confidence(sec.get("confidence"))}
            else:
                entry["count"] += 1
                conf = _safe_confidence(sec.get("confidence"))
                if conf > float(entry.get("best_conf", 0)):
                    entry["best"] = sec
                    entry["best_conf"] = conf
        # Rank signatures, return ALL unique signatures (best per signature)
        ranked = sorted(by_sig.values(), key=lambda e: (int(e.get("count", 0)), _safe_confidence(e.get("best_conf", 0))), reverse=True)
        chosen_sections = [e.get("best") for e in ranked if isinstance(e.get("best"), dict)]
        return {
            "detectionMethod": "css_crawler",
            "pageUrl": base_url,
            "htmlLength": 0,
            "chunksProcessed": 0,
            "sections": chosen_sections,
            "totalSections": len(chosen_sections),
            # Per-page buckets so consumers can attribute sections correctly
            "perPage": per_page,
            "debugFiles": {"pagesVisited": pages_visited, "targetsTried": len(targets)}
        }
    except Exception as e:
        return {
            "ok": False,
            "stage": "css_crawler",
            "errorType": type(e).__name__,
            "errorMessage": str(e)[:120],
            "context": {"pageUrl": base_url}
        }


def _css_crawler_or_single(base_url: str, timeout: float) -> Optional[dict]:
    if CRAWLER_ENABLED:
        return _css_crawler_collect(base_url=base_url, page_timeout=min(timeout, CRAWLER_PAGE_TIMEOUT_SEC))
    return _css_selector_fallback(base_url=base_url, timeout=timeout, headful=False, slowmo_ms=0)


# ========================
# Sitemap Fallback Helpers
# ========================

_FALLBACK_LOGS: int = 0  # cap noisy logs when parsing fails for many candidates
# Per-host budget to avoid excessive browser retries for the same domain
_BROWSER_RETRY_BUDGET: dict = {}

# Diagnostics flags for heuristic fallback (module-scoped, reset before each fallback call)
_DIAG_FB_BROWSER_ATTEMPTED: bool = False
_DIAG_FB_BROWSER_SUCCESS: bool = False
_DIAG_FB_BLOCK_DETECTED: bool = False

DEFAULT_SITEMAP_PATHS_EXTRA: List[str] = [
    '1', '1-1', '1_index_sitemap', '1_de_0_sitemap', '1_en_0_sitemap', '01', '001', '0001',
    '2020', '2021', '2022', '2023', '2024', '2025', 'index', 'index-files', 's_1', 's_01', 's_001',
    's_0001', 's-1', 's-01', 's-001', 's-0001', 's_1-1', 's1', 's01', 's001', 's0001', 's1-1',
    'site_1', 'site_01', 'site_001', 'site_0001', 'site-1', 'site-01', 'site-001', 'site-0001',
    'site-1-1', 'site', 'site1', 'site01', 'site001', 'site0001', 'sites', 'siteindex', 'siteindex1',
    'siteindex01', 'siteindex001', 'siteindex0001', 'site-map', 'site_map', 'sitemap', 'sitemapmain',
    'sitemapMain', 'sitemap.all', 'sitemap.index', 'sitemap-index.sitemap', 'sitemap.website',
    'sitemap-shop', 'sitemap.pages', 'sitemap.default', 'sitemap.main', 'sitemap.ssl', 'sitemap.root',
    'sitemap.de', 'sitemap.en', 'sitemap.1', 'sitemap.01', 'sitemap.001', 'sitemap.0001', 'sitemap_0',
    'sitemap_1', 'sitemap_01', 'sitemap_001', 'sitemap_0001', 'sitemap_1_1', 'sitemap_01_01',
    'sitemap_content', 'sitemap_default', 'sitemap_en', 'sitemap_de', 'sitemap_index_de',
    'sitemap_index_en', 'sitemap_index', 'sitemap_sites', 'sitemap_ssl', 'sitemap-1', 'sitemap-1-1',
    'sitemap-01', 'sitemap-001', 'sitemap-0001', 'sitemap_hreflang', 'sitemapindex', 'sitemap-index',
    'index-sitemap', 'sitemap-index-1', 'sitemap-index-de', 'sitemap-index-en', 'sitemap-complete',
    'sitemap-default', 'sitemap-root', 'sitemap-root-1', 'sitemap-main', 'sitemap-pages',
    'sitemap-posts', 'sitemap-sections', 'sitemap-sites', 'sitemap-ssl', 'sitemap-de', 'sitemap-de-de',
    'sitemap-de_de', 'sitemap-deu', 'sitemap-en', 'sitemap-en-us', 'sitemap-en_us', 'sitemap-eng',
    'sitemap-web', 'sitemap-website', 'sitemap-www', 'sitemap-secure', 'sitemap-secure-www',
    'secure-sitemap', 'sitemaps', 'sitemaps-1-sitemap', 'sitemapsindex', 'Sitemap', 'SiteMap',
    'sitemap1', 'sitemap2', 'sitemap01', 'sitemap001', 'sitemap0001', 'sitemap-files', 'sitemap-items',
    'sitemap-items-1', 'sitemappages', 'sitemapproducts', 'sitemap-4seo', 'default', 'standard_sitemap',
    'items', 'files', 'sm', 'google-sitemap', 'google-sitemap-index', 'google-sitemap-1', 'google_sitemap',
    'googlesitemap', 'google.sitemap', 'gsitemap', 'GSiteMap', 'gsiteindex', 'xml-sitemap', 'xml_sitemap',
    'main-sitemap', 'name-ihrer-sitemap', 'news', 'news-sitemap', 'news_sitemap', 'newssitemap',
    'googleNewsList', 'gNewsSiteMap', 'googlenews', 'google-news-sitemap', 'google-news-index.sitemap',
    'googlenews-sitemap', 'sitemap_gnews', 'sitemap-google-news', 'sitemap-googlenews', 'sitemap-news',
    'sitemap_news', 'sitemapnews', 'sitemapNews', 'sitemap-archives', 'sitemap_articles', 'sitemap-cms',
    'sitemap-global', 'sitemap_static', 'sitemap_global', 'sitemap-home', 'sitemap_https', 'sitemap_xml',
    'sitemap_xml_de', 'sitemap_xml_en', 'sitemaps-all-pages', 'sitemap_neu', 'sitemap_new',
    'sitemap-master-index', 'List', 'list', 'website', 'web', 'wp-sitemap', 'main', 'map', 'maps',
    'global', 'geositemap', 'content', 'content_index', 'page', 'pages', 'page-sitemap', 'post-sitemap',
    'product-sitemap', 'product_index', 'root-sitemap', 'all', 'all-sitemaps-xml', 'add-sitemap', 'de',
    'de.sitemap', 'en', 'en.sitemap', 'wpms-sitemap', 'seo_sitemap', 'toprank-sitemap_index_01-aa', 'urllist',
    'xmlsitemap', 'map/1', 'map/index', 'map/global', 'map/default', 'map/s_1', 'map/s-1', 'map/s1',
    'map/site_1', 'map/site-1', 'map/site', 'map/siteindex', 'map/site1', 'map/sitemap_1',
    'map/sitemap_index_de', 'map/sitemap_index', 'map/sitemap_sites', 'map/sitemap-1', 'map/sitemap-index',
    'map/sitemap-sites', 'map/sitemap', 'map/sitemap1', 'map/sitemap01', 'map/sitemap001', 'map/sitemap0001',
    'map/sm', 'map/main', 'sitemap/1', 'sitemap/01', 'sitemap/001', 'sitemap/0001', 'sitemap/de-sitemap',
    'sitemap/en-sitemap', 'sitemap/index', 'sitemap/index-files', 'sitemap/global', 'sitemap/news-sitemap',
    'sitemap/s_1', 'sitemap/s-1', 'sitemap/s1', 'sitemap/site_1', 'sitemap/site-1', 'sitemap/site',
    'sitemap/siteindex', 'sitemap/site1', 'sitemap/sitemap', 'sitemap/sitemap_1', 'sitemap/sitemap_01',
    'sitemap/sitemap_001', 'sitemap/sitemap_0001', 'sitemap/sitemap_index_de', 'sitemap/sitemap_index',
    'sitemap/sitemap_sites', 'sitemap/sitemap-0', 'sitemap/sitemap-1', 'sitemap/sitemap-01',
    'sitemap/sitemap-001', 'sitemap/sitemap-0001', 'sitemap/sitemap-index', 'sitemap/sitemap-sections',
    'sitemap/sitemap-sites', 'sitemap/sitemap-main', 'sitemap/sitemap-de', 'sitemap/sitemap_de',
    'sitemap/sitemap-en', 'sitemap/sitemap_en', 'sitemap/sitemap_news', 'sitemap/sitemap1',
    'sitemap/sitemap01', 'sitemap/sitemap001', 'sitemap/sitemap0001', 'sitemap/map', 'sitemap/map1',
    'sitemap/map01', 'sitemap/map001', 'sitemap/map0001', 'sitemap/main', 'sitemap/sitemap_global',
    'sitemap/sitemapmain', 'sitemap/default', 'sitemap/full', 'sitemap/items', 'sitemap/root', 'sitemap/sm',
    'sitemap/web', 'sitemap/pages', 'sitemap/page-a', 'sitemap/files', 'sitemap/file-a', 'sitemap/de/add_sitemap',
    'sitemap/de/sitemap', 'sitemap/en/sitemap', 'sitemap/google/index', 'sitemaps/1', 'sitemaps/01',
    'sitemaps/001', 'sitemaps/0001', 'sitemaps/index', 'sitemaps/pages', 'sitemaps/default', 'sitemaps/main',
    'sitemaps/news', 'sitemaps/sitemap.news', 'sitemaps/s_1', 'sitemaps/s-1', 'sitemaps/s1', 'sitemaps/site_1',
    'sitemaps/site-1', 'sitemaps/site', 'sitemaps/siteindex', 'sitemaps/site1', 'sitemaps/sitemap_1',
    'sitemaps/sitemap_01', 'sitemaps/sitemap_001', 'sitemaps/sitemap_0001', 'sitemaps/sitemap_1_1',
    'sitemaps/sitemap_index_de', 'sitemaps/sitemap_de', 'sitemaps/sitemap_en', 'sitemaps/sitemap_index',
    'sitemaps/sitemap_sites', 'sitemaps/sitemap-1', 'sitemaps/sitemap-01', 'sitemaps/sitemap-001',
    'sitemaps/sitemap-0001', 'sitemaps/sitemap-index', 'sitemaps/sitemap-sites', 'sitemaps/sitemap-main',
    'sitemaps/sitemap', 'sitemaps/sitemaps', 'sitemaps/sitemap1', 'sitemaps/sitemap01', 'sitemaps/sitemap001',
    'sitemaps/sitemap0001', 'sitemaps/sitemappages', 'sitemaps/sm', 'sitemaps/de/sitemap', 'sitemaps/en/sitemap',
    'sitemaps2/index', 'sitemaps-2/index', 'sitemap_xml/sitemap', 'sitemapxml/sitemap', 'sitemapxmllist/1',
    'sitemapxmllist/index', 'sitemapxmllist/s_1', 'sitemapxmllist/s-1', 'sitemapxmllist/s1',
    'sitemapxmllist/site_1', 'sitemapxmllist/site-1', 'sitemapxmllist/site', 'sitemapxmllist/siteindex',
    'sitemapxmllist/site1', 'sitemapxmllist/sitemap_1', 'sitemapxmllist/sitemap_index_de',
    'sitemapxmllist/sitemap_index', 'sitemapxmllist/sitemap_sites', 'sitemapxmllist/sitemap-1',
    'sitemapxmllist/sitemap-index', 'sitemapxmllist/sitemap-sites', 'sitemapxmllist/Sitemap',
    'sitemapxmllist/sitemap', 'sitemapxmllist/sitemap1', 'sitemapxmllist/sitemap01', 'sitemapxmllist/sitemap001',
    'sitemapxmllist/sitemap0001', 'sitemapxmllist/sm', 'sitemapxmllist-var/index', 's/sitemap.xml', 'sm/1',
    'sm/index', 'sm/s_1', 'sm/s-1', 'sm/s1', 'sm/site_1', 'sm/site-1', 'sm/site', 'sm/siteindex', 'sm/site1',
    'sm/sitemap_1', 'sm/sitemap_index_de', 'sm/sitemap_index', 'sm/sitemap_sites', 'sm/sitemap-1',
    'sm/sitemap-index', 'sm/sitemap-sites', 'sm/Sitemap', 'sm/sitemap', 'sm/sitemap1', 'sm/sitemap01',
    'sm/sitemap001', 'sm/sitemap0001', 'sm/sm', 'xml/1', 'xml/index', 'xml/s_1', 'xml/s-1', 'xml/s1',
    'xml/site_1', 'xml/site-1', 'xml/site', 'xml/siteindex', 'xml/site1', 'xml/sitemap_1',
    'xml/sitemap_index_de', 'xml/sitemap_index', 'xml/sitemap_sites', 'xml/sitemap-1', 'xml/sitemap-index',
    'xml/sitemap-sites', 'xml/sitemap-pages', 'xml/sitemappages', 'xml/Sitemap', 'xml/sitemap', 'xml/sitemap1',
    'xml/sitemap01', 'xml/sitemap001', 'xml/sitemap0001', 'xml/sm', 'xml/main', 'xml/sitemapmain.xml',
    'xml/SitemapMain.xml', 'xml-sitemap/xml-sitemap', 'export/sitemap', 'export/sitemap_0', 'export/sitemap_index',
    'export/google_sitemap_de', 'export/google_sitemap_en', 'export/sitemapindex_de', 'files/sitemap',
    'files/sitemap-index', 'files/sitemap_index', 'files/sitemap/sitemap', 'files/sitemap/sitemap-index',
    'files/sitemap/sitemap_index', 'files/sitemap/index', 'files/sitemaps/sitemap-de', 'files/sitemaps/sitemap-en',
    'files/xml/sitemap-index', 'files/xml/sitemap', 'files/xml/sitemap.pages', 'files/others/sitemap',
    'sites/default/files/sitemap', 'sites/default/files/sitemap/1', 'sites/default/files/sitemap/sitemap',
    'sites/default/files/sitemap/sitemap_1', 'sites/default/files/sitemaps',
    'sites/default/files/sitemaps/sitemap', 'sites/default/files/sitemaps/sitemapindex',
    'sites/default/files/sitemaps/sitemap-index', 'sites/default/files/sitemaps/sitemap_index',
    'sites/default/files/sitemaps/sitemapmonthly-1', 'de/main', 'de/sitemap-content', 'de/sitemap-de',
    'de/sitemap_index', 'de/sitemap', 'de/sitemaps-1-sitemap', 'de/sitemaps/index', 'de/wp-sitemap',
    'de/googlesitemap', 'de-de/main', 'de-de/sitemap', 'en/main', 'en/sitemap-content', 'en/sitemap-en',
    'en/sitemap_index', 'en/sitemap', 'en/sitemaps-1-sitemap', 'en/sitemaps/index', 'en/wp-sitemap',
    'en/googlesitemap', 'en-us/main', 'en-us/sitemap', 'share/sitemap', 'share/sitemap-de', 'share/sitemap_de',
    'share/sitemap-en', 'share/sitemap_en', 'share/sitemap-xml', 'public/sitemap', 'public/sitemap-main',
    'public/sitemap-de', 'public/sitemap-en', 'public/sitemap-1', 'public/sitemap-01', 'public/sitemap-001',
    'public/sitemap-0001', 'public/sitemap_index', 'public/sitemap/index', 'public/sitemap/de/siteindex',
    'public/sitemap/en/siteindex', 'public/sitemap-xml/sitemap', 'pub/sitemap', 'pub/sitemap-1', 'pub/sitemap-01',
    'pub/sitemap-001', 'pub/sitemap-0001', 'pub/sitemap-1-1', 'pub/sitemap_de', 'pub/sitemap_en',
    'pub/sitemap/sitemap', 'pub/sitemaps/sitemap', 'pub/media/sitemap', 'pub/media/sitemap-1-1',
    'pub/media/sitemap/sitemap', 'updated/sitemap_index', 'items/sitemap', 'cms/sitemap', 'cms/sitemap_index',
    'myinterfaces/cms/googlesitemap-overview', 'blog/post-sitemap', 'blog/sitemap', 'blog/sitemap_index',
    'blog/page-sitemap', 'wp/sitemap', 'wordpress/sitemap', 'fileadmin/sitemap/sitemap',
    'typo3temp/dd_googlesitemap/sitemap', 'typo3temp/sitemap_seiten', 'temp/sitemap', 'temp/sitemap-https',
    'userdata/sitemap', 'incms_files/sitemap', 'navigation/ws/xmlsitemap/sitemap', 'nav-sitemap/sitemap_index',
    'site/sitemap', 'static/sitemap', 'system/sitemap', 'media/sitemap', 'media/sitemap_de', 'media/sitemap_en',
    'media/sitemap/sitemap', 'docs/sitemap', 'seo/sitemap', 'full', 'rss', 'rss2', 'atom', 'feed',
    'feed/google_sitemap.xml', 'feeds/sitemap', 'datafeed/sitemap-search-de'
]

def _build_base_variants(root_url: str) -> List[str]:
    try:
        parts = urlsplit(root_url)
        host = (parts.netloc or "").lower()
        if host.startswith("www."):
            bare = host[4:]
        else:
            bare = host
        variants = [
            f"https://{bare}",
            f"https://www.{bare}",
            f"http://{bare}",
        ]
        return list(dict.fromkeys(variants))
    except Exception:
        return [root_url]


def _fetch_and_validate_sitemap(candidate_url: str, timeout: float) -> Optional[dict]:
    global _FALLBACK_LOGS, _DIAG_FB_BLOCK_DETECTED, _DIAG_FB_BROWSER_ATTEMPTED, _DIAG_FB_BROWSER_SUCCESS
    try:
        raw = fetch_bytes(candidate_url, timeout)
        raw = maybe_decompress(candidate_url, raw)
        if not raw:
            # Log status if available
            if _FALLBACK_LOGS < 5:
                status = "unknown"
                if _rq is not None:
                    try:
                        r = _rq.get(candidate_url, timeout=timeout, allow_redirects=True, stream=True)
                        status = str(getattr(r, "status_code", "unknown"))
                    except Exception:
                        status = "error"
                print(f"[fallback] Empty response for {candidate_url} (status={status})")
                _FALLBACK_LOGS += 1
            # Consider browser retry for access-blocked scenarios
            should_try_browser = False
            probe = None
            if SITEMAP_BROWSER_RETRY_ENABLED and not candidate_url.lower().endswith('.gz'):
                try:
                    probe = _http_probe(candidate_url, timeout=min(5.0, float(timeout) if timeout else 5.0))
                except Exception:
                    probe = None
                try:
                    status_int = int((probe or {}).get("status_code") or 0)
                except Exception:
                    status_int = 0
                body_low = str((probe or {}).get("body_snippet") or "").lower()
                # Triggers: 3xx, 401/403, 429, common WAF text
                if (300 <= status_int < 400) or status_int in (401, 403, 429) or ("cloudflare" in body_low or "attention required" in body_low):
                    should_try_browser = True
                    _DIAG_FB_BLOCK_DETECTED = True
            if should_try_browser:
                try:
                    host = (urlsplit(candidate_url).netloc or "").lower()
                except Exception:
                    host = ""
                budget_used = int(_BROWSER_RETRY_BUDGET.get(host) or 0)
                if budget_used < 1:
                    # Single-shot browser retry
                    print(f"[fallback] Browser retry for sitemap (host budget {budget_used+1}/1): {candidate_url}")
                    try:
                        from playwright.sync_api import sync_playwright  # lazy import safety
                        proxy_server = os.getenv("PROXY_SERVER")
                        proxy_kw = {"server": proxy_server} if (proxy_server and proxy_server.strip()) else None
                        with sync_playwright() as p:
                            browser = p.chromium.launch(headless=True, proxy=proxy_kw) if proxy_kw else p.chromium.launch(headless=True)
                            context = browser.new_context(
                                user_agent=(
                                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36"
                                ),
                                viewport={"width": 1366, "height": 864},
                            )
                            page = context.new_page()
                            try:
                                page.goto(candidate_url, wait_until="domcontentloaded", timeout=int(max(5.0, float(timeout)) * 1000))
                                html = page.content() or ""
                            finally:
                                try:
                                    browser.close()
                                except Exception:
                                    pass
                    except Exception as _be:
                        html = ""
                        if _FALLBACK_LOGS < 5:
                            print(f"[fallback] Browser retry failed for {candidate_url}: {type(_be).__name__}")
                            _FALLBACK_LOGS += 1
                    finally:
                        _DIAG_FB_BROWSER_ATTEMPTED = True
                    _BROWSER_RETRY_BUDGET[host] = budget_used + 1
                    if html:
                        try:
                            raw = html.encode("utf-8", errors="ignore")
                            # continue to parse below
                        except Exception:
                            raw = None
                    else:
                        raw = None
                else:
                    if _FALLBACK_LOGS < 5:
                        print(f"[fallback] Skipping browser retry (budget exhausted) for host: {host}")
                        _FALLBACK_LOGS += 1
            if not raw:
                return None
        root = parse_xml_bytes(raw)
        if root is None:
            # Log first bytes and status to help identify WAF/HTML blocks
            if _FALLBACK_LOGS < 5:
                status = "unknown"
                if _rq is not None:
                    try:
                        r = _rq.get(candidate_url, timeout=timeout, allow_redirects=True, stream=True)
                        status = str(getattr(r, "status_code", "unknown"))
                    except Exception:
                        status = "error"
                try:
                    snippet = (raw[:200] if raw else b"")
                    preview = snippet.decode("utf-8", errors="replace").replace("\n", " ")
                except Exception:
                    preview = ""
                print(f"[fallback] Non-XML for {candidate_url} (status={status}); first bytes: {preview[:200]}")
                _FALLBACK_LOGS += 1
            # Attempt browser retry if blocked and enabled
            root = None
            should_try_browser = False
            probe = None
            if SITEMAP_BROWSER_RETRY_ENABLED and not candidate_url.lower().endswith('.gz'):
                try:
                    probe = _http_probe(candidate_url, timeout=min(5.0, float(timeout) if timeout else 5.0))
                except Exception:
                    probe = None
                try:
                    status_int = int((probe or {}).get("status_code") or 0)
                except Exception:
                    status_int = 0
                body_low = str((probe or {}).get("body_snippet") or "").lower()
                if (300 <= status_int < 400) or status_int in (401, 403, 429) or ("cloudflare" in body_low or "attention required" in body_low):
                    should_try_browser = True
                    _DIAG_FB_BLOCK_DETECTED = True
            if should_try_browser:
                try:
                    host = (urlsplit(candidate_url).netloc or "").lower()
                except Exception:
                    host = ""
                budget_used = int(_BROWSER_RETRY_BUDGET.get(host) or 0)
                if budget_used < 1:
                    print(f"[fallback] Browser retry for non-XML sitemap body (host budget {budget_used+1}/1): {candidate_url}")
                    try:
                        from playwright.sync_api import sync_playwright  # lazy import safety
                        proxy_server = os.getenv("PROXY_SERVER")
                        proxy_kw = {"server": proxy_server} if (proxy_server and proxy_server.strip()) else None
                        with sync_playwright() as p:
                            browser = p.chromium.launch(headless=True, proxy=proxy_kw) if proxy_kw else p.chromium.launch(headless=True)
                            context = browser.new_context(
                                user_agent=(
                                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36"
                                ),
                                viewport={"width": 1366, "height": 864},
                            )
                            page = context.new_page()
                            try:
                                page.goto(candidate_url, wait_until="domcontentloaded", timeout=int(max(5.0, float(timeout)) * 1000))
                                html = page.content() or ""
                            finally:
                                try:
                                    browser.close()
                                except Exception:
                                    pass
                    except Exception as _be:
                        html = ""
                        if _FALLBACK_LOGS < 5:
                            print(f"[fallback] Browser retry failed for {candidate_url}: {type(_be).__name__}")
                            _FALLBACK_LOGS += 1
                    finally:
                        _DIAG_FB_BROWSER_ATTEMPTED = True
                    _BROWSER_RETRY_BUDGET[host] = budget_used + 1
                    if html:
                        try:
                            raw2 = html.encode("utf-8", errors="ignore")
                            root = parse_xml_bytes(raw2)
                            if root is not None:
                                _DIAG_FB_BROWSER_SUCCESS = True
                        except Exception:
                            root = None
                else:
                    if _FALLBACK_LOGS < 5:
                        print(f"[fallback] Skipping browser retry (budget exhausted) for host: {host}")
                        _FALLBACK_LOGS += 1
            if root is None:
                return None
        tag = root.tag or ""
        if tag.endswith("sitemapindex"):
            locs = [child_text_any_ns(e, "loc") for e in root.findall(".//{*}sitemap")]
            locs = [u for u in locs if u]
            return {"kind": "sitemapindex", "urls": locs}
        if tag.endswith("urlset"):
            return {"kind": "urlset", "urls": [candidate_url]}
        # Unknown root -> log preview once in a while
        if _FALLBACK_LOGS < 5:
            status = "unknown"
            if _rq is not None:
                try:
                    r = _rq.get(candidate_url, timeout=timeout, allow_redirects=True, stream=True)
                    status = str(getattr(r, "status_code", "unknown"))
                except Exception:
                    status = "error"
            try:
                snippet = (raw[:200] if raw else b"")
                preview = snippet.decode("utf-8", errors="replace").replace("\n", " ")
            except Exception:
                preview = ""
            print(f"[fallback] Unexpected XML root '{tag}' for {candidate_url} (status={status}); first bytes: {preview[:200]}")
            _FALLBACK_LOGS += 1
        return None
    except Exception:
        return None


def _probe_default_paths(root_url: str, timeout: float = 10.0) -> Tuple[List[str], str]:
    bases = _build_base_variants(root_url)
    names = [
        "sitemap.xml",
        "sitemap_index.xml",
        "sitemap1.xml",
        "sitemap-1.xml",
        "sitemaps/sitemap.xml",
        "sitemap/sitemap.xml",
        "news-sitemap.xml",
        "sitemap-news.xml",
        "video-sitemap.xml",
        "image-sitemap.xml",
        "post-sitemap.xml",
        "page-sitemap.xml",
        "category-sitemap.xml",
    ] + DEFAULT_SITEMAP_PATHS_EXTRA
    candidates: List[str] = []
    for b in bases:
        for n in names:
            base_path = urljoin(b + "/", n)
            candidates.append(base_path)
            # If no explicit extension, also try with .xml
            low = n.lower()
            if not (low.endswith('.xml') or low.endswith('.xml.gz')):
                candidates.append(urljoin(b + "/", n + ".xml"))
            # Always try compressed variant for the xml form
            if low.endswith('.xml'):
                candidates.append(urljoin(b + "/", n + ".gz"))
            elif not low.endswith('.xml.gz'):
                candidates.append(urljoin(b + "/", n + ".xml.gz"))

    tried = 0
    for cu in candidates:
        tried += 1
        res = _fetch_and_validate_sitemap(cu, timeout)
        if res and res.get("urls"):
            return (res["urls"], "fallback: default_paths")
        if tried >= 150:
            break
    return ([], "")


def _probe_cms_paths(root_url: str, timeout: float = 10.0) -> Tuple[List[str], str]:
    bases = _build_base_variants(root_url)
    names = [
        "wp-sitemap.xml",
        "sitemap_index.xml",
        "sitemap.xml",
    ]
    for b in bases:
        for n in names:
            cu = urljoin(b + "/", n)
            res = _fetch_and_validate_sitemap(cu, timeout)
            if res and res.get("urls"):
                return (res["urls"], "fallback: cms_paths")
    return ([], "")


def _extract_sitemap_links_from_homepage(root_url: str, timeout: float = 10.0) -> List[str]:
    out: List[str] = []
    try:
        raw = fetch_bytes(root_url, timeout)
        if not raw:
            # Optional browser retry if access-blocked
            html = ""
            should_try_browser = False
            probe = None
            if SITEMAP_BROWSER_RETRY_ENABLED:
                try:
                    probe = _http_probe(root_url, timeout=min(5.0, float(timeout) if timeout else 5.0))
                except Exception:
                    probe = None
                try:
                    status_int = int((probe or {}).get("status_code") or 0)
                except Exception:
                    status_int = 0
                body_low = str((probe or {}).get("body_snippet") or "").lower()
                if (300 <= status_int < 400) or status_int in (401, 403, 429) or ("cloudflare" in body_low or "attention required" in body_low):
                    should_try_browser = True
            if should_try_browser:
                try:
                    from playwright.sync_api import sync_playwright  # lazy import
                    proxy_server = os.getenv("PROXY_SERVER")
                    proxy_kw = {"server": proxy_server} if (proxy_server and proxy_server.strip()) else None
                    with sync_playwright() as p:
                        browser = p.chromium.launch(headless=True, proxy=proxy_kw) if proxy_kw else p.chromium.launch(headless=True)
                        context = browser.new_context(
                            user_agent=(
                                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36"
                            ),
                            viewport={"width": 1366, "height": 864},
                        )
                        page = context.new_page()
                        try:
                            page.goto(root_url, wait_until="domcontentloaded", timeout=int(max(5.0, float(timeout)) * 1000))
                            html = page.content() or ""
                        finally:
                            try:
                                browser.close()
                            except Exception:
                                pass
                except Exception:
                    html = ""
            if not html:
                return out
        else:
            html = raw.decode("utf-8", errors="ignore")
        # Very lightweight extraction: href="...xml" and text contains 'sitemap'
        import re as _re
        hrefs = _re.findall(r'href=["\']([^"\']+)["\']', html, flags=_re.I)
        texts = _re.findall(r">([^<]{1,120})<", html)
        cand: List[str] = []
        for h in hrefs:
            s = h.strip()
            low = s.lower()
            if (low.endswith(".xml") or low.endswith(".xml.gz")) and ("sitemap" in low):
                cand.append(s)
        # Basic rel=sitemap
        rels = _re.findall(r'rel=["\']sitemap["\']\s+href=["\']([^"\']+)["\']', html, flags=_re.I)
        cand.extend(rels)
        # Absolutize and dedupe
        base = urlsplit(root_url)
        base_url = f"{base.scheme}://{base.netloc}"
        for c in cand:
            try:
                absu = urljoin(base_url + "/", c)
                out.append(absu)
            except Exception:
                continue
    except Exception:
        pass
    # Stable unique
    return list(dict.fromkeys(out))[:25]


def _probe_homepage_links(root_url: str, timeout: float = 10.0) -> Tuple[List[str], str]:
    cands = _extract_sitemap_links_from_homepage(root_url, timeout)
    for cu in cands:
        res = _fetch_and_validate_sitemap(cu, timeout)
        if res and res.get("urls"):
            return (res["urls"], "fallback: homepage_links")
    return ([], "")


def _fallback_discover_sitemaps(root_url: str, timeout: float = 10.0) -> Tuple[List[str], str]:
    urls, why = _probe_default_paths(root_url, timeout)
    if urls:
        return (urls, why)
    urls, why = _probe_cms_paths(root_url, timeout)
    if urls:
        return (urls, why)
    urls, why = _probe_homepage_links(root_url, timeout)
    if urls:
        return (urls, why)
    return ([], "")


def _apex_roots(orig_url: str) -> List[str]:
    try:
        parts = urlsplit(orig_url)
        host = (parts.netloc or "").lower()
        if not host:
            return []
        # Drop port if present
        if ':' in host:
            host = host.split(':', 1)[0]
        labels = [l for l in host.split('.') if l]
        # No escalation for hosts with < 3 labels (already apex or www apex)
        if len(labels) < 3:
            return []

        # Offline allowlist for multi-label public suffixes
        MULTI_SUFFIXES = {
            "com.au", "net.au", "org.au",
            "co.uk", "org.uk", "ac.uk", "gov.uk", "sch.uk",
            "com.br", "com.mx", "co.in", "co.jp", "co.nz",
            "com.sg", "com.tr", "com.hk",
        }

        # Determine suffix length (labels) by allowlist match, else default 1
        suffix = None
        for s in MULTI_SUFFIXES:
            if host == s or host.endswith('.' + s):
                suffix = s
                break
        suffix_labels = (suffix.count('.') + 1) if suffix else 1

        # Need at least one label before the suffix to form a registrable domain
        if len(labels) <= suffix_labels:
            return []

        registrable = '.'.join(labels[-(suffix_labels + 1):])

        # Do not escalate if already at registrable or www.registrable
        if host == registrable or host == ("www." + registrable):
            return []

        roots = [
            f"https://{registrable}",
            f"https://www.{registrable}",
            f"http://{registrable}",
        ]
        # Stable unique
        return list(dict.fromkeys(roots))
    except Exception:
        return []


def read_urls_from_excel(file_path: str = "urls.xlsx") -> List[str]:
    """Read URLs from Excel file with broad pattern support (text + hyperlinks + domains)."""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Helper: extract URLs from arbitrary text using regexes
        def _extract_urls_from_text(text: str) -> List[str]:
            if not text:
                return []
            s = str(text)
            results: List[str] = []
            try:
                # 1) http/https (case-insensitive)
                strict = re.findall(r"(?i)https?://[^\s\'\"()<>]+", s)
                results.extend(strict)
                # 2) www.*
                wwws = re.findall(r"(?i)\bwww\.[^\s\'\"()<>]+", s)
                results.extend(wwws)
                # 3) bare domains like example.com or sub.example.co.uk
                domains = re.findall(r"(?i)\b[a-z0-9][a-z0-9-]{0,62}(?:\.[a-z0-9-]{1,63})+\b", s)
                results.extend(domains)
            except Exception:
                pass
            return results

        # Helper: normalize a URL-like string into canonical URL
        def _normalize(raw: str) -> Optional[str]:
            if not raw:
                return None
            t = raw.strip().strip("),.;:]")
            if not t:
                return None
            # Prepend scheme for www/bare domains
            low = t.lower()
            if low.startswith("http://") or low.startswith("https://"):
                norm = t
            elif low.startswith("www."):
                norm = "https://" + t
            else:
                # If looks like bare domain, add scheme
                if re.match(r"(?i)^[a-z0-9][a-z0-9-]{0,62}(?:\.[a-z0-9-]{1,63})+", t):
                    norm = "https://" + t
                else:
                    return None
            try:
                parts = urlsplit(norm)
                if not parts.netloc:
                    return None
                scheme = (parts.scheme or "https").lower()
                host = (parts.netloc or "").lower()
                path = parts.path or "/"
                query = ("?" + parts.query) if parts.query else ""
                frag = ("#" + parts.fragment) if parts.fragment else ""
                return f"{scheme}://{host}{path}{query}{frag}"
            except Exception:
                return None

        # Try to find a URL-like column by header within first row; fallback to scan all columns
        url_col_idx = None
        try:
            header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
            header_tokens = {"url", "site", "website", "homepage", "domain"}
            for idx, cell in enumerate(header_cells):
                val = str(cell.value).strip().lower() if cell and cell.value is not None else ""
                if any(tok in val for tok in header_tokens):
                    url_col_idx = idx
                    break
        except Exception:
            url_col_idx = None

        urls: List[str] = []
        seen = set()

        def _maybe_add(raw: str) -> None:
            n = _normalize(raw)
            if n and n not in seen:
                seen.add(n)
                urls.append(n)

        # Iterate rows as cells (not values_only) to access hyperlinks
        # IMPORTANT: At most ONE URL per row
        for row in ws.iter_rows(min_row=2, values_only=False):
            cells = [row[url_col_idx]] if (url_col_idx is not None and url_col_idx < len(row)) else row

            chosen: Optional[str] = None

            # Preference 1: hyperlink target (any cell in the row)
            try:
                for cell in cells:
                    try:
                        if getattr(cell, "hyperlink", None) and getattr(cell.hyperlink, "target", None):
                            cand = _normalize(str(cell.hyperlink.target))
                            if cand:
                                chosen = cand
                                break
                    except Exception:
                        continue
            except Exception:
                pass

            # Preference 2: first URL-like text in the row (strict -> www -> bare domain order)
            if chosen is None:
                try:
                    for cell in cells:
                        if cell is None or cell.value is None:
                            continue
                        matches = _extract_urls_from_text(str(cell.value))
                        for m in matches:
                            cand = _normalize(m)
                            if cand:
                                chosen = cand
                                break
                        if chosen is not None:
                            break
                except Exception:
                    pass

            if chosen is not None and chosen not in seen:
                seen.add(chosen)
                urls.append(chosen)

        wb.close()
        return urls
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")
        return []


def test_recursive_expansion(url: str, recent_hours: int = 24, timeout: float = 15.0, max_depth: int = 3, llm_concurrency: int = 1) -> dict:
    """
    Test recursive sitemap expansion with filtering (like searching_pipeline.py)
    
    Returns dict with statistics about filtering at each level
    """
    stats = {
        "url": url,
        "robotsTxt": {
            "found": False,
            "sitemapsTotal": 0,
            "afterWordFilter": 0,
            "afterYearFilter": 0,
            "afterDateFilter": 0,
            "rejected": []
        },
        "recursiveExpansion": {
            "childrenFound": 0,
            "childrenRejectedWord": [],
            "childrenRejectedYear": [],
            "childrenRejectedDate": [],
            "childrenRejectedTitle": [],
            "leavesFound": []
        },
        "finalStats": {
            "totalLeaves": 0,
            "afterWordFilter": 0,
            "afterYearFilter": 0,
            "afterDateFilter": 0
        },
        "llmDetection": {
            "totalLeaves": 0,
            "successful": 0,
            "failed": 0,
            "selectors": []
        },
        "cssFallback": {
            "triggered": False,
            "success": False,
            "selectors": None
        },
        "diagnostics": {
            "initial_url_http_status": None,
            "robots_browser_retry_attempted": False,
            "robots_browser_retry_status": "not_attempted",

            "robots_txt_found": False,
            "robots_txt_sitemaps_present": False,

            "heuristic_fallback_used": False,
            "heuristic_fallback_status": "not_attempted",
            "heuristic_block_detected": False,
            "heuristic_browser_retry_attempted": False,
            "heuristic_browser_retry_status": "not_attempted",

            "sitemaps_rejected_count": 0,
            "leaf_sitemaps_accepted_count": 0,

            "css_fallback_status": "not_attempted",
            "css_fallback_error_details": "",
            # Filter counts
            "word_after_count": 0,
            "word_rejected_count": 0,
            "year_after_count": 0,
            "year_rejected_count": 0,
            "date_after_count": 0,
            "date_rejected_count": 0,
            # Expansion counts
            "expansion_children_found": 0,
            "children_rejected_word_count": 0,
            "children_rejected_year_count": 0,
            "children_rejected_date_count": 0,
            "leaf_checked_count": 0,
            "leaf_recent_passed_count": 0,
            # Selector counts
            "selector_total_leaves": 0,
            "selector_success_count": 0,
            "selector_failed_count": 0,
            "selector_first_error": "",
            # Leaf extra metrics
            "leaf_total_count": 0,
            "leaf_recency_rejected_count": 0,
            "leaf_title_rejected_count": 0,
        }
    }
    # Bind diagnostic globals for assignments in this function
    global _DIAG_FB_BROWSER_ATTEMPTED, _DIAG_FB_BROWSER_SUCCESS, _DIAG_FB_BLOCK_DETECTED
    
    print(f"\n{'='*80}")
    print(f"🌐 Testing: {url}")
    print(f"{'='*80}\n")
    
    # If sitemap is disabled, go straight to CSS fallback/crawler
    if DISABLE_SITEMAP:
        print(f"[1] SITEMAP DISABLED (DISABLE_SITEMAP=1) -> running CSS fallback directly")
        fallback_result = _css_crawler_or_single(
            base_url=url,
            timeout=60.0,
        )
        if isinstance(fallback_result, dict) and fallback_result.get("ok", True) and int(fallback_result.get("totalSections", 0)) > 0:
            print(f"\n   ✅ CSS FALLBACK SUCCESS!")
            print(f"      📊 Total Sections: {fallback_result['totalSections']}")
            stats["cssFallback"] = {
                "triggered": True,
                "success": True,
                "selectors": fallback_result,
                "reason": "Sitemap disabled"
            }
            try:
                stats["diagnostics"]["css_fallback_status"] = "success"
                stats["diagnostics"]["css_fallback_error_details"] = ""
                stats["diagnostics"]["robots_txt_found"] = False
                stats["diagnostics"]["robots_txt_sitemaps_present"] = False
                stats["diagnostics"]["heuristic_fallback_used"] = False
                stats["diagnostics"]["heuristic_fallback_status"] = "disabled"
            except Exception:
                pass
        else:
            print(f"\n   ❌ CSS FALLBACK FAILED")
            cfail = fallback_result if isinstance(fallback_result, dict) else None
            stats["cssFallback"] = {
                "triggered": True,
                "success": False,
                "reason": "Sitemap disabled, CSS detection failed",
                "failure": cfail if (cfail and isinstance(cfail, dict) and not cfail.get("ok", True)) else None
            }
            try:
                stats["diagnostics"]["css_fallback_status"] = "error"
                stats["diagnostics"]["css_fallback_error_details"] = str((stats["cssFallback"] or {}).get("reason") or "")
                stats["diagnostics"]["robots_txt_found"] = False
                stats["diagnostics"]["robots_txt_sitemaps_present"] = False
                stats["diagnostics"]["heuristic_fallback_used"] = False
                stats["diagnostics"]["heuristic_fallback_status"] = "disabled"
            except Exception:
                pass
        return stats

    # Step 1: Initial probe (robots.txt) + Fetch robots.txt
    try:
        _parts0 = urlsplit(url)
        _base0 = f"{_parts0.scheme}://{_parts0.netloc}" if (_parts0.scheme and _parts0.netloc) else url
        _robots_url0 = urljoin(_base0, "/robots.txt")
    except Exception:
        _robots_url0 = url
    try:
        _probe0 = _http_probe(_robots_url0, timeout=min(5.0, float(timeout) if timeout else 5.0))
    except Exception:
        _probe0 = None
    _classified0 = _classify_probe(_probe0)
    try:
        if isinstance(_probe0, dict):
            stats["diagnostics"]["initial_url_http_status"] = int(_probe0.get("status_code") or 0)
    except Exception:
        pass
    print(f"[1] Fetching robots.txt...")
    _robots_meta = fetch_robots_txt_meta(url, timeout=timeout)
    robots = _robots_meta.get("text") or None
    # Capture robots browser retry diagnostics from meta
    try:
        stats["diagnostics"]["robots_browser_retry_attempted"] = bool(_robots_meta.get("attempted"))
        stats["diagnostics"]["robots_browser_retry_status"] = str(_robots_meta.get("status") or "not_attempted")
    except Exception:
        pass
    if not robots:
        print(f"   ❌ No robots.txt found")
        try:
            stats["diagnostics"]["robots_txt_found"] = False
        except Exception:
            pass
        # Try sitemap fallbacks before CSS
        try:
            _DIAG_FB_BROWSER_ATTEMPTED = False
            _DIAG_FB_BROWSER_SUCCESS = False
            _DIAG_FB_BLOCK_DETECTED = False
        except Exception:
            pass
        fall_urls, fall_reason = _fallback_discover_sitemaps(url, timeout=timeout)
        if fall_urls:
            print(f"   ✅ Fallback sitemaps discovered ({fall_reason}): {len(fall_urls)} URL(s)")
            # robots.txt still not found; but we discovered sitemaps heuristically
            stats["robotsTxt"]["found"] = False
            stats["robotsTxt"]["sitemapsTotal"] = len(fall_urls)
            stats["robotsTxt"]["afterDateFilter"] = 0
            # Continue with these sitemaps downstream (reuse variables below)
            robots = "#fallback"
            sitemaps = fall_urls
            try:
                stats["diagnostics"]["heuristic_fallback_used"] = True
                stats["diagnostics"]["heuristic_fallback_status"] = "success"
                stats["diagnostics"]["heuristic_block_detected"] = bool(_DIAG_FB_BLOCK_DETECTED)
                stats["diagnostics"]["heuristic_browser_retry_attempted"] = bool(_DIAG_FB_BROWSER_ATTEMPTED)
                stats["diagnostics"]["heuristic_browser_retry_status"] = ("bypassed" if _DIAG_FB_BROWSER_SUCCESS else ("failed" if _DIAG_FB_BROWSER_ATTEMPTED else "not_attempted"))
            except Exception:
                pass
        else:
            # Escalate to apex domain fallbacks
            escalated = False
            for root in _apex_roots(url):
                urls2, why2 = _fallback_discover_sitemaps(root, timeout=timeout)
                if urls2:
                    print(f"   ✅ Fallback sitemaps discovered on apex ({root}) ({why2}): {len(urls2)} URL(s)")
                    stats["robotsTxt"]["found"] = True
                    stats["robotsTxt"]["sitemapsTotal"] = len(urls2)
                    sitemaps = urls2
                    escalated = True
                    break
            if not escalated:
                if CSS_FALLBACK_DISABLED:
                    print(f"\n[2] 🚫 CSS FALLBACK DISABLED (DISABLE_CSS_FALLBACK=1)")
                    stats["cssFallback"] = {
                        "triggered": False,
                        "success": False,
                        "reason": "disabled",
                        "failure": None,
                    }
                    try:
                        stats["diagnostics"]["css_fallback_status"] = "skipped"
                        stats["diagnostics"]["css_fallback_error_details"] = "disabled"
                    except Exception:
                        pass
                    return stats
                print(f"\n[2] ⚠️  NO ROBOTS.TXT - Trying CSS Selector Fallback...")
            # Use browser-based detection (exact selector_scraper approach)
            if CSS_FALLBACK_DISABLED:
                print(f"\n[2] 🚫 CSS FALLBACK DISABLED (DISABLE_CSS_FALLBACK=1)")
                stats["cssFallback"] = {
                    "triggered": False,
                    "success": False,
                    "reason": "disabled",
                    "failure": None,
                }
                try:
                    stats["diagnostics"]["css_fallback_status"] = "skipped"
                    stats["diagnostics"]["css_fallback_error_details"] = "disabled"
                except Exception:
                    pass
                return stats
            fallback_result = _css_crawler_or_single(
                base_url=url,
                timeout=60.0,
            )
            
            if isinstance(fallback_result, dict) and fallback_result.get("ok", True) and int(fallback_result.get("totalSections", 0)) > 0:
                print(f"\n   ✅ CSS FALLBACK SUCCESS!")
                print(f"      📊 Total Sections: {fallback_result['totalSections']}")
                print(f"      📄 HTML Length: {fallback_result['htmlLength']} chars")
                print(f"      🔢 Chunks Processed: {fallback_result['chunksProcessed']}")
                
                stats["cssFallback"] = {
                    "triggered": True,
                    "success": True,
                    "selectors": fallback_result,
                    "reason": "No robots.txt found"
                }
                try:
                    stats["diagnostics"]["css_fallback_status"] = "success"
                except Exception:
                    pass
            else:
                print(f"\n   ❌ CSS FALLBACK FAILED")
                cfail = fallback_result if isinstance(fallback_result, dict) else None
                stats["cssFallback"] = {
                    "triggered": True,
                    "success": False,
                    "reason": "No robots.txt found, CSS detection failed",
                    "failure": cfail if (cfail and isinstance(cfail, dict) and not cfail.get("ok", True)) else None
                }
                try:
                    stats["diagnostics"]["css_fallback_status"] = "error"
                    stats["diagnostics"]["css_fallback_error_details"] = str((stats["cssFallback"] or {}).get("reason") or "")
                except Exception:
                    pass
            
            return stats
    
    if robots == "#fallback":
        print(f"   ℹ️ Using heuristic-discovered sitemaps (no robots.txt text)")
        try:
            stats["diagnostics"]["robots_txt_found"] = False
        except Exception:
            pass
    else:
        print(f"   ✅ robots.txt found ({len(robots)} chars)")
        stats["robotsTxt"]["found"] = True
        try:
            stats["diagnostics"]["robots_txt_found"] = True
        except Exception:
            pass
    
    # Step 2: Parse sitemaps from robots.txt
    sitemaps = locals().get("sitemaps") or parse_sitemaps_from_robots(robots, url, news_only=False)
    if not sitemaps:
        print(f"   ❌ No sitemaps in robots.txt")
        # Try sitemap fallbacks before CSS
        try:
            _DIAG_FB_BROWSER_ATTEMPTED = False
            _DIAG_FB_BROWSER_SUCCESS = False
            _DIAG_FB_BLOCK_DETECTED = False
        except Exception:
            pass
        fall_urls, fall_reason = _fallback_discover_sitemaps(url, timeout=timeout)
        if fall_urls:
            print(f"   ✅ Fallback sitemaps discovered ({fall_reason}): {len(fall_urls)} URL(s)")
            sitemaps = fall_urls
            try:
                stats["diagnostics"]["heuristic_fallback_used"] = True
                stats["diagnostics"]["heuristic_fallback_status"] = "success"
                stats["diagnostics"]["heuristic_block_detected"] = bool(_DIAG_FB_BLOCK_DETECTED)
                stats["diagnostics"]["heuristic_browser_retry_attempted"] = bool(_DIAG_FB_BROWSER_ATTEMPTED)
                stats["diagnostics"]["heuristic_browser_retry_status"] = ("bypassed" if _DIAG_FB_BROWSER_SUCCESS else ("failed" if _DIAG_FB_BROWSER_ATTEMPTED else "not_attempted"))
            except Exception:
                pass
        else:
            # Escalate to apex domain fallbacks
            escalated = False
            for root in _apex_roots(url):
                urls2, why2 = _fallback_discover_sitemaps(root, timeout=timeout)
                if urls2:
                    print(f"   ✅ Fallback sitemaps discovered on apex ({root}) ({why2}): {len(urls2)} URL(s)")
                    sitemaps = urls2
                    escalated = True
                    break
            if not escalated:
                if CSS_FALLBACK_DISABLED:
                    print(f"\n[2] 🚫 CSS FALLBACK DISABLED (DISABLE_CSS_FALLBACK=1)")
                    stats["cssFallback"] = {
                        "triggered": False,
                        "success": False,
                        "reason": "disabled",
                        "failure": None,
                    }
                    return stats
                print(f"\n[2] ⚠️  NO SITEMAPS FOUND - Trying CSS Selector Fallback...")
            
            # Use browser-based detection (exact selector_scraper approach)
            if CSS_FALLBACK_DISABLED:
                print(f"\n[2] 🚫 CSS FALLBACK DISABLED (DISABLE_CSS_FALLBACK=1)")
                stats["cssFallback"] = {
                    "triggered": False,
                    "success": False,
                    "reason": "disabled",
                    "failure": None,
                }
                return stats
            fallback_result = _css_crawler_or_single(
                base_url=url,
                timeout=60.0,
            )
            
            if isinstance(fallback_result, dict) and fallback_result.get("ok", True) and int(fallback_result.get("totalSections", 0)) > 0:
                print(f"\n   ✅ CSS FALLBACK SUCCESS!")
                print(f"      📊 Total Sections: {fallback_result['totalSections']}")
                print(f"      📄 HTML Length: {fallback_result['htmlLength']} chars")
                print(f"      🔢 Chunks Processed: {fallback_result['chunksProcessed']}")
                
                stats["cssFallback"] = {
                    "triggered": True,
                    "success": True,
                    "selectors": fallback_result,
                    "reason": "No sitemaps in robots.txt"
                }
            else:
                print(f"\n   ❌ CSS FALLBACK FAILED")
                cfail = fallback_result if isinstance(fallback_result, dict) else None
                stats["cssFallback"] = {
                    "triggered": True,
                    "success": False,
                    "reason": "No sitemaps in robots.txt, CSS detection failed",
                    "failure": cfail if (cfail and isinstance(cfail, dict) and not cfail.get("ok", True)) else None
                }
            
            return stats
    
    stats["robotsTxt"]["sitemapsTotal"] = len(sitemaps)
    print(f"   ✅ Found {len(sitemaps)} sitemap(s) in robots.txt")
    try:
        stats["diagnostics"]["robots_txt_sitemaps_present"] = bool(len(sitemaps) > 0)
        # Since sitemaps present, heuristic fallback not needed
        stats["diagnostics"]["heuristic_fallback_used"] = False
        stats["diagnostics"]["heuristic_fallback_status"] = "not_needed"
    except Exception:
        pass
    
    # Display sitemaps
    print(f"\n   📋 Sitemaps from robots.txt:")
    for idx, sm in enumerate(sitemaps, 1):
        print(f"      {idx:2d}. {sm}")
    print()
    
    # Step 3: Apply Word Filter
    print(f"[2] Applying WORD FILTER...")
    word_kept = []
    word_rejected = []
    for sm in sitemaps:
        should_keep, matched_word = filter_by_words(sm)
        if should_keep:
            word_kept.append(sm)
        else:
            word_rejected.append((sm, matched_word))
            stats["robotsTxt"]["rejected"].append({"url": sm, "filter": "word", "reason": matched_word})
    
    sitemaps = word_kept
    stats["robotsTxt"]["afterWordFilter"] = len(sitemaps)
    try:
        stats["diagnostics"]["word_after_count"] = int(len(sitemaps))
        stats["diagnostics"]["word_rejected_count"] = int(len(word_rejected))
    except Exception:
        pass
    
    if word_rejected:
        print(f"   ❌ Rejected {len(word_rejected)} by keywords:")
        for sm, word in word_rejected[:3]:
            print(f"      • {sm} ('{word}')")
        if len(word_rejected) > 3:
            print(f"      ... and {len(word_rejected) - 3} more")
    
    print(f"   ✅ After word filter: {len(sitemaps)} sitemap(s)")
    print()
    
    # Step 4: Apply Year Filter
    print(f"[3] Applying YEAR FILTER...")
    current_year = datetime.now().year
    year_pattern = r'(19[5-9]\d|20[0-3]\d)'
    
    year_kept = []
    year_rejected = []
    for sm in sitemaps:
        years_found = [int(y) for y in re.findall(year_pattern, sm)]
        
        if not years_found or all(y == current_year for y in years_found):
            year_kept.append(sm)
        else:
            old_years = [str(y) for y in years_found if y != current_year]
            year_rejected.append((sm, old_years))
            stats["robotsTxt"]["rejected"].append({"url": sm, "filter": "year", "reason": f"old years: {', '.join(old_years)}"})
    
    sitemaps = year_kept
    stats["robotsTxt"]["afterYearFilter"] = len(sitemaps)
    try:
        stats["diagnostics"]["year_after_count"] = int(len(sitemaps))
        stats["diagnostics"]["year_rejected_count"] = int(len(year_rejected))
    except Exception:
        pass
    
    if year_rejected:
        print(f"   ❌ Rejected {len(year_rejected)} by year:")
        for sm, years in year_rejected[:5]:
            print(f"      • {sm} (years: {', '.join(years)})")
        if len(year_rejected) > 5:
            print(f"      ... and {len(year_rejected) - 5} more")
    
    print(f"   ✅ After year filter: {len(sitemaps)} sitemap(s)")
    print()
    
    # Step 5: Apply Date Filter
    print(f"[4] Applying DATE FILTER (query params, path dates)...")
    date_kept = []
    date_rejected = []
    for sm in sitemaps:
        should_keep, reason = filter_by_date(sm, hours_threshold=recent_hours, conservative=True)
        if should_keep:
            date_kept.append(sm)
        else:
            date_rejected.append((sm, reason))
            stats["robotsTxt"]["rejected"].append({"url": sm, "filter": "date", "reason": reason})
    
    sitemaps = date_kept
    stats["robotsTxt"]["afterDateFilter"] = len(sitemaps)
    try:
        stats["diagnostics"]["date_after_count"] = int(len(sitemaps))
        stats["diagnostics"]["date_rejected_count"] = int(len(date_rejected))
    except Exception:
        pass
    
    if date_rejected:
        print(f"   ❌ Rejected {len(date_rejected)} by date:")
        for sm, reason in date_rejected[:3]:
            print(f"      • {sm} ({reason})")
        if len(date_rejected) > 3:
            print(f"      ... and {len(date_rejected) - 3} more")
    
    print(f"   ✅ After date filter: {len(sitemaps)} sitemap(s)")
    print()
    
    # Step 6: Recursive Expansion (like _expand_children_recursive)
    print(f"[5] RECURSIVE EXPANSION (with filtering at each level)...")
    print(f"   Testing all {len(sitemaps)} sitemap(s) for expansion...\n")
    
    # Cache latest item datetime per accepted leaf urlset
    leaf_lastmod_map: Dict[str, Optional[datetime]] = {}

    def expand_recursive(sitemap_url: str, depth: int = 0, max_depth: int = 3) -> List[str]:
        """Recursively expand sitemap with filtering (SIMULATION)"""
        if depth >= max_depth:
            return []
        
        indent = "   " * (depth + 1)
        leaves = []
        
        try:
            # Fetch sitemap
            raw = fetch_bytes(sitemap_url, timeout)
            raw = maybe_decompress(sitemap_url, raw)
            if not raw:
                print(f"{indent}⚠️  No content: {sitemap_url}")
                return []
            
            root = parse_xml_bytes(raw)
            if root is None:
                # Save exact body to debug for inspection
                try:
                    os.makedirs("debug_html", exist_ok=True)
                    parsed = urlsplit(sitemap_url)
                    domain = parsed.netloc or "unknown"
                    ts = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S-%fZ")
                    fname = f"nonxml_{domain}_{ts}.html"
                    fpath = os.path.join("debug_html", fname)
                    with open(fpath, "wb") as f:
                        f.write(raw if isinstance(raw, (bytes, bytearray)) else bytes(str(raw or ""), "utf-8"))
                    print(f"{indent}⚠️  Invalid XML: {sitemap_url} -> saved body {_clickable_path(fpath)}")
                except Exception:
                    print(f"{indent}⚠️  Invalid XML: {sitemap_url}")
                return []
            
            tag = (root.tag or "").lower()
            
            if tag.endswith("sitemapindex"):
                # It's an INDEX - expand children
                children = root.findall(".//{*}sitemap")
                print(f"{indent}📂 INDEX: {len(children)} child sitemap(s) in {sitemap_url}")
                
                for idx, smnode in enumerate(children, 1):
                    loc = child_text_any_ns(smnode, "loc")
                    if not loc:
                        continue
                    
                    child_url = loc.strip()
                    
                    # Filter 1: Word
                    should_keep_word, matched_word = filter_by_words(child_url)
                    if not should_keep_word:
                        print(f"{indent}  [{idx}] ❌ {child_url} (word: '{matched_word}')")
                        stats["recursiveExpansion"]["childrenRejectedWord"].append(child_url)
                        continue
                    
                    # Filter 2: Year
                    years_found = [int(y) for y in re.findall(year_pattern, child_url)]
                    if years_found and not all(y == current_year for y in years_found):
                        old_years = [str(y) for y in years_found if y != current_year]
                        print(f"{indent}  [{idx}] ❌ {child_url} (year: {', '.join(old_years)})")
                        stats["recursiveExpansion"]["childrenRejectedYear"].append(child_url)
                        continue
                    
                    # Filter 3: Date
                    lastmod_str = child_text_any_ns(smnode, "lastmod")
                    child_lastmod = _parse_w3c_datetime(lastmod_str) if lastmod_str else None
                    should_keep_date, date_reason = filter_by_date(
                        child_url,
                        hours_threshold=recent_hours,
                        conservative=True,
                        xml_lastmod=child_lastmod
                    )
                    if not should_keep_date:
                        print(f"{indent}  [{idx}] ❌ {child_url} (date: {date_reason})")
                        stats["recursiveExpansion"]["childrenRejectedDate"].append(child_url)
                        continue
                    
                    # Passed all filters - recurse
                    print(f"{indent}  [{idx}] ✅ {child_url} (expanding...)")
                    child_leaves = expand_recursive(child_url, depth + 1, max_depth)
                    leaves.extend(child_leaves)
                
                stats["recursiveExpansion"]["childrenFound"] += len(children)
                
            elif tag.endswith("urlset"):
                # It's a LEAF urlset — sample items and decide recency using item-level dates
                url_nodes = root.findall(".//{*}url")
                url_count = len(url_nodes)
                print(f"{indent}📄 LEAF: {url_count} article(s) in {sitemap_url}")
                try:
                    stats["diagnostics"]["leaf_total_count"] = int(stats.get("diagnostics", {}).get("leaf_total_count", 0)) + 1
                except Exception:
                    pass

                # Sample top 5 and bottom 5 to handle ascending/descending orders
                recent_count = 0
                latest_dt: Optional[datetime] = None
                title_found: bool = False

                sample_nodes = []
                if url_nodes:
                    top_k = 5
                    bottom_k = 5
                    first_part = url_nodes[:top_k]
                    last_part = url_nodes[-bottom_k:] if len(url_nodes) > top_k else []
                    seen_node_ids = set()
                    for n in (first_part + last_part):
                        if n is None:
                            continue
                        key = id(n)
                        if key in seen_node_ids:
                            continue
                        seen_node_ids.add(key)
                        sample_nodes.append(n)

                for url_node in sample_nodes:
                    # Prefer <lastmod>, else publication_date from news namespace
                    dtxt = child_text_any_ns(url_node, "lastmod")
                    if not dtxt:
                        try:
                            pub_el = url_node.find(".//{*}publication_date")
                            dtxt = (pub_el.text or "").strip() if pub_el is not None else ""
                        except Exception:
                            dtxt = ""
                    # Title-like detection (optional)
                    if SITEMAP_REQUIRE_ANY_TITLE and not title_found:
                        try:
                            for desc in url_node.iter():
                                tag_name = desc.tag.split('}')[-1] if '}' in (desc.tag or '') else (desc.tag or '')
                                if tag_name.lower() in ("title", "headline", "name"):
                                    val = (desc.text or "").strip()
                                    if val:
                                        title_found = True
                                        break
                        except Exception:
                            pass
                    if not dtxt:
                        continue
                    try:
                        dt = _parse_w3c_datetime(dtxt)
                    except Exception:
                        dt = None
                    if not dt:
                        continue
                    # Track latest date
                    if latest_dt is None or dt > latest_dt:
                        latest_dt = dt
                    # Check recency window
                    try:
                        age_hours = (datetime.utcnow() - dt.replace(tzinfo=None)).total_seconds() / 3600.0
                        if age_hours <= float(recent_hours):
                            recent_count += 1
                    except Exception:
                        pass

                if recent_count >= 1 and (not SITEMAP_REQUIRE_ANY_TITLE or title_found):
                    leaves.append(sitemap_url)
                    stats["recursiveExpansion"]["leavesFound"].append(sitemap_url)
                    # Cache latest item dt for final filter stage
                    leaf_lastmod_map[sitemap_url] = latest_dt
                else:
                    if recent_count < 1:
                        print(f"{indent}  ❌ Rejecting leaf (no items within last {recent_hours}h; sampled={len(sample_nodes)})")
                        try:
                            stats["diagnostics"]["leaf_recency_rejected_count"] = int(stats.get("diagnostics", {}).get("leaf_recency_rejected_count", 0)) + 1
                        except Exception:
                            pass
                    elif SITEMAP_REQUIRE_ANY_TITLE and not title_found:
                        print(f"{indent}  ❌ Rejecting leaf (no title-like field in sampled items)")
                        try:
                            stats["recursiveExpansion"]["childrenRejectedTitle"].append(sitemap_url)
                            stats["diagnostics"]["leaf_title_rejected_count"] = int(stats.get("diagnostics", {}).get("leaf_title_rejected_count", 0)) + 1
                        except Exception:
                            pass
            
        except Exception as e:
            print(f"{indent}❌ Error: {type(e).__name__}")
        
        return leaves
    
    # Test expansion on first few sitemaps
    all_leaves = []
    for idx, sm in enumerate(sitemaps, 1):
        print(f"   [{idx}/{min(3, len(sitemaps))}] Expanding: {sm}")
        leaves = expand_recursive(sm, depth=0, max_depth=max_depth)
        all_leaves.extend(leaves)
        print()

    
    # Deduplicate leaves before final filtering
    all_leaves = list(dict.fromkeys(all_leaves))
    stats["finalStats"]["totalLeaves"] = len(all_leaves)
    try:
        stats["diagnostics"]["leaf_checked_count"] = int(len(all_leaves))
    except Exception:
        pass
    
    # Step 7: Final Filtering on Collected Leaves
    if all_leaves:
        print(f"[6] FINAL FILTERING on {len(all_leaves)} collected leaf sitemap(s)...")
        
        # Word filter
        word_kept = []
        for leaf in all_leaves:
            should_keep, matched_word = filter_by_words(leaf)
            if should_keep:
                word_kept.append(leaf)
        stats["finalStats"]["afterWordFilter"] = len(word_kept)
        print(f"   Word Filter: {len(all_leaves)} → {len(word_kept)}")
        
        # Year filter
        year_kept = filter_sitemaps_by_year(word_kept)
        stats["finalStats"]["afterYearFilter"] = len(year_kept)
        print(f"   Year Filter: {len(word_kept)} → {len(year_kept)}")
        
        # Date filter (use cached latest item dt when available)
        date_kept = []
        for leaf in year_kept:
            xml_dt = leaf_lastmod_map.get(leaf)
            should_keep, reason = filter_by_date(leaf, hours_threshold=recent_hours, conservative=True, xml_lastmod=xml_dt)
            if should_keep:
                date_kept.append(leaf)
        stats["finalStats"]["afterDateFilter"] = len(date_kept)
        print(f"   Date Filter: {len(year_kept)} → {len(date_kept)}")
        try:
            stats["diagnostics"]["sitemaps_rejected_count"] = int(len(stats.get("robotsTxt", {}).get("rejected") or []))
            stats["diagnostics"]["leaf_sitemaps_accepted_count"] = int(len(date_kept))
            stats["diagnostics"]["expansion_children_found"] = int(stats.get("recursiveExpansion", {}).get("childrenFound") or 0)
            stats["diagnostics"]["children_rejected_word_count"] = int(len(stats.get("recursiveExpansion", {}).get("childrenRejectedWord") or []))
            stats["diagnostics"]["children_rejected_year_count"] = int(len(stats.get("recursiveExpansion", {}).get("childrenRejectedYear") or []))
            stats["diagnostics"]["children_rejected_date_count"] = int(len(stats.get("recursiveExpansion", {}).get("childrenRejectedDate") or []))
        except Exception:
            pass
        
        print(f"\n   ✅ FINAL: {len(date_kept)} leaf sitemap(s) ready for LLM detection")
        
        if date_kept:
            print(f"\n   📋 Final Leaf Sitemaps:")
            for idx, leaf in enumerate(date_kept, 1):
                print(f"      {idx:2d}. {leaf}")
            
            # 🆕 Step 7: LLM Selector Detection on Kept Leaves
            print(f"\n[7] 🤖 LLM SELECTOR DETECTION on {len(date_kept)} kept leaf sitemap(s)...")
            
            def _detect_one(leaf_url: str) -> Dict[str, Any]:
                print(f"\n   [LLM] Detecting: {leaf_url}")
                try:
                    detected, llm_err = _detect_selectors_with_llm(leaf_url, timeout=timeout)
                    if not detected:
                        reason = f" ({llm_err[:80]})" if llm_err else ""
                        print(f"      ⚠️  LLM failed, trying basic detection{reason}...")
                        detected = _detect_selectors_basic(leaf_url, timeout=timeout)
                    if detected:
                        fields = detected.get('fields', {})
                        method = detected.get('detectionMethod', 'unknown')
                        confidence = detected.get('confidence', 0)
                        # Enforce title-like field presence if required
                        if SITEMAP_REQUIRE_ANY_TITLE:
                            keys_low = set([str(k or '').lower() for k in (fields or {}).keys()])
                            if not ("title" in keys_low or "headline" in keys_low or "name" in keys_low or "news:title" in keys_low):
                                print(f"      ❌ Rejecting detection (no title-like field in selectors)")
                                return {"url": leaf_url, "detectedSelectors": None, "error": "no title-like field in selectors"}
                        print(f"      ✅ Detected {len(fields)} field(s) (method={method}, confidence={confidence})")
                        print(f"      📋 Fields:")
                        for field_name, field_path in list(fields.items())[:5]:
                            print(f"         • {field_name}: {field_path}")
                        if len(fields) > 5:
                            print(f"         ... and {len(fields) - 5} more")
                        return {
                            "url": leaf_url,
                            "detectedSelectors": detected,
                            "fieldCount": len(fields),
                            "method": method,
                            "confidence": confidence
                        }
                    else:
                        print(f"      ❌ Detection failed completely")
                        return {"url": leaf_url, "detectedSelectors": None, "error": "No detection result"}
                except Exception as e:
                    print(f"      ❌ Error: {type(e).__name__}: {str(e)[:50]}")
                    return {"url": leaf_url, "detectedSelectors": None, "error": str(e)}

            detected_selectors: List[Dict[str, Any]] = []
            if max(1, int(llm_concurrency)) > 1 and len(date_kept) > 1:
                max_workers = max(1, int(llm_concurrency))
                with cf.ThreadPoolExecutor(max_workers=max_workers) as ex:
                    futures = [ex.submit(_detect_one, leaf_url) for leaf_url in date_kept]
                    for fut in cf.as_completed(futures):
                        try:
                            detected_selectors.append(fut.result())
                        except Exception as e:
                            detected_selectors.append({"url": None, "detectedSelectors": None, "error": str(e)})
            else:
                for leaf_url in date_kept:
                    detected_selectors.append(_detect_one(leaf_url))
            
            # Update stats
            stats["llmDetection"] = {
                "totalLeaves": len(date_kept),
                "successful": sum(1 for d in detected_selectors if d.get("detectedSelectors")),
                "failed": sum(1 for d in detected_selectors if not d.get("detectedSelectors")),
                "selectors": detected_selectors
            }
            try:
                stats["diagnostics"]["selector_total_leaves"] = int(stats["llmDetection"]["totalLeaves"])
                stats["diagnostics"]["selector_success_count"] = int(stats["llmDetection"]["successful"])
                stats["diagnostics"]["selector_failed_count"] = int(stats["llmDetection"]["failed"])
                _first_err = ""
                for d in detected_selectors:
                    e = (d or {}).get("error")
                    if e:
                        _first_err = str(e)
                        break
                stats["diagnostics"]["selector_first_error"] = _first_err
            except Exception:
                pass
            
            print(f"\n   ✅ LLM Detection Complete: {stats['llmDetection']['successful']}/{len(date_kept)} successful")
    
    # Step 8: Apex escalation then CSS Selector Fallback (if no leaves found)
    if stats["finalStats"]["afterDateFilter"] == 0:
        # Try apex domain discovery before CSS
        escalated_success = False
        for root in _apex_roots(url):
            try:
                print(f"\n[8] 🔼 Escalating to apex domain for discovery: {root}")
                apex_stats = test_recursive_expansion(
                    root,
                    recent_hours=recent_hours,
                    timeout=timeout,
                    max_depth=max_depth,
                    llm_concurrency=llm_concurrency,
                )
                if int((apex_stats.get("finalStats") or {}).get("afterDateFilter") or 0) > 0:
                    print(f"   ✅ Apex discovery produced {int((apex_stats['finalStats'] or {}).get('afterDateFilter') or 0)} leaf sitemap(s).")
                    return apex_stats
            except Exception:
                pass

        if CSS_FALLBACK_DISABLED:
            print(f"\n[8] 🚫 CSS FALLBACK DISABLED (DISABLE_CSS_FALLBACK=1)")
            stats["cssFallback"] = {
                "triggered": False,
                "success": False,
                "reason": "disabled",
                "failure": None,
            }
            try:
                stats["diagnostics"]["css_fallback_status"] = "skipped"
                stats["diagnostics"]["css_fallback_error_details"] = "disabled"
            except Exception:
                pass
        else:
            print(f"\n[8] ⚠️  NO SITEMAP LEAVES FOUND - Trying CSS Selector Fallback...")
            
            # Use browser-based detection (exact selector_scraper approach)
            fallback_result = _css_selector_fallback(
                base_url=url,
                timeout=60.0,
                headful=False,  # Set to True for debugging
                slowmo_ms=0
            )
            
            if isinstance(fallback_result, dict) and fallback_result.get("ok", True) and int(fallback_result.get("totalSections", 0)) > 0:
                print(f"\n   ✅ CSS FALLBACK SUCCESS!")
                print(f"      📊 Total Sections: {fallback_result['totalSections']}")
                print(f"      📄 HTML Length: {fallback_result['htmlLength']} chars")
                print(f"      🔢 Chunks Processed: {fallback_result['chunksProcessed']}")
                
                # Display sections
                print(f"\n      📋 Detected Sections:")
                for idx, section in enumerate(fallback_result["sections"], 1):
                    name = section.get("sectionName", f"Section {idx}")
                    conf = section.get("confidence", 0)
                    selectors = section.get("selectors", {})
                    chunk = section.get("detectedInChunk", "?")
                    print(f"         {idx}. {name} (confidence={conf}, chunk={chunk})")
                    print(f"            • title: {selectors.get('title', 'N/A')}")
                    print(f"            • link: {selectors.get('link', 'N/A')}")
                    if selectors.get('date'):
                        print(f"            • date: {selectors.get('date')}")
                    if selectors.get('description'):
                        print(f"            • description: {selectors.get('description')}")
                    if selectors.get('author'):
                        print(f"            • author: {selectors.get('author')}")
                
                # Display debug files
                print(f"\n      💾 Debug Files:")
                print(f"         • HTML: {_clickable_path(fallback_result['debugFiles']['html'])}")
                print(f"         • LLM Responses: {len(fallback_result['debugFiles']['llmResponses'])} file(s)")
                
                stats["cssFallback"] = {
                    "triggered": True,
                    "success": True,
                    "selectors": fallback_result
                }
                try:
                    stats["diagnostics"]["css_fallback_status"] = "success"
                except Exception:
                    pass
            else:
                print(f"\n   ❌ CSS FALLBACK FAILED")
                cfail = fallback_result if isinstance(fallback_result, dict) else None
                stats["cssFallback"] = {
                    "triggered": True,
                    "success": False,
                    "reason": "Could not detect selectors",
                    "failure": cfail if (cfail and isinstance(cfail, dict) and not cfail.get("ok", True)) else None
                }
                try:
                    stats["diagnostics"]["css_fallback_status"] = "error"
                    stats["diagnostics"]["css_fallback_error_details"] = str((stats["cssFallback"] or {}).get("reason") or "")
                except Exception:
                    pass
    else:
        stats["cssFallback"]["reason"] = f"{stats['finalStats']['afterDateFilter']} sitemap leaves found, fallback not needed"
        try:
            # CSS fallback not needed (leaves already present)
            stats["diagnostics"]["css_fallback_status"] = "not_needed"
            stats["diagnostics"]["css_fallback_error_details"] = ""
        except Exception:
            pass
    
    return stats


def main():
    import json
    from datetime import datetime
    import argparse
    
    parser = argparse.ArgumentParser(description="Selection Extraction Pipeline with optional concurrency")
    parser.add_argument("--site-concurrency", type=int, default=1, dest="site_concurrency", help="Parallel sites to process (default 1)")
    parser.add_argument("--llm-concurrency", type=int, default=3, dest="llm_concurrency", help="Parallel LLM detections per site (default 3)")
    parser.add_argument("--recent-hours", type=int, default=24, dest="recent_hours", help="Recent hours threshold for date filtering in tests (default 24)")
    parser.add_argument("--timeout", type=float, default=15.0, dest="timeout", help="Per-request timeout seconds (default 15)")
    parser.add_argument("--max-depth", type=int, default=2, dest="max_depth", help="Max sitemap recursion depth (default 2)")
    args = parser.parse_args()

    print("=" * 80)
    print("🧪 SITEMAP FILTERING TEST - Recursive Expansion")
    print("=" * 80)
    print(f"⚙️  Concurrency: sites={args.site_concurrency} llm={args.llm_concurrency}")
    print()
    
    # Read URLs from Excel
    print("📂 Reading URLs from urls.xlsx...")
    urls = read_urls_from_excel("urls.xlsx")
    # Log the count of URLs read from Excel (JSONL)
    try:
        from datetime import datetime as _dt
        _log_path = os.path.join(os.path.dirname(__file__) or ".", "urls_read_log.jsonl")
        _rec = {
            "ts": _dt.utcnow().isoformat() + "Z",
            "excel": os.path.abspath("urls.xlsx"),
            "urlsRead": int(len(urls or [])),
        }
        with open(_log_path, "a", encoding="utf-8") as _f:
            json.dump(_rec, _f, ensure_ascii=False)
            _f.write("\n")
            _f.flush()
            try:
                os.fsync(_f.fileno())
            except Exception:
                pass
    except Exception:
        pass
    
    if not urls:
        print("❌ No URLs found in Excel file!")
        return
    
    print(f"✅ Found {len(urls)} URL(s)")
    print()
    
    # Display URLs
    print("📋 URLs from Excel:")
    print("-" * 80)
    for idx, url in enumerate(urls, 1):
        print(f"  {idx:2d}. {url}")
    print()
    
    # Prepare streaming output (overwrite from previous runs)
    stream_file = "selection_extraction_report_stream.jsonl"
    try:
        with open(stream_file, "w", encoding="utf-8") as _f:
            pass
    except Exception:
        pass

    # Test each URL (ALL URLs from Excel) with optional site-level concurrency
    all_stats: List[Dict[str, Any]] = []
    print(f"⚡ Processing ALL {len(urls)} URL(s) from Excel...\n")

    def _site_job(target_url: str) -> Dict[str, Any]:
        print(f"\n{'▓'*80}")
        print(f"Processing: {target_url}")
        print(f"{'▓'*80}")
        try:
            stats_local = test_recursive_expansion(
                target_url,
                recent_hours=args.recent_hours,
                timeout=args.timeout,
                max_depth=args.max_depth,
                llm_concurrency=args.llm_concurrency,
            )
        except Exception as e:
            # Ensure every site emits one stream record
            err_msg = f"{type(e).__name__}: {str(e)[:160]}"
            stats_local = _default_stats(url=target_url, err=err_msg)
        # Stream this website's result immediately
        _append_stream({
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC"),
            "result": stats_local
        }, stream_path=stream_file)
        print(f"\n📝 Streamed result to {_clickable_path(stream_file)}")

        # === CSV upsert for discovery stage ===
        try:
            domain = _normalize_domain(stats_local.get("url") or target_url)
            source_id = stats_local.get("url") or target_url
            robots = stats_local.get("robotsTxt") or {}
            final_stats = stats_local.get("finalStats") or {}
            cssf = stats_local.get("cssFallback") or {}
            llm = stats_local.get("llmDetection") or {}

            leaves = int(final_stats.get("afterDateFilter") or 0)

            # Decide sitemap status
            sitemap_status = "Not Attempted"
            sitemap_detail = ""
            robots_found = bool(robots.get("found"))
            total_sm = int(robots.get("sitemapsTotal") or 0)
            after_date = int(robots.get("afterDateFilter") or 0)

            if not robots_found:
                sitemap_status = "Not Found"
                sitemap_detail = "no robots.txt"
            elif total_sm == 0:
                sitemap_status = "Not Found"
                sitemap_detail = "no sitemaps in robots.txt"
            elif leaves > 0:
                sitemap_status = "Success"
            elif after_date == 0:
                sitemap_status = "Empty"
            else:
                sitemap_status = "Empty"

            # CSS fallback
            css_status = "Not Attempted"
            css_detail = ""
            if bool(cssf.get("triggered")):
                if bool(cssf.get("success")):
                    css_status = "Success"
                else:
                    css_status = "Error"
                    # Build detailed failure message
                    failure = cssf.get("failure") if isinstance(cssf.get("failure"), dict) else None
                    if failure and (failure.get("ok") is False):
                        stage = str(failure.get("stage") or "?")
                        etype = str(failure.get("errorType") or "")
                        emsg = str(failure.get("errorMessage") or "")
                        ctx = failure.get("context") or {}
                        parts = [f"{stage}"]
                        if etype:
                            parts.append(etype)
                        if emsg:
                            parts.append(emsg)
                        # Small context summary
                        try:
                            html_len = ctx.get("htmlLength")
                            chunks = ctx.get("chunksProcessed") or ctx.get("chunks")
                            cand = ctx.get("candidatesTotal")
                            strict_ok = ctx.get("strictAccepted")
                            len_ok = ctx.get("lenientAccepted")
                            ctx_bits = []
                            if html_len is not None:
                                ctx_bits.append(f"html={html_len}")
                            if chunks is not None:
                                ctx_bits.append(f"chunks={chunks}")
                            if cand is not None:
                                ctx_bits.append(f"candidates={cand}")
                            if strict_ok is not None or len_ok is not None:
                                ctx_bits.append(f"accepted={strict_ok or 0}/{len_ok or 0}")
                            if ctx_bits:
                                parts.append("; " + ", ".join(ctx_bits))
                        except Exception:
                            pass
                        css_detail = ": ".join([p for p in parts if p])[:200]
                    else:
                        css_detail = str(cssf.get("reason") or "")
            else:
                css_status = "Skipped" if leaves > 0 else "Not Attempted"

            # Override when CSS fallback is globally disabled
            try:
                if CSS_FALLBACK_DISABLED:
                    css_status = "Skipped"
                    css_detail = "disabled"
            except Exception:
                pass

            # Selector discovery attempt error (why detection failed)
            sel_attempt_error = ""
            total_leaves = int(llm.get("totalLeaves") or 0)
            succ = int(llm.get("successful") or 0)
            failed = int(llm.get("failed") or 0)
            if not robots_found:
                sel_attempt_error = "no robots.txt"
            elif total_sm == 0:
                sel_attempt_error = "no sitemaps in robots.txt"
            elif leaves == 0:
                sel_attempt_error = "no recent leaf sitemaps"
            elif total_leaves > 0 and succ == 0:
                # Try to pull first error detail from selectors list
                first_err = ""
                try:
                    selectors_list = llm.get("selectors") or []
                    for s in selectors_list:
                        e = (s or {}).get("error")
                        if e:
                            first_err = str(e)
                            break
                except Exception:
                    first_err = ""
                sel_attempt_error = ("selector detection failed for all leaves" + (f": {first_err}" if first_err else ""))

            # Build structured error detail (JSON) for the new column
            try:
                err_detail_obj = _build_selector_error_detail(
                    base_url=target_url,
                    robots_found=robots_found,
                    total_sm=total_sm,
                    leaves=leaves,
                    after_date=after_date,
                    llm=llm,
                    cssf=cssf,
                )
            except Exception:
                err_detail_obj = None
            # Summarize only status code and errorMessage for the CSV/DB column
            try:
                err_detail_json = _summarize_error_response(err_detail_obj, llm, cssf)
            except Exception:
                err_detail_json = ""

            # Prepare leaf sitemap URLs string (limit length for CSV readability)
            leaf_urls_list = stats_local.get("finalLeafUrls") or []
            try:
                if not leaf_urls_list and isinstance(stats_local.get("finalStats"), dict):
                    # Derive from leavesFound if present
                    lf = stats_local.get("recursiveExpansion", {}).get("leavesFound") or []
                    leaf_urls_list = list(dict.fromkeys(lf))
            except Exception:
                pass
            leaf_urls_joined = " | ".join(list(map(str, leaf_urls_list))[:25])
            if isinstance(leaf_urls_list, list) and len(leaf_urls_list) > 25:
                leaf_urls_joined += f" | ... (+{len(leaf_urls_list) - 25} more)"

            # Build discovery error segment
            disc_reasons = []
            if not robots_found:
                disc_reasons.append("no_robots")
            elif total_sm == 0:
                disc_reasons.append("no_sitemaps")
            else:
                if leaves == 0 and after_date == 0:
                    disc_reasons.append("no_recent_leaves")

            try:
                total_leaves = int(llm.get("totalLeaves") or 0)
                succ = int(llm.get("successful") or 0)
                if total_leaves > 0 and succ == 0:
                    # Pull first error if present
                    first_err = ""
                    try:
                        selectors_list = llm.get("selectors") or []
                        for s in selectors_list:
                            e = (s or {}).get("error")
                            if e:
                                first_err = str(e)
                                break
                    except Exception:
                        first_err = ""
                    if first_err:
                        disc_reasons.append(f"selector_detection_failed: {first_err[:80]}")
                    else:
                        disc_reasons.append("selector_detection_failed")
            except Exception:
                pass

            try:
                if bool(cssf.get("triggered")):
                    if not bool(cssf.get("success")):
                        failure = cssf.get("failure") if isinstance(cssf.get("failure"), dict) else None
                        if failure and (failure.get("ok") is False):
                            stage = str(failure.get("stage") or "?")
                            etype = str(failure.get("errorType") or "")
                            emsg = str(failure.get("errorMessage") or "")
                            seg = "css_failed: " + ": ".join([p for p in [stage, etype, emsg[:60]] if p])
                            disc_reasons.append(seg)
                        else:
                            reason = str(cssf.get("reason") or "")
                            disc_reasons.append("css_failed" + (f": {reason[:60]}" if reason else ""))
                else:
                    if leaves > 0:
                        disc_reasons.append("css_skipped")
            except Exception:
                pass

            disc_seg = ""
            if disc_reasons:
                ctx_bits = []
                try:
                    ctx_bits.append(f"sitemaps={total_sm}")
                    ctx_bits.append(f"afterDate={after_date}")
                    ctx_bits.append(f"leaves={leaves}")
                except Exception:
                    pass
                ctx = "; ".join(ctx_bits)
                disc_seg = "discovery: " + ", ".join(disc_reasons) + (f"; {ctx}" if ctx else "")

            # Derive always-present numeric counts to avoid blanks
            try:
                _rej_len = int(len((stats_local.get("robotsTxt") or {}).get("rejected") or []))
            except Exception:
                _rej_len = 0
            try:
                _leaf_acc = int((stats_local.get("finalStats") or {}).get("afterDateFilter") or 0)
            except Exception:
                _leaf_acc = 0

            diag = stats_local.get("diagnostics") or {}

            updates_map = {
                "Domain (sources)": domain,
                "Selector Discovery Attempted": "Yes",
                "Selector Discovery Attempt Error": sel_attempt_error,
                "Selector Discovery Attempt Error Response": err_detail_json,
                "Sitemap Processing Status": sitemap_status,
                "Sitemap Processing Error Details": sitemap_detail,
                "leaf Sitemap URLs Discovered": str(leaves),
                "CSS Fallback Status": css_status,
                "CSS Fallback error Details": css_detail,
                "Leaf Sitemap URLs": leaf_urls_joined,
                "Selector Discovery Not Attempted Reason": "",
                # Map diagnostics to dedicated columns
                "initial_url_http_status": _to_str(diag.get("initial_url_http_status"), "0"),
                "robots_browser_retry_attempted": _to_str(diag.get("robots_browser_retry_attempted"), "false"),
                "robots_browser_retry_status": _to_str(diag.get("robots_browser_retry_status"), "not_needed"),
                "robots_txt_found": _to_str(diag.get("robots_txt_found"), "false"),
                "robots_txt_sitemaps_present": _to_str(diag.get("robots_txt_sitemaps_present"), "false"),
                "heuristic_fallback_used": _to_str(diag.get("heuristic_fallback_used"), "false"),
                "heuristic_fallback_status": _to_str(diag.get("heuristic_fallback_status"), "not_needed"),
                "heuristic_block_detected": _to_str(diag.get("heuristic_block_detected"), "false"),
                "heuristic_browser_retry_attempted": _to_str(diag.get("heuristic_browser_retry_attempted"), "false"),
                "heuristic_browser_retry_status": _to_str(diag.get("heuristic_browser_retry_status"), "not_attempted"),
                "sitemaps_rejected_count": _to_str(diag.get("sitemaps_rejected_count", _rej_len), "0"),
                "leaf_sitemaps_accepted_count": _to_str(diag.get("leaf_sitemaps_accepted_count", _leaf_acc), "0"),
                "css_fallback_status": _to_str(diag.get("css_fallback_status"), "not_needed"),
                "css_fallback_error_details": _to_str(diag.get("css_fallback_error_details"), ""),
                # Filter counts
                "word_after_count": _to_str(diag.get("word_after_count"), "0"),
                "word_rejected_count": _to_str(diag.get("word_rejected_count"), "0"),
                "year_after_count": _to_str(diag.get("year_after_count"), "0"),
                "year_rejected_count": _to_str(diag.get("year_rejected_count"), "0"),
                "date_after_count": _to_str(diag.get("date_after_count"), "0"),
                "date_rejected_count": _to_str(diag.get("date_rejected_count"), "0"),
                # Expansion counts
                "expansion_children_found": _to_str(diag.get("expansion_children_found"), "0"),
                "children_rejected_word_count": _to_str(diag.get("children_rejected_word_count"), "0"),
                "children_rejected_year_count": _to_str(diag.get("children_rejected_year_count"), "0"),
                "children_rejected_date_count": _to_str(diag.get("children_rejected_date_count"), "0"),
                "leaf_checked_count": _to_str(diag.get("leaf_checked_count"), "0"),
                "leaf_recent_passed_count": _to_str(diag.get("leaf_sitemaps_accepted_count"), "0"),
                # Selector detection counts
                "selector_total_leaves": _to_str(diag.get("selector_total_leaves"), "0"),
                "selector_success_count": _to_str(diag.get("selector_success_count"), "0"),
                "selector_failed_count": _to_str(diag.get("selector_failed_count"), "0"),
                "selector_first_error": _to_str(diag.get("selector_first_error"), ""),
                # Leaf extra metrics
                "leaf_total_count": _to_str(diag.get("leaf_total_count"), "0"),
                "leaf_recency_rejected_count": _to_str(diag.get("leaf_recency_rejected_count"), "0"),
                "leaf_title_rejected_count": _to_str(diag.get("leaf_title_rejected_count"), "0"),
            }
            if disc_seg:
                updates_map["Overall pipelines Error Details"] = disc_seg

            # Friendly human explanation (single best sentence)
            disc_msg = ""
            try:
                # Priority: no_robots > no_sitemaps > no_recent_leaves > selector_detection_failed > css_failed > css_skipped
                if not robots_found:
                    disc_msg = "Discovery: robots.txt not found."
                elif total_sm == 0:
                    disc_msg = "Discovery: robots.txt found but no sitemaps listed."
                elif leaves == 0 and after_date == 0:
                    disc_msg = "Discovery: sitemaps found but none looked recent."
                else:
                    # selector detection failure
                    total_leaves = int(llm.get("totalLeaves") or 0)
                    succ = int(llm.get("successful") or 0)
                    if total_leaves > 0 and succ == 0:
                        disc_msg = "Discovery: couldn't detect fields in sitemap items."
                    elif not bool(cssf.get("triggered")) and leaves > 0:
                        disc_msg = "Discovery: CSS fallback skipped (recent sitemaps available)."
            except Exception:
                pass

            # Also append CSS fallback failure message (if any), even when robots/sitemap messages take priority
            try:
                css_msg = ""
                if bool(cssf.get("triggered")) and not bool(cssf.get("success")):
                    failure = cssf.get("failure") if isinstance(cssf.get("failure"), dict) else None
                    if failure and (failure.get("ok") is False):
                        stage = str(failure.get("stage") or "?")
                        etype = str(failure.get("errorType") or "")
                        if stage or etype:
                            css_msg = f"Discovery: CSS fallback failed during {stage} ({etype}).".strip()
                        else:
                            css_msg = "Discovery: CSS fallback failed."
                    else:
                        css_msg = "Discovery: CSS fallback failed."
                elif not bool(cssf.get("triggered")) and leaves > 0:
                    css_msg = "Discovery: CSS fallback skipped (recent sitemaps available)."
                if css_msg:
                    disc_msg = _merge_friendly_explanation(disc_msg, css_msg)
            except Exception:
                pass

            if disc_msg:
                updates_map["Overall pipelines Explanation"] = disc_msg

            # If error occurred at top-level, add a friendly explanation segment
            try:
                top_err = str(stats_local.get("error") or "").strip()
                if top_err:
                    updates_map["Overall pipelines Error Details"] = _merge_overall_error(updates_map.get("Overall pipelines Error Details", ""), f"top_error: {top_err}")
                    # Friendly message
                    updates_map["Overall pipelines Explanation"] = _merge_friendly_explanation(updates_map.get("Overall pipelines Explanation", ""), "Discovery: pipeline error occurred during processing.")
            except Exception:
                pass

            # Use original source URL as the CSV key instead of domain
            # Initialize DB once (idempotent)
            try:
                ov_init_db()
            except Exception:
                pass

            updates_map["Domain (sources)"] = source_id
            try:
                ov_upsert(source_id, updates_map)
            except Exception:
                pass
        except Exception:
            pass
        return stats_local

    site_workers = max(1, int(args.site_concurrency))
    if site_workers == 1 or len(urls) <= 1:
        for idx, u in enumerate(urls, 1):
            stats = _site_job(u)
            all_stats.append(stats)
            percentage = (idx / len(urls)) * 100
            print(f"\n✅ Progress: {idx}/{len(urls)} ({percentage:.1f}%) completed")
    else:
        completed = 0
        with cf.ThreadPoolExecutor(max_workers=site_workers) as ex:
            futures = [ex.submit(_site_job, u) for u in urls]
            total = len(futures)
            for fut in cf.as_completed(futures):
                try:
                    res = fut.result()
                except Exception as e:
                    res = _default_stats(url=None, err=str(e))
                all_stats.append(res)
                completed += 1
                percentage = (completed / total) * 100
                print(f"\n✅ Progress: {completed}/{total} ({percentage:.1f}%) completed")
    
    # Final CSV Export from SQLite
    try:
        ov_export_csv()
        print("\n📄 Exported pipelines_overview.csv from SQLite store")
    except Exception:
        print("\n⚠️ Failed to export CSV from SQLite store")

    # Overall Summary
    print("\n" + "=" * 80)
    print("📊 OVERALL SUMMARY")
    print("=" * 80)
    
    total_sitemaps = sum(int((s.get("robotsTxt") or {}).get("sitemapsTotal") or 0) for s in all_stats)
    total_after_filters = sum(int((s.get("robotsTxt") or {}).get("afterDateFilter") or 0) for s in all_stats)
    total_children_rejected_word = sum(len((s.get("recursiveExpansion") or {}).get("childrenRejectedWord") or []) for s in all_stats)
    total_children_rejected_year = sum(len((s.get("recursiveExpansion") or {}).get("childrenRejectedYear") or []) for s in all_stats)
    total_children_rejected_date = sum(len((s.get("recursiveExpansion") or {}).get("childrenRejectedDate") or []) for s in all_stats)
    total_final_leaves = sum(int((s.get("finalStats") or {}).get("afterDateFilter") or 0) for s in all_stats)
    total_llm_calls = sum(int((s.get("llmDetection") or {}).get("totalLeaves") or 0) for s in all_stats)
    total_llm_success = sum(int((s.get("llmDetection") or {}).get("successful") or 0) for s in all_stats)
    total_llm_failed = sum(int((s.get("llmDetection") or {}).get("failed") or 0) for s in all_stats)
    total_css_triggered = sum(1 for s in all_stats if (s.get("cssFallback") or {}).get("triggered"))
    total_css_success = sum(1 for s in all_stats if (s.get("cssFallback") or {}).get("success"))
    total_css_failed = total_css_triggered - total_css_success
    
    print(f"\nTotal URLs Tested:                    {len(all_stats)}")
    print(f"Total Sitemaps in robots.txt:         {total_sitemaps}")
    print(f"After robots.txt filtering:           {total_after_filters}")
    print()
    print(f"Recursive Expansion Rejections:")
    print(f"  • By Word Filter:                   {total_children_rejected_word}")
    print(f"  • By Year Filter:                   {total_children_rejected_year}")
    print(f"  • By Date Filter:                   {total_children_rejected_date}")
    print()
    print(f"✅ Final Leaf Sitemaps Ready:         {total_final_leaves}")
    print(f"❌ Total Filtered Out:                {total_sitemaps - total_final_leaves}")
    print()
    print(f"🤖 LLM Selector Detection (Sitemap):")
    print(f"  • Total LLM Calls:                  {total_llm_calls}")
    print(f"  • Successful Detections:            {total_llm_success}")
    print(f"  • Failed Detections:                {total_llm_failed}")
    if total_llm_calls > 0:
        success_rate = (total_llm_success / total_llm_calls) * 100
        print(f"  • Success Rate:                     {success_rate:.1f}%")
        estimated_cost = total_llm_calls * 0.01
        print(f"  • Estimated API Cost:               ${estimated_cost:.2f}")
    print()
    print(f"🔄 CSS Selector Fallback (No Sitemap):")
    print(f"  • Triggered:                        {total_css_triggered}")
    print(f"  • Successful:                       {total_css_success}")
    print(f"  • Failed:                           {total_css_failed}")
    if total_css_triggered > 0:
        css_success_rate = (total_css_success / total_css_triggered) * 100
        print(f"  • Success Rate:                     {css_success_rate:.1f}%")
        # Each CSS fallback makes 2 LLM calls (homepage + articles)
        estimated_cost = total_css_triggered * 2 * 0.01
        print(f"  • Estimated API Cost:               ${estimated_cost:.2f}")
    print()
    
    # Efficiency
    if total_sitemaps > 0:
        efficiency = (total_final_leaves / total_sitemaps) * 100
        filtered_pct = ((total_sitemaps - total_final_leaves) / total_sitemaps) * 100
        print(f"📈 Filtering Efficiency:")
        print(f"  • Kept:     {efficiency:.1f}%")
        print(f"  • Filtered: {filtered_pct:.1f}%")
    
    print()
    print("=" * 80)
    print("✅ Test Complete!")
    print("=" * 80)
    
    # Save detailed report to file
    report_file = "selection_extraction_report.json"
    targets_file = "selection_extraction_targets.json"
    report = {
        "testInfo": {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC"),
            "excelFile": "urls.xlsx",
            "totalUrls": len(urls),
            "urlsTested": len(all_stats),
            "recentHoursThreshold": 24
        },
        "summary": {
            "totalSitemapsFound": total_sitemaps,
            "afterRobotsTxtFiltering": total_after_filters,
            "recursiveRejections": {
                "byWordFilter": total_children_rejected_word,
                "byYearFilter": total_children_rejected_year,
                "byDateFilter": total_children_rejected_date
            },
            "finalLeafSitemaps": total_final_leaves,
            "totalFilteredOut": total_sitemaps - total_final_leaves,
            "filteringEfficiency": f"{((total_sitemaps - total_final_leaves) / total_sitemaps * 100):.1f}%" if total_sitemaps > 0 else "N/A",
            "llmDetection": {
                "totalCalls": total_llm_calls,
                "successful": total_llm_success,
                "failed": total_llm_failed,
                "successRate": f"{(total_llm_success / total_llm_calls * 100):.1f}%" if total_llm_calls > 0 else "N/A",
                "estimatedCost": f"${total_llm_calls * 0.01:.2f}"
            },
            "cssFallback": {
                "triggered": total_css_triggered,
                "successful": total_css_success,
                "failed": total_css_failed,
                "successRate": f"{(total_css_success / total_css_triggered * 100):.1f}%" if total_css_triggered > 0 else "N/A",
                "estimatedCost": f"${total_css_triggered * 2 * 0.01:.2f}"
            }
        },
        "detailedResults": []
    }
    
    # Add detailed results for each URL
    for stat in all_stats:
        url_report = {
            "url": stat["url"],
            "robotsTxt": {
                "sitemapsFound": stat["robotsTxt"]["sitemapsTotal"],
                "afterWordFilter": stat["robotsTxt"]["afterWordFilter"],
                "afterYearFilter": stat["robotsTxt"]["afterYearFilter"],
                "afterDateFilter": stat["robotsTxt"]["afterDateFilter"],
                "rejectedSitemaps": stat["robotsTxt"]["rejected"]
            },
            "recursiveExpansion": {
                "childrenFound": stat["recursiveExpansion"]["childrenFound"],
                "rejectedByWord": [{"url": u, "filter": "word"} for u in stat["recursiveExpansion"]["childrenRejectedWord"]],
                "rejectedByYear": [{"url": u, "filter": "year"} for u in stat["recursiveExpansion"]["childrenRejectedYear"]],
                "rejectedByDate": [{"url": u, "filter": "date"} for u in stat["recursiveExpansion"]["childrenRejectedDate"]],
                "leavesFound": stat["recursiveExpansion"]["leavesFound"]
            },
            "finalFiltering": {
                "totalLeaves": stat["finalStats"]["totalLeaves"],
                "afterWordFilter": stat["finalStats"]["afterWordFilter"],
                "afterYearFilter": stat["finalStats"]["afterYearFilter"],
                "afterDateFilter": stat["finalStats"]["afterDateFilter"]
            },
            "llmDetection": {
                "totalLeaves": stat["llmDetection"]["totalLeaves"],
                "successful": stat["llmDetection"]["successful"],
                "failed": stat["llmDetection"]["failed"],
                "detectedSelectors": stat["llmDetection"]["selectors"]
            },
            "cssFallback": stat["cssFallback"]
        }
        report["detailedResults"].append(url_report)
    
    # Save report
    with open(report_file, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    
    print(f"\n💾 Detailed report saved to: {_clickable_path(report_file)}")
    print(f"   This file contains complete filtering statistics and rejected URLs")
    print()

    # Build scraper-friendly targets file
    try:
        targets: List[Dict[str, Any]] = []
        for stat in all_stats:
            src = stat.get("url") or ""
            # Sitemap targets from LLM detection
            llm = stat.get("llmDetection") or {}
            selectors_list = llm.get("selectors") or []
            leafs: List[Dict[str, Any]] = []
            for it in selectors_list:
                try:
                    leaf_url = (it or {}).get("url")
                    det = (it or {}).get("detectedSelectors") or None
                    if leaf_url and det and isinstance(det, dict) and (det.get("fields") or {}):
                        leafs.append({
                            "url": str(leaf_url),
                            "selectors": det,
                        })
                except Exception:
                    continue
            if src and leafs:
                targets.append({
                    "source": str(src),
                    "sourceType": "sitemap",
                    "leafSitemaps": leafs,
                })

            # CSS targets from fallback
            cssf = stat.get("cssFallback") or {}
            if bool(cssf.get("triggered")) and bool(cssf.get("success")):
                csssel = cssf.get("selectors") or {}
                page_url = csssel.get("pageUrl") or src
                sections = csssel.get("sections") or []
                if src and page_url and sections:
                    targets.append({
                        "source": str(src),
                        "sourceType": "css",
                        "pageUrl": str(page_url),
                        "sections": sections,
                    })

        with open(targets_file, 'w', encoding='utf-8') as tf:
            json.dump(targets, tf, indent=2, ensure_ascii=False)
        print(f"💾 Targets for scraper saved to: {_clickable_path(targets_file)}")
        print(f"   Format: [ {{ source, sourceType (sitemap/css), leafSitemaps[] or sections[] }} ]")
    except Exception as _e:
        print(f"⚠️  Failed to write targets file: {type(_e).__name__}: {str(_e)[:120]}")


if __name__ == "__main__":
    main()


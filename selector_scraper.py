import os
import re
import json
import time
import uuid
import random
import hashlib
from datetime import datetime
from urllib.parse import urlparse
from typing import Dict, Any, List, Tuple, Optional

import httpx
from urllib.parse import urlparse as _urlparse
import concurrent.futures as cf
try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore
from playwright.sync_api import sync_playwright
try:
    from dotenv import load_dotenv  # type: ignore
except Exception:  # pragma: no cover
    load_dotenv = None  # type: ignore

try:
    from openai import OpenAI  # Official OpenAI Python SDK (>=1.0)
except Exception:  # pragma: no cover
    OpenAI = None  # type: ignore


ADS_WIDGET_KEYWORDS = re.compile(
    r"taboola|outbrain|trc_|videocube|native-ad|widget|sponsored|ad-|promo",
    re.IGNORECASE,
)


def _ensure_dirs():
    os.makedirs("debug_llm", exist_ok=True)
    os.makedirs("debug_html", exist_ok=True)


def _random_user_agent() -> str:
    chrome_versions = [
        "127.0.6533.72",
        "128.0.6613.84",
        "129.0.6668.90",
    ]
    version = random.choice(chrome_versions)
    platforms = [
        "Windows NT 10.0; Win64; x64",
        "Macintosh; Intel Mac OS X 10_15_7",
        "X11; Linux x86_64",
    ]
    platform = random.choice(platforms)
    return f"Mozilla/5.0 ({platform}) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{version} Safari/537.36"


def _sanitize_html(html: str) -> str:
    # Remove scripts and inline handlers
    html = re.sub(r"<script[\s\S]*?</script>", "", html, flags=re.IGNORECASE)
    html = re.sub(r" on[a-zA-Z]+=\"[^\"]*\"", "", html)
    html = re.sub(r" on[a-zA-Z]+='[^']*'", "", html)
    return html


def _snapshot_dom(page) -> str:
    html = page.evaluate("() => document.documentElement.outerHTML")
    return _sanitize_html(html or "")


def _readiness_loop(page, max_cycles: int = 8, sleep_ms: int = 250) -> Dict[str, Any]:
    last_text_len = 0
    last_links = 0
    for _ in range(max_cycles):
        page.evaluate("() => window.scrollBy(0, Math.floor(window.innerHeight * 0.8))")
        page.wait_for_timeout(sleep_ms)
        metrics = page.evaluate(
            "() => { const t = document.body?.innerText?.length || 0; const l = document.querySelectorAll('a').length; return { textLength: t, linkMatches: l }; }"
        )
        if not isinstance(metrics, dict):
            metrics = {"textLength": 0, "linkMatches": 0}
        if metrics["textLength"] <= last_text_len and metrics["linkMatches"] <= last_links:
            break
        last_text_len = metrics["textLength"]
        last_links = metrics["linkMatches"]
    return {"textLength": last_text_len, "linkMatches": last_links}


def _chunk_html(html: str, chunk_size: int = 120_000) -> List[str]:
    chunks: List[str] = []
    for i in range(0, len(html), chunk_size):
        chunks.append(html[i : i + chunk_size])
    return chunks


def _openai_client(timeout: Optional[float] = None):
    if OpenAI is None:
        raise RuntimeError(
            "OpenAI SDK not installed. Please `pip install openai` and set OPENAI_API_KEY."
        )
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY is not set in environment.")
    base_url = os.environ.get("OPENAI_BASE_URL")
    if base_url:
        return OpenAI(api_key=api_key, base_url=base_url, timeout=timeout)
    return OpenAI(api_key=api_key, timeout=timeout)


def _llm_provider_from_env(model: Optional[str]) -> Tuple[str, str, str, str]:
    """Return (provider, api_key, base_url, model).

    Provider is 'deepseek' if DEEPSEEK_API_KEY is set; otherwise 'openai'.
    """
    ds_key = os.environ.get("DEEPSEEK_API_KEY")
    if ds_key:
        base_url = os.environ.get("DEEPSEEK_BASE_URL", "https://api.deepseek.com")
        mdl = model or os.environ.get("DEEPSEEK_MODEL", "deepseek-chat")
        return ("deepseek", ds_key, base_url, mdl)
    # default to OpenAI
    oa_key = os.environ.get("OPENAI_API_KEY")
    if not oa_key:
        raise RuntimeError("Set DEEPSEEK_API_KEY or OPENAI_API_KEY in environment.")
    base_url = os.environ.get("OPENAI_BASE_URL", "https://api.openai.com")
    mdl = model or os.environ.get("OPENAI_MODEL", "gpt-4o-mini")
    return ("openai", oa_key, base_url, mdl)


def _llm_prompt(batch_index: int, chunk: str) -> str:
    return (
        "Goal: Find news article lists/tiles on this page and RETURN CSS SELECTORS ONLY for fields.\n\n"
        "Return JSON ONLY. Prefer multiple sections if present.\n\n"
        "Output format (either):\n"
        "- Single: { sectionName?, title, link, confidence }\n"
        "- Multi: { sections: [ { sectionName, selectors: { title, link, date?, description?, author?, category?, ticker? }, confidence } ] }\n\n"
        "STRICT RULES:\n"
        "- Return CSS SELECTORS (strings) for elements, NOT example values or text.\n"
        "- Avoid unstable hashed classes (e.g., .emotion-*) and avoid :nth-child.\n"
        "- At least one of title/link must use parent>child (>) for specificity.\n"
        "- Target visible DOM (no meta tags).\n"
        "- Aim for selectors that match many items (>=5).\n"
        "- Focus on editorial/news items; ignore ads/recommendations/widgets/sponsored/native sections if possible.\n\n"
        f"HTML CHUNK ({batch_index + 1}):\n{chunk}"
    )


def _call_llm(prompt: str, model: Optional[str] = None) -> str:
    provider, api_key, base_url, resolved_model = _llm_provider_from_env(model)
    # Read per-call timeout from env (seconds)
    try:
        _timeout_env = os.environ.get("LLM_TIMEOUT_SECONDS")
        llm_timeout: Optional[float] = float(_timeout_env) if _timeout_env else 60.0
    except Exception:
        llm_timeout = 60.0
    if provider == "deepseek":
        # Use OpenAI-compatible chat completions via HTTP
        url = base_url.rstrip("/") + "/v1/chat/completions"
        headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
        payload = {
            "model": resolved_model,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.1,
        }
        try:
            with httpx.Client(timeout=llm_timeout) as client:
                r = client.post(url, headers=headers, json=payload)
                r.raise_for_status()
                data = r.json()
                return (data.get("choices", [{}])[0].get("message", {}).get("content") or "{}").strip()
        except Exception as e:  # pragma: no cover
            raise RuntimeError(f"DeepSeek request failed: {e}")
    # OpenAI branch
    client = _openai_client(timeout=llm_timeout)
    # Prefer chat.completions (widely supported)
    try:
        chat = client.chat.completions.create(
            model=resolved_model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
            # Some SDKs also honor per-request timeout; constructor timeout covers the client
        )
        return chat.choices[0].message.content or "{}"
    except Exception:
        # Fallback to Responses API
        try:
            resp = client.responses.create(
                model=resolved_model,
                input=prompt,
                temperature=0.1,
                # Constructor timeout applies here as well
            )
            parts = []
            for item in resp.output_text.split("\n"):
                parts.append(item)
            return "\n".join(parts)
        except Exception as e:  # pragma: no cover
            raise RuntimeError(f"LLM request failed: {e}")


def _extract_json_from_text(text: str) -> Optional[Any]:
    """Extract a JSON object from text that may include markdown fences or prose."""
    s = text.strip()
    # 1) Try direct JSON
    try:
        return json.loads(s)
    except Exception:
        pass
    # 2) Fenced block ```json ... ``` or ``` ... ```
    fence = re.search(r"```(?:json)?\s*([\s\S]*?)```", s, flags=re.IGNORECASE)
    if fence:
        inner = fence.group(1).strip()
        try:
            return json.loads(inner)
        except Exception:
            pass
    # 3) First "{" to last "}" slice
    lb = s.find("{")
    rb = s.rfind("}")
    if lb != -1 and rb != -1 and rb > lb:
        candidate = s[lb : rb + 1]
        try:
            return json.loads(candidate)
        except Exception:
            return None
    return None


def _parse_candidates(text: str) -> List[Dict[str, Any]]:
    data = _extract_json_from_text(text)
    if data is None:
        return []

    candidates: List[Dict[str, Any]] = []
    if isinstance(data, dict):
        if "sections" in data and isinstance(data["sections"], list):
            for s in data["sections"]:
                if not isinstance(s, dict):
                    continue
                section_name = s.get("sectionName")
                sel = s.get("selectors") or {}
                conf = s.get("confidence")
                if not isinstance(sel, dict):
                    continue
                if not (isinstance(sel.get("title"), str) and isinstance(sel.get("link"), str)):
                    continue
                # Drop null optional fields
                sel = {k: v for k, v in sel.items() if isinstance(v, str)}
                candidates.append({
                    "sectionName": section_name,
                    "selectors": sel,
                    "confidence": conf,
                })
        else:
            # single format
            title = data.get("title")
            link = data.get("link")
            if isinstance(title, str) and isinstance(link, str):
                sel = {"title": title, "link": link}
                for k in ("date", "description", "author", "category", "ticker"):
                    v = data.get(k)
                    if isinstance(v, str):
                        sel[k] = v
                candidates.append({
                    "sectionName": data.get("sectionName"),
                    "selectors": sel,
                    "confidence": data.get("confidence"),
                })
    return candidates


def _signature(selectors: Dict[str, str]) -> str:
    parts = [selectors.get(k, "") for k in ["title", "link", "date", "description", "author", "category", "ticker"]]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


def _looks_like_selector(s: str) -> bool:
    has_syntax = bool(re.search(r"[.#\[\] >:+]", s)) or bool(re.search(r"^(h\d|a|p|li|article|section|div|span)$", s, re.I))
    is_url = s.startswith("http://") or s.startswith("https://") or s.startswith("//")
    too_long = len(s) > 200
    has_nth = re.search(r":nth-(child|of-type)\s*\(", s, re.I)
    has_contains = re.search(r":contains\s*\(", s, re.I)
    has_has = re.search(r":has\s*\(", s, re.I)
    is_meta = s.lower().startswith("meta[")
    return has_syntax and not (is_url or too_long or has_nth or has_contains or has_has or is_meta)


def _acceptance_checks(title_sel: str, link_sel: str) -> Tuple[bool, bool]:
    # strict: both look like selectors, link targets anchor, and parent>child present
    link_targets_anchor = bool(re.search(r"(\b|\s|>|^)a(\b|[.#\[:\s>])", link_sel, re.I))
    has_parent_child = (">" in title_sel) or (">" in link_sel)
    strict_invalid = (not _looks_like_selector(title_sel)) or (not _looks_like_selector(link_sel)) or (not link_targets_anchor) or (not has_parent_child)
    lenient_invalid = (not _looks_like_selector(title_sel)) or (not _looks_like_selector(link_sel))
    return (not strict_invalid, not lenient_invalid)


def discover_selectors(
    url: str,
    model: Optional[str] = None,
    headful: bool = False,
    slowmo_ms: int = 0,
    exclude_ads: bool = False,
) -> Dict[str, Any]:
    # Load .env once
    if load_dotenv:
        try:
            load_dotenv()
        except Exception:
            pass
    _ensure_dirs()
    print(f"[discover] Navigating to {url} (headful={headful}, slowmo={slowmo_ms})")
    parsed = urlparse(url)
    domain = parsed.netloc or parsed.hostname or "unknown"
    ts = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S-%fZ")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=not headful, slow_mo=slowmo_ms or None)
        context = browser.new_context(
            user_agent=_random_user_agent(),
            viewport={"width": random.randint(1200, 1440), "height": random.randint(800, 1000)},
            extra_http_headers={"Referer": f"https://{domain}/"},
        )
        page = context.new_page()
        page.goto(url, wait_until="domcontentloaded", timeout=60_000)
        metrics = _readiness_loop(page)
        html = _snapshot_dom(page)
        browser.close()

    html_path = os.path.join("debug_html", f"{domain}_{ts}_HTML.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[discover] Snapshot saved: {html_path} (chars={len(html)}), metrics={metrics}")

    chunks = _chunk_html(html)
    print(f"[discover] Total batches: {len(chunks)} (chunk size ~120k chars)")
    batches: List[Dict[str, Any]] = []
    all_candidates: List[Dict[str, Any]] = []

    for i, chunk in enumerate(chunks):
        print(f"[discover] Batch {i+1}/{len(chunks)}: sending {len(chunk)} chars to LLM")
        prompt = _llm_prompt(i, chunk)
        raw_text = _call_llm(prompt, model=model)
        raw_path = os.path.join("debug_llm", f"{domain}_{ts}_batch{i+1}_raw.json")
        with open(raw_path, "w", encoding="utf-8") as f:
            f.write(raw_text)
        parsed_candidates = _parse_candidates(raw_text)

        # Split into acceptedStrict vs acceptedLenient heuristically
        accepted_strict: List[Dict[str, Any]] = []
        accepted_lenient: List[Dict[str, Any]] = []
        for c in parsed_candidates:
            sel = c.get("selectors") or {}
            if exclude_ads and ADS_WIDGET_KEYWORDS.search(" ".join(sel.values())):
                continue
            title_sel = str(sel.get("title", "")).strip()
            link_sel = str(sel.get("link", "")).strip()
            if not title_sel or not link_sel:
                continue
            is_strict, is_lenient = _acceptance_checks(title_sel, link_sel)
            item = {"sectionName": c.get("sectionName") or f"Batch {i+1}", "selectors": sel, "confidence": c.get("confidence", 0.7)}
            if is_strict:
                accepted_strict.append(item)
            elif is_lenient:
                accepted_lenient.append(item)
        print(f"[discover] Batch {i+1}: candidates={len(parsed_candidates)}, strict={len(accepted_strict)}, lenient={len(accepted_lenient)}")
        batches.append(
            {
                "batch": i + 1,
                "rawCandidateCount": len(parsed_candidates),
                "rawCandidates": [],  # keep small to avoid bloat; raw saved on disk
                "acceptedStrict": accepted_strict,
                "acceptedLenient": accepted_lenient,
            }
        )
        all_candidates.extend(accepted_strict + accepted_lenient)

    # Dedupe by selector signature
    seen: Dict[str, Dict[str, Any]] = {}
    for c in all_candidates:
        sig = _signature(c["selectors"])
        if sig not in seen:
            seen[sig] = c
        else:
            # keep higher confidence
            if (c.get("confidence") or 0) > (seen[sig].get("confidence") or 0):
                seen[sig] = c

    sections = list(seen.values())
    strict_count = sum(1 for b in batches for _ in b["acceptedStrict"])  # heuristic summary
    lenient_count = sum(1 for b in batches for _ in b["acceptedLenient"])  # heuristic summary

    aggregation = {
        "domain": domain,
        "totalAccepted": len(sections),
        "strictAccepted": strict_count,
        "lenientAccepted": lenient_count,
        "batches": batches,
    }
    agg_path = os.path.join("debug_llm", f"{domain}_{ts}_aggregation.json")
    with open(agg_path, "w", encoding="utf-8") as f:
        json.dump(aggregation, f, ensure_ascii=False, indent=2)
    print(f"[discover] Aggregation saved: {agg_path} (total={len(sections)}, strict={strict_count}, lenient={lenient_count})")

    result = {
        "success": True,
        "domain": domain,
        "engine": "llm",
        "sections": sections,
        "logs": {"html": html_path, "aggregation": agg_path},
        "metrics": metrics,
    }
    return result


def _process_single_url(u: str, model: Optional[str], headful: bool, slowmo: int, exclude_ads: bool) -> Dict[str, Any]:
    """Worker function to process a single URL. Intended for use in ProcessPool."""
    out = discover_selectors(
        url=u,
        model=model,
        headful=headful,
        slowmo_ms=slowmo,
        exclude_ads=exclude_ads,
    )
    out["_input_url"] = u
    return out


def _cli():  # simple CLI
    import argparse

    parser = argparse.ArgumentParser(description="Discover news section selectors via LLM (OpenAI or DeepSeek)")
    parser.add_argument("url", nargs="?", help="Target URL (omit when using --excel)")
    parser.add_argument("--model", default=None, help="Model name (OpenAI or DeepSeek). Default inferred by provider")
    parser.add_argument("--provider", choices=["auto", "openai", "deepseek"], default="auto", help="Force provider; default auto by env (DEEPSEEK_API_KEY preferred)")
    parser.add_argument("--headful", action="store_true")
    parser.add_argument("--slowmo", type=int, default=0)
    parser.add_argument("--exclude-ads", action="store_true")
    parser.add_argument("--output", default="selectors.json", help="Output file (single mode) or combined output (excel mode)")
    parser.add_argument("--excel", help="Path to an Excel .xlsx file with URLs (first column or header 'url')")
    parser.add_argument("--concurrency", type=int, default=1, help="Number of parallel workers for Excel mode (default 1)")
    args = parser.parse_args()

    # If provider is forced to openai or deepseek, validate env quickly
    if args.provider == "openai" and not os.environ.get("OPENAI_API_KEY"):
        raise SystemExit("OPENAI_API_KEY is not set.")
    if args.provider == "deepseek" and not os.environ.get("DEEPSEEK_API_KEY"):
        raise SystemExit("DEEPSEEK_API_KEY is not set.")

    # Batch mode via Excel
    if args.excel:
        if load_workbook is None:
            raise SystemExit("openpyxl is required. Please pip install openpyxl or run pip install -r python_tools/requirements.txt")
        xls_path = args.excel
        if not os.path.isabs(xls_path) and not os.path.exists(xls_path):
            script_dir = os.path.dirname(__file__)
            cand = os.path.join(script_dir, os.path.basename(xls_path))
            if os.path.exists(cand):
                xls_path = cand
        if not os.path.exists(xls_path):
            raise SystemExit(f"Excel file not found: {args.excel}")

        print(f"[batch] Loading Excel: {xls_path}")
        wb = load_workbook(xls_path)
        ws = wb.active
        urls: List[str] = []
        header: List[str] = []
        url_idx = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 and row:
                header = [str(c).strip().lower() if c is not None else "" for c in row]
                # If there's a 'url' header, use that column index; else default to first column
                try:
                    url_idx = header.index("url")
                except ValueError:
                    url_idx = 0
                continue
            # Identify url cell
            # url_idx already determined from header row
            if not row:
                continue
            cell = row[url_idx] if url_idx < len(row) else None
            if not cell:
                continue
            u = str(cell).strip()
            if not u:
                continue
            # basic URL validation
            pr = _urlparse(u)
            if not pr.scheme:
                # assume https
                u = "https://" + u
                pr = _urlparse(u)
            if not pr.netloc:
                continue
            urls.append(u)
        # de-duplicate
        seen = set()
        urls = [u for u in urls if not (u in seen or seen.add(u))]
        print(f"[batch] Total URLs: {len(urls)}")

        results: List[Dict[str, Any]] = []
        entries: List[Dict[str, Any]] = []
        combined: Dict[str, Any] = {"success": True, "mode": "excel", "processed": 0, "results": results, "entries": entries}
        script_dir = os.path.dirname(__file__)
        # Determine effective headful for batch: restrict headful to concurrency=1 to avoid many windows
        effective_headful = args.headful and (args.concurrency <= 1)
        if args.headful and args.concurrency > 1:
            print(f"[batch] Headful disabled for concurrency={args.concurrency}. Set --concurrency 1 to view browser.")

        max_workers = max(1, int(args.concurrency))
        # Run in parallel using process pool
        with cf.ProcessPoolExecutor(max_workers=max_workers) as executor:
            future_to_url = {
                executor.submit(_process_single_url, u, args.model, effective_headful, args.slowmo, args.exclude_ads): u
                for u in urls
            }
            completed = 0
            for future in cf.as_completed(future_to_url):
                u = future_to_url[future]
                try:
                    out = future.result()
                    domain = (out.get("domain") or (_urlparse(u).netloc or "unknown")).lower()
                    ts = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S-%fZ")
                    per_path = os.path.join(script_dir, f"selectors_{domain}_{ts}.json")
                    with open(per_path, "w", encoding="utf-8") as f:
                        json.dump(out, f, ensure_ascii=False, indent=2)
                    results.append({
                        "url": u,
                        "domain": domain,
                        "success": True,
                        "output": per_path,
                        "sections": len(out.get("sections") or []),
                        "logs": out.get("logs") or {},
                    })
                    entries.append({
                        "url": u,
                        "domain": domain,
                        "sections": out.get("sections") or [],
                        "logs": out.get("logs") or {},
                        "engine": out.get("engine") or "llm",
                    })
                except Exception as e:
                    results.append({"url": u, "success": False, "error": str(e)})
                completed += 1
                print(f"[batch] Completed {completed}/{len(urls)}")
        combined["processed"] = len(results)

        # Write combined output
        out_path = args.output if args.output else os.path.join(script_dir, "selectors_combined.json")
        if not os.path.isabs(out_path):
            if os.path.dirname(out_path):
                out_path = os.path.abspath(out_path)
            else:
                out_path = os.path.join(script_dir, out_path)
        out_parent = os.path.dirname(out_path) or script_dir
        os.makedirs(out_parent, exist_ok=True)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(combined, f, ensure_ascii=False, indent=2)
        print(json.dumps({"success": True, "mode": "excel", "output": out_path}, ensure_ascii=False))
        return

    # Single URL mode
    if not args.url:
        raise SystemExit("Provide a URL or use --excel <file.xlsx>.")
    out = discover_selectors(
        url=args.url,
        model=args.model,
        headful=args.headful,
        slowmo_ms=args.slowmo,
        exclude_ads=args.exclude_ads,
    )
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(json.dumps({"success": True, "output": args.output}, ensure_ascii=False))


if __name__ == "__main__":
    _cli()



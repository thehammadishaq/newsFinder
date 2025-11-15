import os
import sys
import json
import time
import threading
import concurrent.futures as cf
from typing import Dict, Any, List, Optional, Iterator, Tuple
from urllib.parse import urlparse
from queue import Queue, Empty
from datetime import datetime, timezone
import csv
from overview_store import init_db as ov_init_db, upsert_overview as ov_upsert, export_csv as ov_export_csv
import httpx
import xml.etree.ElementTree as ET

# Force unbuffered output for real-time logs (MUST be before any other stdout modifications)
os.environ['PYTHONUNBUFFERED'] = '1'
try:
    sys.stdout.reconfigure(encoding='utf-8', line_buffering=True, write_through=True)
    sys.stderr.reconfigure(encoding='utf-8', line_buffering=True, write_through=True)
except Exception:
    pass

import news_scraper_via_selectors as ns
from sitemap_discovery import expand_sitemap_entries_recent  # type: ignore


def _append_jsonl(record: Dict[str, Any], out_path: str) -> None:
    os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)
    with open(out_path, 'a', encoding='utf-8') as f:
        json.dump(record, f, ensure_ascii=False)
        f.write('\n')
        f.flush()
        try:
            os.fsync(f.fileno())
        except Exception:
            pass


def _iter_jsonl(path: str) -> Iterator[Dict[str, Any]]:
    """Yield JSON objects from a JSONL file as they arrive, blocking for new lines."""
    with open(path, 'r', encoding='utf-8') as f:
        f.seek(0, os.SEEK_SET)
        while True:
            pos = f.tell()
            line = f.readline()
            if not line:
                time.sleep(0.5)
                f.seek(pos)
                continue
            line = line.strip()
            if not line:
                continue
            try:
                yield json.loads(line)
            except Exception:
                continue

def _read_jsonl_once(path: str) -> List[Dict[str, Any]]:
    """Read entire JSONL file once and return parsed rows."""
    rows: List[Dict[str, Any]] = []
    if not os.path.exists(path):
        return rows
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except Exception:
                continue
    return rows


# ========================
# Proxy + Fallback Helpers
# ========================
PROXY_SERVER = os.getenv('PROXY_SERVER')
proxy_hosts_promoted: set = set()
proxy_hosts_lock = threading.Lock()
SCRAPE_TIMEOUT: float = 15.0  # default; set from CLI


def _http_probe(url: str, timeout: float = 5.0) -> Optional[Dict[str, Any]]:
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
    }
    try:
        with httpx.Client(timeout=timeout, headers=headers, follow_redirects=True) as client:
            r = client.get(url)
            try:
                status = int(getattr(r, 'status_code', 0) or 0)
            except Exception:
                status = 0
            try:
                raw = r.content[:1024] if getattr(r, 'content', None) is not None else b''
            except Exception:
                raw = b''
            try:
                body_snippet = raw.decode('utf-8', errors='replace').replace('\n', ' ')
            except Exception:
                body_snippet = ''
            headers_out = {}
            try:
                headers_out = dict(getattr(r, 'headers', {}) or {})
            except Exception:
                headers_out = {}
            return {
                'url': url,
                'status_code': status,
                'headers': headers_out,
                'body_snippet': (body_snippet[:512] if isinstance(body_snippet, str) else '')
            }
    except Exception:
        return None


def _classify_probe(probe: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    try:
        if not isinstance(probe, dict):
            return None
        status = int(probe.get('status_code') or 0)
        headers = probe.get('headers') or {}
        body = (probe.get('body_snippet') or '').lower()
        server = str(headers.get('server', '')).lower()

        if 'cloudflare' in server or 'cf-ray' in headers or 'attention required' in body:
            return {'category': 'access_blocked', 'subtype': 'cloudflare', 'retryable': True}
        if status == 429:
            return {'category': 'http_error', 'subtype': '429', 'retryable': True}
        if status in (401, 403):
            return {'category': 'http_error', 'subtype': str(status), 'retryable': False}
        if 300 <= status < 400:
            return {'category': 'http_redirect', 'subtype': str(status), 'retryable': True}
        return None
    except Exception:
        return None


def _browser_fetch_sitemap_text(url: str, timeout: float = 15.0) -> str:
    """
    Fetch sitemap text using Playwright browser.
    This function runs in a thread-safe manner to avoid event loop conflicts.
    """
    def _run_playwright():
        """Inner function that runs Playwright in a separate thread context"""
        try:
            from playwright.sync_api import sync_playwright
        except Exception:
            return ''
        try:
            with sync_playwright() as p:
                proxy_kw = {'server': PROXY_SERVER} if PROXY_SERVER and PROXY_SERVER.strip() else None
                browser = p.chromium.launch(headless=True, proxy=proxy_kw) if proxy_kw else p.chromium.launch(headless=True)
                context = browser.new_context(
                    user_agent=(
                        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                        'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36'
                    ),
                    viewport={'width': 1366, 'height': 864},
                )
                page = context.new_page()
                try:
                    page.goto(url, wait_until='domcontentloaded', timeout=int(max(5.0, float(timeout)) * 1000))
                    xml_text = page.evaluate('( ) => document && (document.body && document.body.innerText) ? document.body.innerText : (document.documentElement ? document.documentElement.innerText : "")') or ''
                finally:
                    try:
                        browser.close()
                    except Exception:
                        pass
                return xml_text or ''
        except Exception:
            return ''
    
    # Run Playwright in a thread executor to avoid event loop conflicts
    # This is necessary when called from async contexts (like FastAPI with uvloop)
    try:
        with cf.ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(_run_playwright)
            return future.result(timeout=timeout + 5.0)  # Add buffer for thread overhead
    except Exception:
        # Fallback: try direct execution if thread executor fails
        return _run_playwright()


def _parse_sitemap_entries_from_text(xml_text: str) -> List[Dict[str, Any]]:
    if not xml_text:
        return []
    try:
        root = ET.fromstring(xml_text)
    except Exception:
        return []
    tag = (root.tag or '').lower()
    out: List[Dict[str, Any]] = []
    if tag.endswith('urlset'):
        for u in root.findall('.//{*}url'):
            try:
                loc = u.findtext('{*}loc') or u.findtext('loc')
            except Exception:
                loc = None
            if loc:
                out.append({'url': str(loc).strip()})
    return out


def _promote_host(host: str) -> None:
    if not host:
        return
    try:
        with proxy_hosts_lock:
            proxy_hosts_promoted.add(host)
        print(f"[proxy] Promoted host: {host} (proxy-first)")
    except Exception:
        pass

def _normalize_targets(stream_obj: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Convert one stream result object into zero or more scraping targets."""
    targets: List[Dict[str, Any]] = []
    result = stream_obj.get('result') or {}
    site_url = result.get('url') or result.get('domain') or ''
    if not site_url:
        return targets

    # Sitemap targets from llmDetection.selectors
    llm = result.get('llmDetection') or {}
    sel_list = llm.get('selectors') or []
    for item in sel_list:
        leaf_url = item.get('url')
        detected = item.get('detectedSelectors') or {}
        fields = detected.get('fields') or {}
        item_tag = detected.get('item') or 'url'
        if leaf_url and fields:
            targets.append({
                'type': 'sitemap',
                'site': site_url,
                'sitemapUrl': leaf_url,
                'itemTag': item_tag,
                'fields': fields,
                'detectedBy': detected.get('detectionMethod', 'llm')
            })

    # CSS targets from cssFallback.selectors.sections
    cssf = result.get('cssFallback') or {}
    if cssf.get('triggered') and cssf.get('success'):
        csssel = cssf.get('selectors') or {}
        page_url = csssel.get('pageUrl') or site_url
        sections = csssel.get('sections') or []
        if page_url and sections:
            targets.append({
                'type': 'css',
                'site': site_url,
                'pageUrl': page_url,
                'sections': sections,
                'detectedBy': csssel.get('detectionMethod', 'css_fallback')
            })

    return targets


# ========================
# CSV upsert helper (shared contract)
# ========================
_CSV_PATH = os.path.join(os.path.dirname(__file__) or '.', 'pipelines_overview.csv')
_CSV_HEADER = [
    'Domain (sources)',
    'Selector Discovery Attempted',
    'Selector Discovery Not Attempted Reason',
    'Selector Discovery Attempt Error',
    'Selector Discovery Attempt Error Response',
    'Sitemap Processing Status',
    'Sitemap Processing Error Details',
    'leaf Sitemap URLs Discovered',
    'CSS Fallback Status',
    'CSS Fallback error Details',
    'Which Path Used for Final Extraction',
    'Total Time (sec) in scraping',
    'Raw Articles scraped',
    'Zero Raw Articles Reason',
    'Cleaning Status',
    'Cleaned Articles (Final)',
    'Duplicates Removed',
    'Missing Dates Removed',
    'Out of Range/Old Date Removed',
    'Overall pipelines Status',
    'Overall pipelines Error Details',
    'Overall pipelines Explanation',
    'Leaf Sitemap URLs',
]


def _read_csv_map(path: str) -> Dict[str, Dict[str, str]]:
    rows: Dict[str, Dict[str, str]] = {}
    if not os.path.exists(path):
        return rows
    try:
        with open(path, 'r', encoding='utf-8', newline='') as f:
            r = csv.DictReader(f)
            for row in r:
                d = (row.get(_CSV_HEADER[0]) or '').strip()
                if d:
                    rows[d] = {k: (row.get(k) or '') for k in _CSV_HEADER}
    except Exception:
        pass
    return rows


def _default_row(domain: str) -> Dict[str, str]:
    return {
        'Domain (sources)': domain,
        'Selector Discovery Attempted': 'No',
        'Selector Discovery Not Attempted Reason': '',
        'Selector Discovery Attempt Error': '',
        'Sitemap Processing Status': 'Not Attempted',
        'Sitemap Processing Error Details': '',
        'leaf Sitemap URLs Discovered': '0',
        'CSS Fallback Status': 'Not Attempted',
        'CSS Fallback error Details': '',
        'Which Path Used for Final Extraction': 'Neither',
        'Total Time (sec) in scraping': '0',
        'Raw Articles scraped': '0',
        'Cleaning Status': 'Not Attempted',
        'Cleaned Articles (Final)': '0',
        'Duplicates Removed': '0',
        'Missing Dates Removed': '0',
        'Out of Range/Old Date Removed': '0',
        'Overall pipelines Status': 'Pending',
        'Leaf Sitemap URLs': '',
    }


def _write_csv_map(path: str, rows: Dict[str, Dict[str, str]]) -> None:
    tmp = path + '.tmp'
    try:
        with open(tmp, 'w', encoding='utf-8', newline='') as f:
            w = csv.DictWriter(f, fieldnames=_CSV_HEADER)
            w.writeheader()
            for domain in sorted(rows.keys()):
                row = rows[domain]
                for h in _CSV_HEADER:
                    row.setdefault(h, '')
                w.writerow(row)
        try:
            os.replace(tmp, path)
        except Exception:
            try:
                if os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass
            try:
                os.rename(tmp, path)
            except Exception:
                pass
    except Exception:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass


def _upsert_csv_row(domain: str, updates: Dict[str, str]) -> None:
    for _ in range(3):
        try:
            rows = _read_csv_map(_CSV_PATH)
            row = rows.get(domain) or _default_row(domain)
            row[_CSV_HEADER[0]] = domain
            for k, v in (updates or {}).items():
                if k in _CSV_HEADER and v is not None:
                    if k == 'Overall pipelines Error Details':
                        # Merge rather than overwrite
                        prev = row.get(k) or ''
                        try:
                            parts = [s.strip() for s in (prev or '').split(' | ') if s.strip()]
                            if str(v).strip() and str(v).strip() not in parts:
                                parts.append(str(v).strip())
                            row[k] = ' | '.join(parts)[:300]
                        except Exception:
                            row[k] = str(v)[:300]
                    elif k == 'Overall pipelines Explanation':
                        prev = row.get(k) or ''
                        try:
                            parts = [s.strip() for s in (prev or '').split(' | ') if s.strip()]
                            sent = str(v).strip()
                            if sent and sent not in parts:
                                parts.append(sent)
                            row[k] = ' | '.join(parts)[:300]
                        except Exception:
                            row[k] = str(v)[:300]
                    else:
                        row[k] = str(v)
            rows[domain] = row
            _write_csv_map(_CSV_PATH, rows)
            return
        except PermissionError:
            import time as _t
            _t.sleep(0.2)
        except Exception:
            return


def _scrape_sitemap_target(t: Dict[str, Any]) -> List[Dict[str, Any]]:
    leaf = t.get('sitemapUrl')
    fields = t.get('fields') or None
    host = _domain_of(leaf or '')

    # Proxy-first for promoted hosts
    if host and host in proxy_hosts_promoted:
        try:
            print(f"[proxy] Proxy-first sitemap fetch: {leaf}")
            xml_text = _browser_fetch_sitemap_text(leaf, timeout=SCRAPE_TIMEOUT)
            ents = _parse_sitemap_entries_from_text(xml_text)
            if ents:
                return ents
        except Exception:
            pass
        # Optional: try normal as a backup
        try:
            ents = expand_sitemap_entries_recent(leaf, recent_hours=0, field_selectors=fields)
        except Exception:
            ents = []
        out: List[Dict[str, Any]] = []
        for e in ents:
            if e.get('url'):
                out.append(e)
        return out

    # Normal-first for non-promoted hosts
    try:
        ents = expand_sitemap_entries_recent(leaf, recent_hours=0, timeout=SCRAPE_TIMEOUT, field_selectors=fields)
    except Exception:
        ents = []
    out: List[Dict[str, Any]] = []
    for e in ents:
        if e.get('url'):
            out.append(e)
    if out:
        return out

    # If empty, probe and attempt one proxy-backed browser retry on likely blocks
    probe = None
    try:
        probe = _http_probe(leaf, timeout=min(5.0, max(1.0, SCRAPE_TIMEOUT/3.0)))
    except Exception:
        probe = None
    classified = _classify_probe(probe)
    if isinstance(classified, dict) and (classified.get('category') in ('access_blocked', 'http_error', 'http_redirect')):
        try:
            print(f"[proxy] Fallback sitemap via browser+proxy: {leaf} ({classified.get('subtype')})")
            xml_text = _browser_fetch_sitemap_text(leaf, timeout=SCRAPE_TIMEOUT)
            ents2 = _parse_sitemap_entries_from_text(xml_text)
            if ents2:
                try:
                    print(f"[proxy] Fallback SUCCESS (sitemap): {leaf} -> entries={len(ents2)}")
                except Exception:
                    pass
                _promote_host(host)
                return ents2
            else:
                try:
                    print(f"[proxy] Fallback FAILED (sitemap): {leaf} -> no entries")
                except Exception:
                    pass
        except Exception:
            pass
    return out


def _scrape_css_target(t: Dict[str, Any], headful: bool = False, slowmo_ms: int = 0, max_items: int = 10000) -> List[Dict[str, Any]]:
    url = t.get('pageUrl')
    sections = t.get('sections') or []
    host = _domain_of(url or '')

    def _run_extract_with_possible_proxy() -> List[Dict[str, Any]]:
        if PROXY_SERVER:
            try:
                os.environ.setdefault('PROXY_SERVER', PROXY_SERVER)
            except Exception:
                pass
        res = ns.extract_via_selectors(url=url, sections=sections, headful=headful, slowmo_ms=slowmo_ms, max_items=max_items)
        arts = res.get('articles') or []
        return arts if isinstance(arts, list) else []

    # Proxy-first for promoted hosts
    if host and host in proxy_hosts_promoted:
        try:
            print(f"[proxy] Proxy-first CSS extract: {url}")
            arts = _run_extract_with_possible_proxy()
            if arts:
                return arts
        except Exception:
            pass
        # Optional: try without proxy once
        try:
            res_np = ns.extract_via_selectors(url=url, sections=sections, headful=headful, slowmo_ms=slowmo_ms, max_items=max_items)
            arts_np = res_np.get('articles') or []
            return arts_np if isinstance(arts_np, list) else []
        except Exception:
            return []

    # Normal-first for non-promoted hosts
    try:
        res = ns.extract_via_selectors(url=url, sections=sections, headful=headful, slowmo_ms=slowmo_ms, max_items=max_items)
        arts = res.get('articles') or []
        arts = arts if isinstance(arts, list) else []
    except Exception:
        arts = []
    if arts:
        return arts

    # If empty/failed, probe and fallback with proxy once if blocked
    probe = None
    try:
        probe = _http_probe(url, timeout=min(5.0, max(1.0, SCRAPE_TIMEOUT/3.0)))
    except Exception:
        probe = None
    classified = _classify_probe(probe)
    if isinstance(classified, dict) and (classified.get('category') in ('access_blocked', 'http_error', 'http_redirect')):
        try:
            print(f"[proxy] Fallback CSS via proxy: {url} ({classified.get('subtype')})")
            arts2 = _run_extract_with_possible_proxy()
            if arts2:
                try:
                    print(f"[proxy] Fallback SUCCESS (css): {url} -> items={len(arts2)}")
                except Exception:
                    pass
                _promote_host(host)
                return arts2
            else:
                try:
                    print(f"[proxy] Fallback FAILED (css): {url} -> no items")
                except Exception:
                    pass
        except Exception:
            pass
    return arts


def _domain_of(url: str) -> str:
    try:
        return urlparse(url).netloc.lower()
    except Exception:
        return ''


class Writer:
    def __init__(self, out_path: str, queue_size: int = 1000, batch_size: int = 50, flush_interval_sec: float = 0.5) -> None:
        self.out_path = out_path
        self.q: Queue = Queue(maxsize=max(10, queue_size))
        self._stop = threading.Event()
        self._thr = threading.Thread(target=self._run, daemon=True)
        self._f = None  # type: ignore
        self._batch_size = max(1, int(batch_size))
        self._flush_interval_sec = max(0.05, float(flush_interval_sec))

    def start(self) -> None:
        # Truncate existing file and open once for batched appends
        try:
            os.makedirs(os.path.dirname(self.out_path) or '.', exist_ok=True)
            self._f = open(self.out_path, 'w', encoding='utf-8')
        except Exception:
            self._f = None
        self._thr.start()

    def _run(self) -> None:
        buffer: List[Dict[str, Any]] = []
        last_flush = time.perf_counter()
        while not self._stop.is_set():
            try:
                rec = self.q.get(timeout=0.2)
                buffer.append(rec)
                self.q.task_done()
            except Empty:
                pass

            now = time.perf_counter()
            should_flush = (len(buffer) >= self._batch_size) or ((buffer) and (now - last_flush >= self._flush_interval_sec))
            if not should_flush:
                continue

            try:
                if self._f is None:
                    # Fallback to per-record append if file couldn't be opened
                    for r in buffer:
                        _append_jsonl(r, self.out_path)
                else:
                    for r in buffer:
                        try:
                            self._f.write(json.dumps(r, ensure_ascii=False))
                            self._f.write('\n')
                        except Exception:
                            # Fallback to safe path for this record
                            _append_jsonl(r, self.out_path)
                    try:
                        self._f.flush()
                        # Optional: fsync per batch for durability
                        try:
                            os.fsync(self._f.fileno())
                        except Exception:
                            pass
                    except Exception:
                        pass
            finally:
                buffer.clear()
                last_flush = now

    def submit(self, record: Dict[str, Any]) -> None:
        self.q.put(record)

    def close(self) -> None:
        # First, wait for all queued records to be written
        self.q.join()
        # Then signal the thread to exit and wait briefly
        self._stop.set()
        try:
            self._thr.join(timeout=5)
        except Exception:
            pass
        # Flush any remaining buffered writes and close file
        try:
            if getattr(self, '_f', None) is not None:
                try:
                    self._f.flush()
                except Exception:
                    pass
                try:
                    os.fsync(self._f.fileno())
                except Exception:
                    pass
                try:
                    self._f.close()
                except Exception:
                    pass
        except Exception:
            pass


class StatsCollector:
    def __init__(self, sites_log_path: str, summary_path: str, append: bool = False) -> None:
        self.sites_log_path = sites_log_path
        self.summary_path = summary_path
        self.append = append
        self._lock = threading.Lock()
        self._global_start: Optional[float] = None
        self._global_end: Optional[float] = None
        # Aggregates
        self.total_sites_processed: int = 0
        self.total_articles: int = 0
        self.articles_by_source: Dict[str, int] = {"sitemap": 0, "css": 0}
        self.sites_by_approach: Dict[str, int] = {"sitemapOnly": 0, "cssOnly": 0, "both": 0, "none": 0}
        self._last_snapshot_sites_written: int = 0

    def start_global(self) -> None:
        # Prepare files
        os.makedirs(os.path.dirname(self.sites_log_path) or '.', exist_ok=True)
        os.makedirs(os.path.dirname(self.summary_path) or '.', exist_ok=True)
        if not self.append:
            try:
                with open(self.sites_log_path, 'w', encoding='utf-8') as _f:
                    pass
            except Exception:
                pass
            try:
                with open(self.summary_path, 'w', encoding='utf-8') as _f:
                    pass
            except Exception:
                pass
        self._global_start = time.perf_counter()

    def end_global(self) -> None:
        self._global_end = time.perf_counter()

    def record_site(self, *, site: str, started_at_iso: str, ended_at_iso: str, duration_sec: float, items_by_source: Dict[str, int], approaches_used: List[str]) -> None:
        with self._lock:
            self.total_sites_processed += 1
            site_total = int(items_by_source.get('sitemap', 0)) + int(items_by_source.get('css', 0))
            self.total_articles += site_total
            self.articles_by_source['sitemap'] += int(items_by_source.get('sitemap', 0))
            self.articles_by_source['css'] += int(items_by_source.get('css', 0))
            # Approach categorization
            used = set(approaches_used)
            if 'sitemap' in used and 'css' in used:
                self.sites_by_approach['both'] += 1
            elif 'sitemap' in used:
                self.sites_by_approach['sitemapOnly'] += 1
            elif 'css' in used:
                self.sites_by_approach['cssOnly'] += 1
            else:
                self.sites_by_approach['none'] += 1
            # Also write per-site log record
            rec = {
                'site': site,
                'startedAt': started_at_iso,
                'endedAt': ended_at_iso,
                'durationSec': round(float(duration_sec), 3),
                'itemsTotal': site_total,
                'itemsBySource': {
                    'sitemap': int(items_by_source.get('sitemap', 0)),
                    'css': int(items_by_source.get('css', 0))
                },
                'approachesUsed': list(used)
            }
        # Append outside lock using existing helper
        try:
            _append_jsonl(rec, self.sites_log_path)
        except Exception:
            pass

    def write_summary(self) -> None:
        try:
            total_duration = None
            if self._global_start is not None and self._global_end is not None:
                total_duration = self._global_end - self._global_start
            summary = {
                'totalDurationSec': round(float(total_duration or 0.0), 3),
                'totalSitesProcessed': int(self.total_sites_processed),
                'totalArticles': int(self.total_articles),
                'articlesBySource': {
                    'sitemap': int(self.articles_by_source.get('sitemap', 0)),
                    'css': int(self.articles_by_source.get('css', 0))
                },
                'sitesByApproach': {
                    'sitemapOnly': int(self.sites_by_approach.get('sitemapOnly', 0)),
                    'cssOnly': int(self.sites_by_approach.get('cssOnly', 0)),
                    'both': int(self.sites_by_approach.get('both', 0)),
                    'none': int(self.sites_by_approach.get('none', 0))
                }
            }
            with open(self.summary_path, 'w', encoding='utf-8') as f:
                json.dump(summary, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def write_summary_snapshot(self) -> None:
        try:
            with self._lock:
                now_perf = time.perf_counter()
                total_duration = (now_perf - self._global_start) if self._global_start is not None else 0.0
                summary = {
                    'totalDurationSec': round(float(total_duration or 0.0), 3),
                    'totalSitesProcessed': int(self.total_sites_processed),
                    'totalArticles': int(self.total_articles),
                    'articlesBySource': {
                        'sitemap': int(self.articles_by_source.get('sitemap', 0)),
                        'css': int(self.articles_by_source.get('css', 0))
                    },
                    'sitesByApproach': {
                        'sitemapOnly': int(self.sites_by_approach.get('sitemapOnly', 0)),
                        'cssOnly': int(self.sites_by_approach.get('cssOnly', 0)),
                        'both': int(self.sites_by_approach.get('both', 0)),
                        'none': int(self.sites_by_approach.get('none', 0))
                    }
                }
            with open(self.summary_path, 'w', encoding='utf-8') as f:
                json.dump(summary, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def maybe_write_summary_snapshot(self, interval_sites: int) -> None:
        if interval_sites <= 0:
            return
        try:
            with self._lock:
                should_write = (self.total_sites_processed >= self._last_snapshot_sites_written + interval_sites)
                if not should_write:
                    return
                self._last_snapshot_sites_written = self.total_sites_processed
        except Exception:
            return
        self.write_summary_snapshot()


def _cli():
    import argparse
    p = argparse.ArgumentParser(description='Stream scraping pipeline: consume selection_extraction_report_stream.jsonl and extract data')
    p.add_argument('--stream', default=os.path.join(os.path.dirname(__file__), 'selection_extraction_report_stream.jsonl'))
    p.add_argument('--output', default=os.path.join(os.path.dirname(__file__), 'stream_scraped_articles.jsonl'))
    p.add_argument('--targets-json', type=str, default=None, help='Optional: JSON file exported by selection_extraction_pipeline (selection_extraction_targets.json)')
    p.add_argument('--mode', choices=['auto', 'sitemap', 'css', 'both'], default='auto')
    # Concurrency knobs
    p.add_argument('--site-concurrency', type=int, default=1)
    p.add_argument('--target-concurrency', type=int, default=6)
    p.add_argument('--sitemap-concurrency', type=int, default=12)
    p.add_argument('--css-concurrency', type=int, default=1)
    p.add_argument('--http-concurrency', type=int, default=24)
    p.add_argument('--per-domain-cap', type=int, default=1)
    p.add_argument('--queue-size', type=int, default=100)
    p.add_argument('--once', action='store_true', default=True, help='Process current file contents then exit')
    # Auto-tuning
    p.add_argument('--auto-tune', dest='auto_tune', action='store_true', default=True, help='Auto-pick concurrency based on site count (default on)')
    p.add_argument('--no-auto-tune', dest='auto_tune', action='store_false')
    # CSS controls
    p.add_argument('--headful', action='store_true')
    p.add_argument('--slowmo', type=int, default=0)
    p.add_argument('--max', dest='max_items', type=int, default=500)
    p.add_argument('--sitemap-max-urls', dest='sitemap_max_urls', type=int, default=0, help='Max items to take per leaf sitemap (0 = no cap, default)')
    # Logging controls
    p.add_argument('--log-sites', default=os.path.join(os.path.dirname(__file__), 'stream_scrape_sites_log.jsonl'))
    p.add_argument('--log-summary', default=os.path.join(os.path.dirname(__file__), 'stream_scrape_summary.json'))
    p.add_argument('--log-append', action='store_true')
    p.add_argument('--summary-interval-sites', type=int, default=1, help='Write summary JSON every N completed sites (1 = after each site, 0 = only on exit)')
    # Timeouts
    p.add_argument('--timeout', type=float, default=15.0, help='Per-request timeout seconds for sitemap HTTP and browser retries (default 15)')
    # Optional site filter via XLSX
    p.add_argument('--sites-xlsx', type=str, default=None, help='If provided, only scrape sites listed in this XLSX')
    p.add_argument('--sites-sheet', type=str, default=None, help='Sheet name or 0-based index (as string). Default: first sheet')
    p.add_argument('--sites-column', type=str, default=None, help='Column name to read (auto-detect url/site/domain if omitted)')
    args = p.parse_args()
    # Apply timeout globally for sitemap HTTP and browser retries
    try:
        global SCRAPE_TIMEOUT
        SCRAPE_TIMEOUT = float(args.timeout) if args.timeout is not None else 15.0
        if SCRAPE_TIMEOUT <= 0:
            SCRAPE_TIMEOUT = 15.0
        # Optional clamp
        SCRAPE_TIMEOUT = max(5.0, min(SCRAPE_TIMEOUT, 120.0))
    except Exception:
        SCRAPE_TIMEOUT = 15.0

    print(f"[stream] Watching: {args.stream}")
    print(f"[stream] Output JSONL: {args.output}")
    print(f"[stream] Mode={args.mode}")
    # Optional XLSX site filter
    allowed_urls: Optional[set] = None
    allowed_domains: Optional[set] = None
    if args.sites_xlsx:
        try:
            from openpyxl import load_workbook  # type: ignore
            wb = load_workbook(filename=args.sites_xlsx, read_only=True, data_only=True)
            ws = None
            if args.sites_sheet is None:
                ws = wb.worksheets[0]
            else:
                # Try by name first, else index
                if args.sites_sheet in wb.sheetnames:
                    ws = wb[args.sites_sheet]
                else:
                    try:
                        idx = int(args.sites_sheet)
                        ws = wb.worksheets[idx]
                    except Exception:
                        ws = wb.worksheets[0]
            # Read header
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            col_idx = 0
            if isinstance(header, (list, tuple)):
                header_lower = [str(h).strip().lower() if h is not None else '' for h in header]
                if args.sites_column:
                    try:
                        col_idx = header_lower.index(str(args.sites_column).strip().lower())
                    except Exception:
                        col_idx = 0
                else:
                    for cand in ('url', 'site', 'domain'):
                        if cand in header_lower:
                            col_idx = header_lower.index(cand)
                            break
            # Collect
            allowed_urls = set()
            allowed_domains = set()
            for row in rows_iter:
                try:
                    val = row[col_idx]
                except Exception:
                    val = None
                if not val:
                    continue
                s = str(val).strip()
                if not s:
                    continue
                # If looks like URL with scheme, keep as URL; always also add domain
                try:
                    parsed = urlparse(s)
                    if parsed.scheme and parsed.netloc:
                        allowed_urls.add(s)
                        allowed_domains.add(parsed.netloc.lower())
                    else:
                        # treat as domain/host
                        allowed_domains.add(s.lower())
                except Exception:
                    allowed_domains.add(s.lower())
            print(f"[stream] Site filter active: urls={len(allowed_urls or [])} domains={len(allowed_domains or [])}")
        except Exception as e:
            print(f"[stream] Failed to load sites XLSX: {e}")
            allowed_urls, allowed_domains = None, None

    def _is_allowed(site_url: str) -> bool:
        if not args.sites_xlsx:
            return True
        if not site_url:
            return False
        try:
            d = urlparse(site_url).netloc.lower()
        except Exception:
            d = ''
        if allowed_urls and site_url in allowed_urls:
            return True
        if allowed_domains and d in allowed_domains:
            return True
        return False

    # Estimate site count (existing lines only, respecting filter if present)
    est_site_count = 0
    try:
        rows_now = _read_jsonl_once(args.stream)
        seen_sites = set()
        for r in rows_now:
            s = ((r.get('result') or {}).get('url') or '').strip()
            if s and _is_allowed(s):
                seen_sites.add(s)
        est_site_count = len(seen_sites)
    except Exception:
        est_site_count = 0

    # Auto-tune only if enabled and flags are still at defaults
    def _at_default():
        return (
            args.site_concurrency == 1 and
            args.target_concurrency == 6 and
            args.sitemap_concurrency == 12 and
            args.css_concurrency == 1 and
            args.http_concurrency == 24 and
            args.per_domain_cap == 1
        )

    if args.auto_tune and _at_default():
        sc, tc, smc, cc, hc, pdc = 3, 6, 12, 1, 24, 1  # base for small
        if est_site_count <= 0:
            sc, tc, smc, cc, hc, pdc = 3, 6, 12, 1, 24, 1
        elif est_site_count <= 10:
            sc, tc, smc, cc, hc, pdc = 3, 6, 12, 1, 24, 1
        elif est_site_count <= 100:
            sc, tc, smc, cc, hc, pdc = 4, 6, 12, 2, 24, 1
        elif est_site_count <= 500:
            sc, tc, smc, cc, hc, pdc = 6, 8, 16, 2, 32, 1
        else:
            sc, tc, smc, cc, hc, pdc = 8, 8, 24, 2, 32, 1
        args.site_concurrency = sc
        args.target_concurrency = tc
        args.sitemap_concurrency = smc
        args.css_concurrency = cc
        args.http_concurrency = hc
        args.per_domain_cap = pdc
        print(f"[stream] Auto-tuned for ~{est_site_count} site(s)")

    print(f"[stream] Concurrency: sites={args.site_concurrency} target={args.target_concurrency} sitemap={args.sitemap_concurrency} css={args.css_concurrency} http={args.http_concurrency} perDomain={args.per_domain_cap}")

    processed_sites = set()
    enqueued_sites = set()

    # Global semaphores
    http_sem = threading.Semaphore(max(1, int(args.http_concurrency)))
    css_sem = threading.Semaphore(max(1, int(args.css_concurrency)))
    sitemap_sem = threading.Semaphore(max(1, int(args.sitemap_concurrency)))
    domain_sems: Dict[str, threading.Semaphore] = {}

    def _get_domain_sem(domain: str) -> threading.Semaphore:
        if domain not in domain_sems:
            domain_sems[domain] = threading.Semaphore(max(1, int(args.per_domain_cap)))
        return domain_sems[domain]

    # Stats collector
    collector = StatsCollector(args.log_sites, args.log_summary, append=bool(args.log_append))
    collector.start_global()

    # Writer
    writer = Writer(args.output, queue_size=max(10, int(args.queue_size)))
    writer.start()
    # Write initial empty snapshot so file is never empty
    collector.write_summary_snapshot()

    # Site queue and reader
    site_q: Queue = Queue(maxsize=max(10, int(args.queue_size)))
    stop_reader = threading.Event()

    def reader_once():
        for row in _read_jsonl_once(args.stream):
            site = ((row.get('result') or {}).get('url') or '').strip()
            if not site or site in processed_sites or site in enqueued_sites:
                continue
            if _is_allowed(site):
                site_q.put(row)
                enqueued_sites.add(site)
        stop_reader.set()

    def reader_tail():
        for row in _iter_jsonl(args.stream):
            site = ((row.get('result') or {}).get('url') or '').strip()
            if not site or site in processed_sites or site in enqueued_sites:
                continue
            if _is_allowed(site):
                site_q.put(row)
                enqueued_sites.add(site)

    def reader_once_targets():
        # Read once from targets JSON (if provided)
        try:
            with open(args.targets_json, 'r', encoding='utf-8') as f:
                arr = json.load(f)
        except Exception:
            arr = []
        # Group by source
        by_source: Dict[str, Dict[str, Any]] = {}
        for obj in arr if isinstance(arr, list) else []:
            try:
                src = (obj or {}).get('source') or ''
                st = (obj or {}).get('sourceType') or ''
                if not src or st not in ('sitemap', 'css'):
                    continue
                cur = by_source.get(src) or {'result': {'url': src, 'llmDetection': {'selectors': []}, 'cssFallback': {}}}
                if st == 'sitemap':
                    for leaf in (obj.get('leafSitemaps') or []):
                        lu = (leaf or {}).get('url')
                        sel = (leaf or {}).get('selectors') or {}
                        if lu and (sel.get('fields') or {}):
                            cur['result'].setdefault('llmDetection', {}).setdefault('selectors', []).append({'url': lu, 'detectedSelectors': sel})
                else:
                    sections = obj.get('sections') or []
                    page_url = obj.get('pageUrl') or src
                    if sections:
                        cur['result']['cssFallback'] = {'triggered': True, 'success': True, 'selectors': {'pageUrl': page_url, 'sections': sections}}
                by_source[src] = cur
            except Exception:
                continue
        for _, row in by_source.items():
            try:
                site = ((row.get('result') or {}).get('url') or '').strip()
            except Exception:
                site = ''
            if not site or site in processed_sites or site in enqueued_sites:
                continue
            site_q.put(row)
            enqueued_sites.add(site)
        stop_reader.set()

    reader_thr = threading.Thread(target=(reader_once_targets if args.targets_json else (reader_once if args.once else reader_tail)), daemon=True)
    reader_thr.start()

    def process_site(row: Dict[str, Any]) -> Tuple[str, int]:
        site = ((row.get('result') or {}).get('url') or '').strip()
        targets = _normalize_targets(row)
        if not targets:
            print(f"[site] No targets for: {site}")
            # Record an empty site with timing
            started_iso = datetime.now(timezone.utc).isoformat()
            start_perf = time.perf_counter()
            end_perf = time.perf_counter()
            ended_iso = datetime.now(timezone.utc).isoformat()
            collector.record_site(site=site, started_at_iso=started_iso, ended_at_iso=ended_iso, duration_sec=(end_perf - start_perf), items_by_source={"sitemap": 0, "css": 0}, approaches_used=[])
            return site, 0

        # Prioritize sitemap if mode=auto; include CSS if mode=both or css-only
        def _submit_targets(ex: cf.ThreadPoolExecutor) -> Tuple[List[cf.Future], Dict[cf.Future, str]]:
            futures: List[cf.Future] = []
            fut_type: Dict[cf.Future, str] = {}
            for t in targets:
                t_type = t.get('type')
                use_mode = args.mode
                if use_mode == 'auto':
                    # Only run target's natural type; both will run if both exist
                    use_mode = t_type
                if use_mode not in ('sitemap', 'css', 'both'):
                    use_mode = t_type

                if use_mode in ('sitemap', 'both') and t_type == 'sitemap':
                    domain = _domain_of(t.get('sitemapUrl') or '')
                    dom_sem = _get_domain_sem(domain)
                    def _run_sm(tt=t, dsem=dom_sem):
                        # Acquire order: sitemap -> http -> domain
                        with sitemap_sem:
                            with http_sem:
                                with dsem:
                                    return _scrape_sitemap_target(tt)
                    f = ex.submit(_run_sm)
                    futures.append(f)
                    fut_type[f] = 'sitemap'

                if use_mode in ('css', 'both') and t_type == 'css':
                    domain = _domain_of(t.get('pageUrl') or '')
                    dom_sem = _get_domain_sem(domain)
                    effective_headful = args.headful and (args.css_concurrency == 1)
                    def _run_css(tt=t, dsem=dom_sem):
                        with css_sem:
                            with http_sem:
                                with dsem:
                                    return _scrape_css_target(tt, headful=effective_headful, slowmo_ms=args.slowmo, max_items=args.max_items)
                    f = ex.submit(_run_css)
                    futures.append(f)
                    fut_type[f] = 'css'
            return futures, fut_type

        total_items = 0
        started_iso = datetime.now(timezone.utc).isoformat()
        start_perf = time.perf_counter()
        items_by_source: Dict[str, int] = {"sitemap": 0, "css": 0}
        approaches_used: List[str] = []
        had_sitemap = any((t.get('type') == 'sitemap') for t in targets)
        had_css = any((t.get('type') == 'css') for t in targets)
        exception_types: set = set()
        with cf.ThreadPoolExecutor(max_workers=max(1, int(args.target_concurrency))) as ex:
            futures, fut_type = _submit_targets(ex)
            approaches_used = sorted(list(set(fut_type.values())))
            for fut in cf.as_completed(futures):
                try:
                    items = fut.result() or []
                except Exception as e:
                    try:
                        exception_types.add(type(e).__name__)
                    except Exception:
                        pass
                    items = []
                # Apply per-leaf sitemap cap if applicable
                src = fut_type.get(fut, 'unknown')
                if src == 'sitemap':
                    cap = int(getattr(args, 'sitemap_max_urls', 0) or 0)
                    if cap > 0 and len(items) > cap:
                        items = items[:cap]
                total_items += len(items)
                if src in items_by_source:
                    items_by_source[src] += len(items)
                for it in items:
                    # Trust the executor type we scheduled (fut_type) instead of inferring from fields
                    writer.submit({
                        'site': site,
                        'sourceType': src,
                        'item': it,
                        'ts': time.strftime('%Y-%m-%d %H:%M:%S')
                    })
        end_perf = time.perf_counter()
        ended_iso = datetime.now(timezone.utc).isoformat()
        collector.record_site(site=site, started_at_iso=started_iso, ended_at_iso=ended_iso, duration_sec=(end_perf - start_perf), items_by_source=items_by_source, approaches_used=approaches_used)
        # Periodic summary write
        collector.maybe_write_summary_snapshot(int(args.summary_interval_sites))
        print(f"[site] Completed: {site} -> items={total_items} (sitemap={items_by_source['sitemap']}, css={items_by_source['css']}), time={round(end_perf - start_perf, 3)}s")

        # === CSV upsert for extraction stage ===
        try:
            domain = _domain_of(site)
            source_id = site
            path_used = 'Neither'
            if items_by_source.get('sitemap', 0) > 0 and items_by_source.get('css', 0) > 0:
                path_used = 'Both'
            elif items_by_source.get('sitemap', 0) > 0:
                path_used = 'Sitemap'
            elif items_by_source.get('css', 0) > 0:
                path_used = 'CSS'

            # Build extraction error details + zero reason
            reasons: List[str] = []
            if not targets:
                reasons.append('selection_not_run_or_no_targets')
            else:
                if had_sitemap and int(items_by_source.get('sitemap', 0)) == 0:
                    reasons.append('sitemap_zero')
                if had_css and int(items_by_source.get('css', 0)) == 0:
                    reasons.append('css_zero')
            if exception_types:
                reasons.append('target_exceptions: ' + ','.join(sorted(exception_types))[:60])
            ctx_bits = [
                f"targets={len(targets)}",
                f"sitemapItems={int(items_by_source.get('sitemap', 0))}",
                f"cssItems={int(items_by_source.get('css', 0))}"
            ]
            extraction_seg = ''
            if reasons:
                extraction_seg = 'extraction: ' + ', '.join(reasons) + '; ' + '; '.join(ctx_bits)

            zero_reason = ''
            if int(total_items) == 0:
                if not targets:
                    zero_reason = 'selection_not_run_or_no_targets'
                else:
                    zparts: List[str] = []
                    if had_sitemap:
                        zparts.append('sitemap_zero')
                    if had_css:
                        zparts.append('css_zero')
                    if exception_types:
                        zparts.append('target_exceptions')
                    zero_reason = ', '.join(zparts) if zparts else 'unknown'

            # Human-friendly explanation for extraction
            human_extraction = ''
            if not targets:
                human_extraction = 'Extraction: selection produced no targets to scrape.'
            else:
                msgs: List[str] = []
                if had_sitemap and int(items_by_source.get('sitemap', 0)) == 0:
                    msgs.append('sitemap extraction returned 0 items')
                if had_css and int(items_by_source.get('css', 0)) == 0:
                    msgs.append('CSS extraction returned 0 items')
                if exception_types:
                    msgs.append(f"errors occurred ({', '.join(sorted(exception_types))})")
                if msgs:
                    human_extraction = 'Extraction: ' + '; '.join(msgs) + '.'

            # Initialize DB (idempotent)
            try:
                ov_init_db()
            except Exception:
                pass

            try:
                ov_upsert(source_id, {
                'Domain (sources)': source_id,
                'Which Path Used for Final Extraction': path_used,
                'Total Time (sec) in scraping': str(round(end_perf - start_perf, 3)),
                'Raw Articles scraped': str(int(total_items)),
                'Zero Raw Articles Reason': zero_reason,
                'Overall pipelines Error Details': extraction_seg,
                'Overall pipelines Explanation': human_extraction,
                # If selection never ran for this domain, hint a reason
                'Selector Discovery Not Attempted Reason': '' if path_used != 'Neither' else ('selection not run / no targets'),
                })
            except Exception:
                pass
        except Exception:
            pass
        return site, total_items

    # Site workers
    completed = 0
    try:
        with cf.ThreadPoolExecutor(max_workers=max(1, int(args.site_concurrency))) as site_pool:
            futures: List[cf.Future] = []
            while True:
                try:
                    row = site_q.get(timeout=0.5)
                except Empty:
                    if args.once and stop_reader.is_set() and site_q.empty():
                        break
                    continue
                futures.append(site_pool.submit(process_site, row))
                site_q.task_done()

            # Drain outstanding futures
            for fut in cf.as_completed(futures):
                try:
                    site, _ = fut.result()
                except Exception:
                    site = None
                if site:
                    processed_sites.add(site)
                completed += 1

    except KeyboardInterrupt:
        print("[stream] Stopped")
    finally:
        writer.close()
        try:
            collector.end_global()
            collector.write_summary()
            print(f"[stream] Summary written -> {args.log_summary}")
        except Exception:
            pass

        # Final CSV Export from SQLite
        try:
            ov_export_csv()
            print("[stream] Exported pipelines_overview.csv from SQLite store")
        except Exception:
            print("[stream] Warning: failed to export CSV from SQLite store")


if __name__ == '__main__':
    _cli()



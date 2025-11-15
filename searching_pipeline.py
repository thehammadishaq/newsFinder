import os
import re
import json
import time
import sys
import threading
import shutil
import concurrent.futures as cf
from typing import Dict, Any, List, Optional
from urllib.parse import urlparse
from datetime import datetime, timezone

# Force unbuffered output for real-time logs (MUST be before any other stdout modifications)
os.environ['PYTHONUNBUFFERED'] = '1'
try:
    sys.stdout.reconfigure(line_buffering=True)
    sys.stderr.reconfigure(line_buffering=True)
except Exception:
    pass  # Python < 3.7 doesn't have reconfigure

# Fix Windows console encoding for Unicode/emoji support
if sys.platform == 'win32':
    try:
        import codecs
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')
    except Exception:
        pass  # Fallback: continue without fix

from sitemap_discovery import (
    fetch_robots_txt,
    parse_sitemaps_from_robots,
    fetch_bytes,
    maybe_decompress,
    parse_xml_bytes,
    child_text_any_ns,
    _parse_w3c_datetime,
)
import selector_scraper as ss

try:
    from dotenv import load_dotenv  # type: ignore
except Exception:
    load_dotenv = None  # type: ignore


# ============================================================================
# INCREMENTAL FILE WRITING
# ============================================================================

# Global lock for thread-safe file operations
_file_lock = threading.Lock()


def _initialize_output_file(file_path: str, total_sites: int, recent_hours: int = 24, concurrency: int = 1) -> None:
    """Initialize output JSON file with empty structure"""
    data = {
        "success": True,
        "mode": "searching",
        "summary": {
            "totalWebsites": total_sites,
            "processed": 0,
            "withSitemap": 0,
            "withCssOnly": 0,
            "failed": 0,
            "inProgress": True,
            "startTime": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            "recentHours": recent_hours,
            "concurrency": concurrency,
            "totalSitemapUrls": 0,
            "totalCssSections": 0
        },
        "entries": []
    }
    
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    print(f"[file] Initialized output file: {file_path}")


def _append_entry_to_file(file_path: str, new_entry: Dict[str, Any]) -> None:
    """Thread-safe append of entry to JSON file"""
    
    with _file_lock:
        try:
            # Read current file
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError) as e:
            print(f"[ERROR] Could not read file: {e}")
            return
        
        # Add new entry
        data["entries"].append(new_entry)
        
        # Update summary
        data["summary"]["processed"] = len(data["entries"])
        data["summary"]["lastUpdate"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        
        # Update counts based on entry status
        sitemap_obj = new_entry.get("sitemap")
        css_obj = new_entry.get("css")
        if isinstance(sitemap_obj, dict) and sitemap_obj.get("present"):
            data["summary"]["withSitemap"] = data["summary"].get("withSitemap", 0) + 1
        elif isinstance(css_obj, dict) and css_obj.get("present"):
            data["summary"]["withCssOnly"] = data["summary"].get("withCssOnly", 0) + 1
        else:
            data["summary"]["failed"] = data["summary"].get("failed", 0) + 1
        
        # Update totals
        if isinstance(sitemap_obj, dict) and sitemap_obj.get("present"):
            leaf_urls = sitemap_obj.get("leafSitemapUrls", []) or []
            data["summary"]["totalSitemapUrls"] = data["summary"].get("totalSitemapUrls", 0) + len(leaf_urls)
        
        if isinstance(css_obj, dict) and css_obj.get("present"):
            css_sections = css_obj.get("sections", []) or []
            data["summary"]["totalCssSections"] = data["summary"].get("totalCssSections", 0) + len(css_sections)
        
        # Write atomically using temp file
        temp_file = file_path + ".tmp"
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        # Atomic rename (Windows safe)
        try:
            os.replace(temp_file, file_path)
        except Exception as e:
            print(f"[ERROR] Could not rename temp file: {e}")
            # Fallback: direct write
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)


def _finalize_output_file(file_path: str, start_time: float) -> None:
    """Mark output file as complete and add final statistics"""
    
    with _file_lock:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            print(f"[ERROR] Could not read file for finalization")
            return
        
        # Update summary
        data["summary"]["inProgress"] = False
        data["summary"]["completedTime"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        
        # Add performance metrics
        total_time = time.time() - start_time
        total_sites = len(data["entries"])
        
        data["summary"]["performance"] = {
            "totalTimeSeconds": round(total_time, 2),
            "averageTimePerSite": round(total_time / total_sites, 2) if total_sites > 0 else 0,
            "concurrency": data["summary"].get("concurrency", 1),
            "estimatedSpeedup": f"{round(total_sites / total_time, 1)}x" if total_time > 0 else "N/A"
        }
        
        data["summary"]["timestamp"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        
        # Write final file
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        print(f"\n[file] ✅ Finalized output file: {file_path}")


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def _ensure_dirs() -> None:
    os.makedirs("debug_html", exist_ok=True)
    os.makedirs("debug_llm", exist_ok=True)


def _domain_from_url(url: str) -> str:
    p = urlparse(url)
    return p.netloc or p.hostname or "unknown"


def _prettify_xml(xml_str: str) -> str:
    """Add indentation to XML for readability."""
    try:
        from xml.dom import minidom
        dom = minidom.parseString(xml_str)
        return dom.toprettyxml(indent="  ")
    except Exception:
        return xml_str


def _extract_sample_urls_from_sitemap(sitemap_url: str, timeout: float = 15.0, sample_count: int = 3) -> List[str]:
    """Extract first N <url> entries as pretty XML strings for LLM analysis."""
    try:
        raw = fetch_bytes(sitemap_url, timeout)
        raw = maybe_decompress(sitemap_url, raw)
        if not raw:
            return []
        
        root = parse_xml_bytes(raw)
        if root is None:
            return []
        
        # Find first N url elements
        url_elements = root.findall(".//{*}url")[:sample_count]
        
        if not url_elements:
            return []
        
        # Convert each to pretty XML string
        samples = []
        from xml.etree.ElementTree import tostring
        for url_elem in url_elements:
            xml_str = tostring(url_elem, encoding='unicode')
            pretty_xml = _prettify_xml(xml_str)
            samples.append(pretty_xml)
        
        return samples
    except Exception:
        return []


def _sitemap_llm_prompt(samples: List[str], sitemap_url: str) -> str:
    """Build LLM prompt for sitemap field detection."""
    samples_text = "\n\n---SAMPLE SEPARATOR---\n\n".join(samples)
    
    return f"""Goal: Analyze this XML sitemap structure and identify ALL useful field mappings for data extraction.

Sitemap URL: {sitemap_url}

Here are {len(samples)} sample <url> entries from the sitemap:

{samples_text}

---

Task: Return a JSON object with XPath/tag paths for extracting ALL useful fields (including nested ones).

Output format:
{{
  "type": "urlset",
  "item": "url",
  "fields": {{
    "url": "loc",
    "date": "news:news/news:publication_date",
    "title": "news:news/news:title",
    "language": "news:news/news:publication/news:language",
    "publisher": "news:news/news:publication/news:name",
    "imageUrl": "image:image/image:loc",
    "changefreq": "changefreq",
    "priority": "priority"
  }},
  "confidence": 0.9
}}

STRICT RULES:
1. Return FIELD PATHS/TAGS as strings, NOT example values or actual content
2. For nested tags, use XPath-style paths: "parent:tag/child:tag" (e.g., "news:news/news:title")
3. For direct children of <url>, use just the tag name: "loc", "lastmod"
4. Include namespace prefixes if present: "news:title", "image:loc"
5. Extract ALL useful fields you find:
   - Core: url, date, title
   - Metadata: author, publisher, language, description, category
   - Media: imageUrl, videoUrl, thumbnailUrl
   - SEO: changefreq, priority, keywords
6. If multiple date fields exist, include both as "date" and "dateAlternative"
7. DO NOT include: xmlns attributes, technical metadata, empty tags
8. Confidence: 0.9 if structure is very clear, 0.7-0.8 if some ambiguity

Return ONLY valid JSON, no markdown fences or explanation.
"""


def _parse_sitemap_selector_response(llm_response: str) -> Optional[Dict[str, Any]]:
    """Parse LLM response for sitemap field mappings."""
    # Reuse existing JSON extraction from selector_scraper
    import selector_scraper as ss
    data = ss._extract_json_from_text(llm_response)
    
    if not isinstance(data, dict):
        return None
    
    # Validate structure
    fields = data.get("fields")
    if not isinstance(fields, dict) or not fields:
        return None
    
    # Must have at least 'url' field
    if "url" not in fields:
        return None
    
    # Clean and validate field paths
    valid_fields = {}
    for key, xpath in fields.items():
        if isinstance(xpath, str) and xpath.strip():
            # Basic validation - should look like tag path (allow namespaces, slashes, colons)
            if re.match(r'^[a-zA-Z0-9_:/.\\-]+$', xpath):
                valid_fields[key] = xpath.strip()
    
    if not valid_fields:
        return None
    
    return {
        "type": data.get("type", "urlset"),
        "item": data.get("item", "url"),
        "fields": valid_fields,
        "confidence": data.get("confidence", 0.7),
        "detectionMethod": "llm"
    }


def _detect_selectors_from_xml_with_llm(
    sitemap_url: str, 
    timeout: float = 15.0,
    model: Optional[str] = None
) -> Optional[Dict[str, Any]]:
    """Use LLM to intelligently detect sitemap field mappings (including nested fields)."""
    import selector_scraper as ss
    from urllib.parse import urlparse as up
    
    # Load .env for API keys (search current and parent directories)
    if load_dotenv:
        try:
            # dotenv automatically searches up the directory tree
            load_dotenv(override=False)  # Don't override existing env vars
        except Exception:
            pass
    
    try:
        # Step 1: Extract samples
        print(f"[detect-llm] Sampling sitemap: {sitemap_url}")
        samples = _extract_sample_urls_from_sitemap(sitemap_url, timeout=timeout, sample_count=3)
        
        if not samples:
            print(f"[detect-llm] No samples found, skipping LLM detection")
            return None
        
        # Save INPUT samples to file (for review/debugging)
        parsed_url = up(sitemap_url)
        domain = parsed_url.netloc or "unknown"
        from datetime import datetime
        ts = datetime.utcnow().strftime("%Y-%m-%dT%H-%M-%S-%fZ")
        
        # Extract sitemap name from URL and sanitize for Windows
        sitemap_name = sitemap_url.split('/')[-1].split('?')[0]  # Remove query params
        sitemap_name = sitemap_name.replace('.xml', '').replace('.gz', '')
        # Remove invalid Windows filename characters: < > : " / \ | ? *
        for char in '<>:"/\\|?*':
            sitemap_name = sitemap_name.replace(char, '_')
        sitemap_name = sitemap_name[:50]  # Limit length
        
        input_debug_path = os.path.join("debug_llm", f"INPUT_{domain}_{sitemap_name}_{ts}.json")
        with open(input_debug_path, "w", encoding="utf-8") as f:
            json.dump({
                "sitemap_url": sitemap_url,
                "timestamp": ts,
                "samples_count": len(samples),
                "samples": samples,
                "total_chars": sum(len(s) for s in samples)
            }, f, indent=2, ensure_ascii=False)
        
        print(f"[detect-llm] 💾 Samples saved: {input_debug_path}")
        
        # Step 2: Build LLM prompt
        prompt = _sitemap_llm_prompt(samples, sitemap_url)
        
        # Step 3: Call LLM
        print(f"[detect-llm] 🤖 Calling LLM for intelligent field detection...")
        llm_response = ss._call_llm(prompt, model=model)
        
        # Save LLM response for debugging
        output_debug_path = os.path.join("debug_llm", f"OUTPUT_{domain}_{sitemap_name}_{ts}.json")
        with open(output_debug_path, "w", encoding="utf-8") as f:
            json.dump({
                "sitemap_url": sitemap_url,
                "timestamp": ts,
                "llm_response": llm_response,
                "response_length": len(llm_response)
            }, f, indent=2, ensure_ascii=False)
        
        print(f"[detect-llm] 💾 LLM response saved: {output_debug_path}")
        
        # Step 4: Parse LLM response
        detected = _parse_sitemap_selector_response(llm_response)
        
        if detected:
            field_count = len(detected.get('fields', {}))
            confidence = detected.get('confidence', 0)
            print(f"[detect-llm] ✅ Detected {field_count} fields (confidence={confidence})")
            return detected
        else:
            print(f"[detect-llm] ❌ Failed to parse LLM response")
            return None
            
    except Exception as e:
        print(f"[detect-llm] ❌ Error: {e}")
        return None


def _detect_selectors_from_xml(sitemap_url: str, timeout: float = 15.0) -> Optional[Dict[str, Any]]:
    """Basic detection - inspect XML tags without LLM (fallback)."""
    try:
        raw = fetch_bytes(sitemap_url, timeout)
        raw = maybe_decompress(sitemap_url, raw)
        if not raw:
            return None
        root = parse_xml_bytes(raw)
        if root is None:
            return None
        
        tag = (root.tag or "").lower()
        
        # Check if it's sitemapindex or urlset
        if tag.endswith("sitemapindex"):
            # For index, standard tags are always the same
            return {
                "type": "index",
                "item": "sitemap",
                "fields": {
                    "url": "loc",
                    "date": "lastmod"
                },
                "detectionMethod": "basic"
            }
        elif tag.endswith("urlset"):
            # For urlset, basic detection of direct children only
            first_url = root.find(".//{*}url")
            if first_url is None:
                return None
            
            # Get all child tags (strip namespace)
            tags_present = set()
            for child in first_url:
                tag_name = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                tags_present.add(tag_name.lower())
            
            # Build basic selector mapping
            fields: Dict[str, str] = {}
            
            # URL
            if "loc" in tags_present:
                fields["url"] = "loc"
            
            # Date
            if "lastmod" in tags_present:
                fields["date"] = "lastmod"
            
            # Optional
            if "changefreq" in tags_present:
                fields["changefreq"] = "changefreq"
            if "priority" in tags_present:
                fields["priority"] = "priority"
            
            return {
                "type": "urlset",
                "item": "url",
                "fields": fields,
                "detectionMethod": "basic"
            }
        else:
            return None
    except Exception:
        return None


def _is_leaf_sitemap_recent(root, recent_hours: int, sample_size: int = 1) -> bool:
    """
    Check if a leaf sitemap contains recent articles by sampling first N entries.
    
    Args:
        root: Parsed XML root element (urlset)
        recent_hours: Age threshold in hours
        sample_size: Number of articles to sample (default: 10)
        
    Returns:
        True if sitemap has at least one recent article, False otherwise
    """
    from datetime import datetime, timezone
    
    # DEBUG: Check how many URL elements found
    all_urls = root.findall(".//{*}url")
    print(f"[DEBUG-LEAF] Found {len(all_urls)} <url> elements, sampling first {sample_size}")
    
    # Sample first N article entries
    article_dates = []
    parsed_count = 0
    failed_count = 0
    
    for idx, url_elem in enumerate(all_urls[:sample_size], 1):
        lastmod = child_text_any_ns(url_elem, "lastmod")
        
        if lastmod:
            dt = _parse_w3c_datetime(lastmod)
            if dt:
                article_dates.append(dt)
                parsed_count += 1
            else:
                failed_count += 1
        else:
            failed_count += 1
    
    print(f"[DEBUG-LEAF] Sampled {sample_size} articles: {parsed_count} with valid dates, {failed_count} without")
    
    # If no dates found, be conservative (keep it)
    if not article_dates:
        print(f"[DEBUG-LEAF] ⚠️ CONSERVATIVE MODE - no dates found, returning True (keep sitemap)")
        return True
    
    # Check if most recent article is within threshold
    most_recent = max(article_dates)
    now = datetime.now(timezone.utc)
    age_hours = (now - most_recent).total_seconds() / 3600
    
    print(f"[DEBUG-LEAF] Most recent article: {most_recent}")
    print(f"[DEBUG-LEAF] Age: {age_hours:.1f} hours (threshold: {recent_hours}h)")
    print(f"[DEBUG-LEAF] Result: {'✅ PASS (keep)' if age_hours <= recent_hours else '❌ REJECT (too old)'}")
    
    return age_hours <= recent_hours


def build_sitemap_selectors(url: str, recent_hours: int = 24, timeout: float = 15.0, use_llm_filter: bool = False) -> Dict[str, Any]:
    """Return sitemap selector metadata without fetching article data.

    - Reads robots.txt; collects sitemap URLs
    - Applies word-based filtering (video/sports/weather rejected)
    - Applies date-based filtering (old sitemaps rejected)
    - Recursively expands to leaf sitemaps (checks article freshness)
    - Detects XML field mappings for each leaf
    """
    _ensure_dirs()
    robots = fetch_robots_txt(url, timeout=timeout)
    if not robots:
        return {"present": False}
    # Get all sitemaps without filtering (filtering done by recursive expansion)
    sitemaps = parse_sitemaps_from_robots(robots, url, news_only=False)
    if not sitemaps:
        return {"present": False}

    print(f"[sitemap] Found {len(sitemaps)} sitemap URL(s) in robots.txt")
    
    # 🚀 FAST PRE-FILTERING: Apply filters on robots.txt URLs before downloading
    from sitemap_filters import filter_by_words, filter_by_date, filter_sitemaps_by_year
    
    # Step 1: Word filter (reject video, image, sports, etc.)
    sitemaps_before_word = len(sitemaps)
    sitemaps_word_filtered = []
    rejected_by_word = []
    for sm in sitemaps:
        should_keep, matched_word = filter_by_words(sm)
        if should_keep:
            sitemaps_word_filtered.append(sm)
        else:
            rejected_by_word.append((sm, matched_word))
    
    sitemaps = sitemaps_word_filtered
    
    if rejected_by_word:
        print(f"[word-filter] Rejected {len(rejected_by_word)} sitemap(s) by keywords:")
        for sm_url, word in rejected_by_word[:5]:
            print(f"[word-filter]   ❌ {sm_url} ('{word}')")
        if len(rejected_by_word) > 5:
            print(f"[word-filter]   ... and {len(rejected_by_word) - 5} more")
    
    if len(sitemaps) < sitemaps_before_word:
        print(f"[word-filter] URLs: {sitemaps_before_word} → {len(sitemaps)} (rejected {sitemaps_before_word - len(sitemaps)} by keywords)")
    
    # 🆕 Step 2: Year filter (reject any URL with old years in filename)
    sitemaps = filter_sitemaps_by_year(sitemaps)
    
    # Step 3: Date filter (additional date checks, now mostly redundant after year filter)
    sitemaps_before_date = len(sitemaps)
    sitemaps_filtered = []
    rejected_by_date = []
    for sm in sitemaps:
        keep, reason = filter_by_date(sm, hours_threshold=recent_hours, conservative=True)
        if keep:
            sitemaps_filtered.append(sm)
        else:
            rejected_by_date.append((sm, reason))
    
    sitemaps = sitemaps_filtered
    
    if rejected_by_date:
        print(f"[url-filter] 🚀 Fast rejection of {len(rejected_by_date)} old sitemap(s) by URL pattern:")
        for sm_url, reason in rejected_by_date[:5]:  # Show first 5
            print(f"[url-filter]   ❌ {sm_url} ({reason})")
        if len(rejected_by_date) > 5:
            print(f"[url-filter]   ... and {len(rejected_by_date) - 5} more")
    
    if len(sitemaps) < sitemaps_before_date:
        print(f"[date-filter] URLs: {sitemaps_before_date} → {len(sitemaps)} (rejected {sitemaps_before_date - len(sitemaps)} by URL date)")
    
    print(f"[sitemap] After pre-filtering: {len(sitemaps)} sitemap(s) remaining\n")
    
    if not sitemaps:
        print(f"[sitemap] ⚠️ All sitemaps filtered out (too old or wrong keywords)")
        return {"present": False}
    
    def _expand_children_recursive(sitemap_url: str) -> List[str]:
        """
        Recursively expand sitemapindex to leaf urlsets.
        Applies word + year + date filtering at every level using modular functions.
        """
        from sitemap_filters import filter_by_words, filter_by_date
        from datetime import datetime
        
        leaves: List[str] = []
        raw = fetch_bytes(sitemap_url, timeout)
        raw = maybe_decompress(sitemap_url, raw)
        if not raw:
            return leaves
        root = parse_xml_bytes(raw)
        if root is None:
            return leaves
        tag = (root.tag or "").lower()
        
        if tag.endswith("sitemapindex"):
            # It's an INDEX - process children
            for smnode in root.findall(".//{*}sitemap"):
                loc = child_text_any_ns(smnode, "loc")
                if not loc:
                    continue
                
                child_url = loc.strip()
                
                # Extract lastmod from XML (same iteration - zero extra cost!)
                lastmod_str = child_text_any_ns(smnode, "lastmod")
                child_lastmod = _parse_w3c_datetime(lastmod_str) if lastmod_str else None
                
                # ===== FILTER 1: Word Filter on Child =====
                should_keep_word, matched_word = filter_by_words(child_url)
                if not should_keep_word:
                    print(f"[word-filter] ❌ Child rejected: {child_url} ('{matched_word}')")
                    continue  # Skip this child
                
                # ===== FILTER 2: Year Filter on Child (reject old years) =====
                current_year = datetime.now().year
                year_pattern = r'(19[5-9]\d|20[0-3]\d)'
                years_found = [int(y) for y in re.findall(year_pattern, child_url)]
                
                if years_found and not all(y == current_year for y in years_found):
                    old_years = [str(y) for y in years_found if y != current_year]
                    print(f"[year-filter] ❌ Child rejected: {child_url} (old year(s): {', '.join(old_years)})")
                    continue  # Skip this child
                
                # ===== FILTER 3: Date Filter on Child (with XML lastmod) =====
                should_keep_date, date_reason = filter_by_date(
                    child_url,
                    hours_threshold=recent_hours,
                    conservative=True,  # Keep if no date found
                    xml_lastmod=child_lastmod  # ✅ Pass XML lastmod!
                )
                if not should_keep_date:
                    print(f"[date-filter] ❌ Child rejected: {child_url} ({date_reason})")
                    continue  # Skip this child
                
                # ✅ Both filters passed - RECURSE into child
                try:
                    child_leaves = _expand_children_recursive(child_url)
                    if child_leaves:
                        for cl in child_leaves:
                            if cl not in leaves:
                                leaves.append(cl)
                    # If empty, child was either rejected or has no content - don't add
                except Exception as e:
                    print(f"[expand] ⚠️  Recursion failed for {child_url}: {type(e).__name__}")
                    # On error, don't add anything (safer than guessing)
        else:
            # It's a LEAF urlset - check article freshness before adding
            print(f"[DEBUG-LEAF] Checking leaf sitemap: {sitemap_url}")
            if _is_leaf_sitemap_recent(root, recent_hours):
                # Has recent articles - keep it
                print(f"[leaf-filter] ✅ Leaf kept (has recent articles): {sitemap_url}")
            leaves.append(sitemap_url)
            else:
                # All articles are old - reject it
                print(f"[leaf-filter] ❌ Leaf rejected - all sampled articles old: {sitemap_url}")
        
        return leaves

    selected: List[Dict[str, Any]] = []
    aggregated_selector_urls: List[str] = []
    for sm in sitemaps:
        # Skip peek - let recursive expansion with filters handle everything
        is_recent = True  # Filters will determine actual recency
        
        # Recursively collect leaf sitemap URLs (with word+date filtering)
        leaf_selectors = _expand_children_recursive(sm)
        # Build selectors list for this node (leaf sitemaps to use)
        sel_urls: List[str] = leaf_selectors or ([sm] if is_recent else [])
        
        # Final filtering on collected leaves (word + year + date filters)
        # ALWAYS apply filters, regardless of use_llm_filter flag
        try:
            from sitemap_filters import filter_by_words, filter_by_date, filter_sitemaps_by_year
            
            print(f"[filter] Starting final filtering on {len(sel_urls)} collected leaf sitemap(s)...")
            
            # ===== FILTER 1: Word Filter on Leaves =====
            word_filtered = []
            for leaf_url in sel_urls:
                should_keep, matched_word = filter_by_words(leaf_url)
                if should_keep:
                    word_filtered.append(leaf_url)
                            else:
                    print(f"[word-filter] ❌ Leaf: {leaf_url} ('{matched_word}')")
            
            print(f"[word-filter] Leaves: {len(sel_urls)} → {len(word_filtered)} (rejected {len(sel_urls) - len(word_filtered)})")
            
            # ===== FILTER 2: Year Filter on Leaves =====
            year_filtered = filter_sitemaps_by_year(word_filtered)
            
            # ===== FILTER 3: Date Filter on Leaves =====
            date_filtered = []
            for leaf_url in year_filtered:
                should_keep, date_reason = filter_by_date(
                    leaf_url,
                    hours_threshold=recent_hours,
                    conservative=True  # Keep if no date in URL
                )
                if should_keep:
                    date_filtered.append(leaf_url)
                            else:
                    print(f"[date-filter] ❌ Leaf: {leaf_url} ({date_reason})")
            
            print(f"[date-filter] Leaves: {len(year_filtered)} → {len(date_filtered)} (rejected {len(year_filtered) - len(date_filtered)})")
            
            sel_urls = date_filtered
            
        except ImportError:
            # Fallback if sitemap_filters.py not available
            print("[filter] ⚠️  sitemap_filters.py not found, keeping all URLs")
            pass
        
        # Detect selectors for each FILTERED leaf sitemap using LLM (with error handling)
        leaf_with_selectors = []
        for idx, leaf_url in enumerate(sel_urls, 1):
            try:
                print(f"[detect] 🔍 [{idx}/{len(sel_urls)}] Detecting selectors: {leaf_url}")
                
                # Try LLM detection first (comprehensive), fallback to basic if fails
                detected = _detect_selectors_from_xml_with_llm(leaf_url, timeout=timeout)
                if not detected:
                    # Fallback to basic detection
                    print(f"[detect] ⚠️  LLM detection failed, trying basic...")
                    detected = _detect_selectors_from_xml(leaf_url, timeout=timeout)
                
                leaf_obj = {
                    "url": leaf_url,
                    "likelyRecent": True,
                    "detectedSelectors": detected
                }
                leaf_with_selectors.append(leaf_obj)
                
                # Aggregate for top-level
                if leaf_url and leaf_url not in aggregated_selector_urls:
                    aggregated_selector_urls.append(leaf_url)
                    
            except Exception as e:
                print(f"[detect] ❌ Detection failed for {leaf_url}: {type(e).__name__}: {str(e)[:80]}")
                # Add with empty selectors on error
                leaf_with_selectors.append({
                    "url": leaf_url,
                    "likelyRecent": True,
                    "detectedSelectors": None
                })
        
        # Store parent sitemap with its leaf sitemaps (no duplication)
        selected.append({
            "url": sm, 
            "likelyRecent": is_recent,
            "leafSitemaps": leaf_with_selectors  # Detailed info with detected selectors
        })
    
    return {
        "present": True,
        "recentHours": recent_hours,
        "sitemaps": selected,
        "leafSitemapUrls": aggregated_selector_urls,  # For quick iteration during scraping
    }


def build_css_selectors(url: str, headful: bool = False, slowmo_ms: int = 0) -> Dict[str, Any]:
    _ensure_dirs()
    try:
        discovered = ss.discover_selectors(url=url, headful=headful, slowmo_ms=slowmo_ms)
        sections = discovered.get("sections") or []
        if isinstance(sections, list) and sections:
            return {"present": True, "sections": sections}
    except Exception:
        pass
    return {"present": False, "sections": []}


def _cli():
    import argparse

    parser = argparse.ArgumentParser(description="Searching pipeline: build sitemap+CSS selectors from URL or Excel")
    parser.add_argument("url", nargs="?", help="Site URL (optional if --excel provided)")
    parser.add_argument("--excel", help="Path to .xlsx with URLs (first column or header 'url')")
    parser.add_argument("--output", default="selectors_search.json", help="Output JSON path")
    parser.add_argument("--recent-hours", dest="recent_hours", type=int, default=24)
    parser.add_argument("--sitemaps-only", action="store_true", help="Skip CSS discovery; only output sitemap selectors")
    parser.add_argument("--always-css", action="store_true", help="Always discover CSS selectors even if sitemap exists (for fallback)")
    # LLM filter removed - now using word-based filter only (FREE & fast)
    parser.add_argument("--export-like", action="store_true", help="Use sitemap_export-style strategy (no recency heuristics); recurse to ALL leaf urlsets and emit as leafSitemapUrls")
    parser.add_argument("--headful", action="store_true")
    parser.add_argument("--slowmo", type=int, default=0)
    parser.add_argument("--concurrency", type=int, default=1)
    args = parser.parse_args()

    out_path = args.output
    script_dir = os.path.dirname(__file__)
    project_root = os.path.abspath(os.path.join(script_dir, ".."))
    base_dir_name = os.path.basename(script_dir)
    if not os.path.isabs(out_path):
        norm = out_path.replace("/", os.sep).replace("\\", os.sep)
        if not os.path.dirname(norm):
            out_path = os.path.join(script_dir, norm)
        else:
            if norm.lower().startswith(base_dir_name.lower() + os.sep):
                out_path = os.path.join(script_dir, os.path.basename(norm))
            else:
                out_path = os.path.join(project_root, norm)
    os.makedirs(os.path.dirname(out_path) or script_dir, exist_ok=True)

    # Excel mode
    if args.excel:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            raise SystemExit("openpyxl is required. Please pip install openpyxl or run pip install -r python_tools/requirements.txt")
        xls_path = args.excel
        if not os.path.isabs(xls_path) and not os.path.exists(xls_path):
            cand = os.path.join(script_dir, os.path.basename(xls_path))
            if os.path.exists(cand):
                xls_path = cand
        if not os.path.exists(xls_path):
            raise SystemExit(f"Excel not found: {args.excel}")

        wb = load_workbook(xls_path)
        ws = wb.active
        urls: List[str] = []
        header: List[str] = []
        url_idx = 0
        row_count = 0
        
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row_count += 1
            if i == 0 and row:
                header = [str(c).strip().lower() if c is not None else "" for c in row]
                try:
                    url_idx = header.index("url")
                except ValueError:
                    url_idx = 0
                print(f"[DEBUG-EXCEL] Header row: {header}, URL column index: {url_idx}")
                continue
            if not row:
                continue
            cell = row[url_idx] if url_idx < len(row) else None
            if not cell:
                continue
            u = str(cell).strip()
            if not u:
                continue
            if not (u.startswith("http://") or u.startswith("https://")):
                u = "https://" + u
            urls.append(u)
        
        # DEBUG: Show URLs before deduplication
        print(f"[DEBUG-EXCEL] Rows processed: {row_count}, URLs extracted: {len(urls)}")
        print(f"[DEBUG-EXCEL] URLs before deduplication: {len(urls)}")
        
        seen = set()
        urls = [u for u in urls if not (u in seen or seen.add(u))]
        
        # DEBUG: Show collected URLs
        print(f"\n[DEBUG-EXCEL] Total URLs collected from Excel: {len(urls)}")
        for idx, url in enumerate(urls[:5], 1):
            print(f"[DEBUG-EXCEL] URL {idx}: {url}")
        if len(urls) > 5:
            print(f"[DEBUG-EXCEL] ... and {len(urls) - 5} more URLs")
        print()
        
        if not urls:
            raise SystemExit("No URLs in Excel")

        entries: List[Dict[str, Any]] = []
        max_workers = max(1, int(args.concurrency))
        effective_headful = args.headful and (max_workers == 1)

        def _leaf_urlsets_export_like(root_url: str) -> List[str]:
            """Recurse sitemap indexes and collect leaf urlsets using modular filters."""
            from sitemap_filters import filter_by_words, filter_by_date
            
            urls_out: List[str] = []
            seen = set()
            robots = fetch_robots_txt(root_url)
            if not robots:
                return []
            sitemaps = parse_sitemaps_from_robots(robots, root_url)
            
            def visit(url: str) -> None:
                raw = fetch_bytes(url, 15.0)
                raw = maybe_decompress(url, raw)
                if not raw:
                    return
                root = parse_xml_bytes(raw)
                if root is None:
                    return
                tag = (root.tag or "").lower()
                
                if tag.endswith("sitemapindex"):
                    # Process children with filtering
                    for smnode in root.findall(".//{*}sitemap"):
                        loc = child_text_any_ns(smnode, "loc")
                        if not loc:
                            continue
                        
                        child_url = loc.strip()
                        
                        # Word filter
                        should_keep_word, matched_word = filter_by_words(child_url)
                        if not should_keep_word:
                            continue
                        
                        # Date filter (2 months threshold for export mode)
                        should_keep_date, date_reason = filter_by_date(child_url, hours_threshold=1440, conservative=True)  # 60 days
                        if not should_keep_date:
                                continue
                        
                        # Passed filters → visit
                        visit(child_url)
                else:
                    # Leaf urlset
                    if url not in seen:
                        seen.add(url)
                        urls_out.append(url)
            
            # Process each sitemap
            for sm in sitemaps:
                # Word filter on parent
                should_keep, word = filter_by_words(sm)
                if not should_keep:
                    continue
                visit(sm)
            
            return urls_out

        def _proc(u: str) -> Dict[str, Any]:
            if args.export_like:
                # Export-like: emit leaf urlsets as leafSitemapUrls with detected selectors
                leafs = _leaf_urlsets_export_like(u)
                # Detect selectors for each leaf using LLM
                leaf_with_selectors = []
                for leaf_url in leafs:
                    # Try LLM first, fallback to basic
                    detected = _detect_selectors_from_xml_with_llm(leaf_url, timeout=15.0)
                    if not detected:
                        detected = _detect_selectors_from_xml(leaf_url, timeout=15.0)
                    leaf_with_selectors.append({
                        "url": leaf_url,
                        "likelyRecent": True,
                        "detectedSelectors": detected
                    })
                sm = {
                    "present": bool(leafs), 
                    "recentHours": args.recent_hours, 
                    "sitemaps": [{
                        "url": u,
                        "likelyRecent": True,
                        "leafSitemaps": leaf_with_selectors
                    }] if leafs else [],
                    "leafSitemapUrls": leafs
                }
                # Fallback to CSS discovery ONLY if no sitemap could be resolved
                css = build_css_selectors(u, headful=effective_headful, slowmo_ms=args.slowmo) if not leafs else {"present": False, "sections": []}
            else:
                sm = build_sitemap_selectors(u, recent_hours=args.recent_hours, use_llm_filter=True)  # Always use word filter
                # ✨ Discover CSS as fallback if:
                # 1. --always-css flag is set, OR
                # 2. Sitemap not present, OR
                # 3. Sitemap present but no usable leaf sitemaps (empty/invalid)
                has_usable_sitemap = sm.get("present") and len(sm.get("sitemaps", [])) > 0 and any(
                    len(s.get("leafSitemaps", [])) > 0 for s in sm.get("sitemaps", [])
                )
                should_discover_css = args.always_css or not has_usable_sitemap
                css = build_css_selectors(u, headful=effective_headful, slowmo_ms=args.slowmo) if should_discover_css else {"present": False, "sections": []}
            # Create human-friendly output
            result = {
                "url": u,
                "domain": _domain_from_url(u),
                "status": "",
                "summary": {},
                "sitemap": sm if sm.get("present") else None,
                "css": css if css.get("present") else None,
            }
            
            # Add human-friendly status and summary
            if sm.get("present"):
                sitemap_count = len(sm.get("leafSitemapUrls", []))
                result["status"] = "✅ Sitemap Found"
                result["summary"] = {
                    "method": "Sitemap",
                    "leafSitemapCount": sitemap_count,
                    "totalSitemaps": len(sm.get("sitemaps", [])),
                    "note": f"{sitemap_count} sitemap(s) ready for scraping"
                }
            elif css.get("present"):
                section_count = len(css.get("sections", []))
                result["status"] = "⚠️ No Sitemap - CSS Fallback"
                result["summary"] = {
                    "method": "CSS Selectors",
                    "sections": section_count,
                    "note": f"{section_count} section(s) discovered via HTML analysis"
                }
            else:
                result["status"] = "❌ No Selectors Found"
                result["summary"] = {
                    "method": "None",
                    "note": "Neither sitemap nor CSS selectors could be found"
                }
            
            return result

        print(f"\n{'='*70}")
        print(f"🚀 STARTING SELECTOR DISCOVERY")
        print(f"{'='*70}")
        print(f"📊 Total Sites: {len(urls)}")
        print(f"⚡ Concurrency: {max_workers} worker(s)")
        print(f"🕐 Recent Hours: {args.recent_hours}h")
        print(f"🔤 Word Filter: ✅ Enabled (video/sports/weather/etc)")
        print(f"{'='*70}\n")
        
        import time
        start_time = time.time()
        
        # 🆕 Initialize output file with empty structure
        _initialize_output_file(args.output, len(urls), args.recent_hours, max_workers)
        
        with cf.ThreadPoolExecutor(max_workers=max_workers) as executor:
            completed = 0
            for entry in executor.map(_proc, urls):
                entries.append(entry)
                completed += 1
                
                # 🆕 Append entry to file immediately
                _append_entry_to_file(args.output, entry)
                
                # Progress indicator with details
                percentage = (completed / len(urls)) * 100
                status_emoji = "✅" if entry.get("sitemap") else ("⚠️" if entry.get("css") else "❌")
                print(f"\n{'─'*70}")
                print(f"{status_emoji} [{completed}/{len(urls)}] ({percentage:.0f}%) | {entry.get('domain', 'unknown')}")
                print(f"   Status: {entry.get('status', 'Unknown')}")
                print(f"   💾 Saved to: {args.output}")
                if entry.get("sitemap"):
                    leaf_count = len(entry.get("sitemap", {}).get("leafSitemapUrls", []))
                    print(f"   📰 Sitemaps: {leaf_count} leaf sitemap(s) discovered")
                if entry.get("css"):
                    section_count = len(entry.get("css", {}).get("sections", []))
                    print(f"   🎨 CSS: {section_count} section(s) discovered")
                print(f"{'─'*70}")

        # Calculate elapsed time
        end_time = time.time()
        elapsed = end_time - start_time
        
        # 🆕 Finalize output file (mark as complete, add final stats)
        _finalize_output_file(args.output, start_time)
        
        print(f"\n{'='*70}")
        print(f"⏱️  PROCESSING TIME")
        print(f"{'='*70}")
        print(f"Total time: {elapsed:.1f}s")
        print(f"Average per site: {elapsed/len(urls):.2f}s")
        print(f"Speedup: ~{len(urls)/max_workers:.1f}x faster (parallel processing)")
        print(f"{'='*70}\n")

        # Create summary statistics for display
        sitemap_count = sum(1 for e in entries if e.get("sitemap"))
        css_count = sum(1 for e in entries if e.get("css"))
        failed_count = sum(1 for e in entries if not e.get("sitemap") and not e.get("css"))
        total_sitemap_urls = sum(len(e.get("sitemap", {}).get("leafSitemapUrls", [])) for e in entries if e.get("sitemap"))
        total_css_sections = sum(len(e.get("css", {}).get("sections", [])) for e in entries if e.get("css"))
        
        # Beautiful summary output
        print(f"\n\n{'='*70}")
        print(f"✨ DISCOVERY COMPLETE!")
        print(f"{'='*70}")
        print(f"📁 Output: {args.output}")
        print(f"📊 Results:")
        print(f"   ✅ With Sitemap:     {sitemap_count}/{len(entries)} ({(sitemap_count/len(entries)*100):.0f}%)")
        print(f"   ⚠️  CSS Fallback:     {css_count}/{len(entries)} ({(css_count/len(entries)*100):.0f}%)")
        print(f"   ❌ Failed:           {failed_count}/{len(entries)} ({(failed_count/len(entries)*100):.0f}%)")
        print(f"📰 Total Leaf Sitemaps: {total_sitemap_urls}")
        print(f"🎨 Total CSS Sections:  {total_css_sections}")
        print(f"{'='*70}\n")
        
        print(json.dumps({
            "success": True, 
            "output": args.output, 
            "sites": len(entries),
            "sitemap": sitemap_count,
            "css": css_count,
            "failed": failed_count
        }, ensure_ascii=False))
        return

    # Single URL
    if not args.url:
        raise SystemExit("Provide a URL or --excel")
    if args.export_like:
        leafs = _leaf_urlsets_export_like(args.url)
        # Detect selectors for each leaf using LLM
        leaf_with_selectors = []
        for leaf_url in leafs:
            # Try LLM first, fallback to basic
            detected = _detect_selectors_from_xml_with_llm(leaf_url, timeout=15.0)
            if not detected:
                detected = _detect_selectors_from_xml(leaf_url, timeout=15.0)
            leaf_with_selectors.append({
                "url": leaf_url,
                "likelyRecent": True,
                "detectedSelectors": detected
            })
        sm = {
            "present": bool(leafs), 
            "recentHours": args.recent_hours, 
            "sitemaps": [{
                "url": args.url,
                "likelyRecent": True,
                "leafSitemaps": leaf_with_selectors
            }] if leafs else [],
            "leafSitemapUrls": leafs
        }
        # Fallback CSS only when no sitemap
        css = build_css_selectors(args.url, headful=args.headful, slowmo_ms=args.slowmo) if not leafs else {"present": False, "sections": []}
    else:
        sm = build_sitemap_selectors(args.url, recent_hours=args.recent_hours, use_llm_filter=True)  # Always use word filter
        # ✨ Discover CSS as fallback if:
        # 1. --always-css flag is set, OR
        # 2. Sitemap not present, OR
        # 3. Sitemap present but no usable leaf sitemaps (empty/invalid)
        has_usable_sitemap = sm.get("present") and len(sm.get("sitemaps", [])) > 0 and any(
            len(s.get("leafSitemaps", [])) > 0 for s in sm.get("sitemaps", [])
        )
        should_discover_css = args.always_css or not has_usable_sitemap
        css = build_css_selectors(args.url, headful=args.headful, slowmo_ms=args.slowmo) if should_discover_css else {"present": False, "sections": []}
    # Create human-friendly output
    entry = {
        "url": args.url,
        "domain": _domain_from_url(args.url),
        "status": "",
        "summary": {},
        "sitemap": sm if sm.get("present") else None,
        "css": css if css.get("present") else None,
    }
    
    # Add human-friendly status and summary
    if sm.get("present"):
        sitemap_count = len(sm.get("leafSitemapUrls", []))
        entry["status"] = "✅ Sitemap Found"
        entry["summary"] = {
            "method": "Sitemap",
            "leafSitemapCount": sitemap_count,
            "totalSitemaps": len(sm.get("sitemaps", [])),
            "note": f"{sitemap_count} sitemap(s) ready for scraping"
        }
    elif css.get("present"):
        section_count = len(css.get("sections", []))
        entry["status"] = "⚠️ No Sitemap - CSS Fallback"
        entry["summary"] = {
            "method": "CSS Selectors",
            "sections": section_count,
            "note": f"{section_count} section(s) discovered via HTML analysis"
        }
    else:
        entry["status"] = "❌ No Selectors Found"
        entry["summary"] = {
            "method": "None",
            "note": "Neither sitemap nor CSS selectors could be found"
        }
    # Create summary for single URL
    sitemap_count = 1 if entry.get("sitemap") else 0
    css_count = 1 if entry.get("css") else 0
    failed_count = 1 if not entry.get("sitemap") and not entry.get("css") else 0
    total_sitemap_urls = len(entry.get("sitemap", {}).get("leafSitemapUrls", [])) if entry.get("sitemap") else 0
    total_css_sections = len(entry.get("css", {}).get("sections", [])) if entry.get("css") else 0
    
    out = {
        "success": True,
        "mode": "searching",
        "summary": {
            "totalWebsites": 1,
            "withSitemap": sitemap_count,
            "withCssOnly": css_count,
            "failed": failed_count,
            "totalSitemapUrls": total_sitemap_urls,
            "totalCssSections": total_css_sections,
            "recentHours": args.recent_hours,
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S UTC", time.gmtime())
        },
        "entries": [entry]
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(json.dumps({
        "success": True, 
        "output": out_path, 
        "sites": 1,
        "sitemap": sitemap_count,
        "css": css_count,
        "failed": failed_count
    }, ensure_ascii=False))


if __name__ == "__main__":
    _cli()


